"""
DOCUMENT INGESTION ENGINE
Created: February 2, 2026
Last Updated: February 26, 2026 - ADDED full PPTX and Excel extractors

CHANGELOG:
- February 26, 2026 (Session 2): ADDED full PPTX and Excel extractors
  * PROBLEM: PowerPoint and Excel extractors were stub placeholders (TODO).
    PPTX route used python-pptx which reads slide TEXT only — completely missing
    chart data embedded as chart XML in the zip structure. Survey results stored
    as bar/pie/column charts were invisible to the system.
    Excel route converted cells to flat text, losing all numeric structure.
  * SOLUTION: Both file types now pass file_bytes directly to new structured
    extractors that operate on the raw file bytes rather than pre-converted text.
  *
  * PPTX EXTRACTORS (_extract_from_survey_pptx, _extract_from_oaf_pptx):
    - Reads .pptx as zip archive (no python-pptx dependency for chart data)
    - Builds slide-to-chart map from _rels files
    - Extracts chart XML from ppt/charts/*.xml
    - Parses <c:ser> series blocks with category/value arrays
    - Separates "Average Shiftworker" series (norms) from client series
    - Stores each question as individual survey_norm and survey_client_result patterns
    - _detect_pptx_subtype() auto-detects survey vs OAF from slide text signals
    - OAF extractor reads slide tables (a:tbl/a:tr/a:tc), extracts operational metrics
  *
  * EXCEL EXTRACTOR (_extract_from_excel):
    - Multi-sheet aware: processes every sheet individually
    - Per-sheet type detection: cost_model, schedule_pattern, schedule_staffing,
      historical_data, general
    - cost_model: scenario names from row 1 cols J-O, labels from col I, two-block layout
    - schedule_pattern: detects "Week" header rows, extracts shift-code grids,
      groups into rotation cycles
    - schedule_staffing: extracts crew names, shift assignments, headcount rows
    - historical_data: stores summary stats only (skips raw date-value rows)
  *
  * BACKWARD COMPATIBILITY: All existing string-based extractors (contract,
    implementation_manual, lessons_learned, scope_of_work, general_word, etc.)
    are completely unchanged. New extractors are additive only. The ingest_document()
    method gains an optional file_bytes kwarg; if absent, behavior is identical
    to before.

- February 26, 2026 (Session 1): COMPLETE REWRITE - Real Knowledge Extraction
  * CONTRACT EXTRACTOR: client, total fee, payment schedule, weeks, interest rate.
  * IMPLEMENTATION MANUAL EXTRACTOR: schedule patterns, shift lengths, start times,
    facility context, appendices, employee selection process.
  * LESSONS LEARNED EXTRACTOR: all 19 lessons parsed individually with title,
    category, situation, key principle, real examples, full text.
  * SCOPE OF WORK EXTRACTOR: client, weekly deliverables, cost, pre-project data.
  * GENERAL WORD DOC: bold headings, dollar amounts, schedule patterns, percentages.

- February 3, 2026: CRITICAL FIX - Use KNOWLEDGE_DB_PATH env var for persistent storage
- February 2, 2026: Original file created

SUPPORTED DOCUMENT TYPES (all functional):
  Word Documents (text/Markdown format on server):
    - contract
    - implementation_manual
    - lessons_learned
    - scope_of_work
    - general_word (fallback)

  PowerPoint (file_bytes required):
    - survey_pptx / eaf   -> per-question chart data, normative vs client series
    - oaf                 -> slide tables, operational metrics, FTE scenarios

  Excel (file_bytes required):
    - excel               -> auto-detects sheet types, extracts all structured data

Author: Jim @ Shiftwork Solutions LLC (managed by Claude Sonnet 4)
"""

import hashlib
import json
import re
import os
import io
import zipfile
from datetime import datetime
from typing import Dict, List, Any, Optional
import sqlite3


# All known schedule pattern names for detection across document types
KNOWN_SCHEDULE_PATTERNS = [
    '2-2-3', '2-3-2', '4-3', '4/3', 'DuPont', 'Panama', 'Pitman',
    '3-on-3-off', 'Continental', '5-2-5-3', '5&2', '5/2', 'Rotating 8',
    '4x10', '4x12', '3x12', '12-hour', '8-hour', '10-hour',
    'fixed days off', 'rotating', 'every other weekend'
]

# Dollar amount pattern for Markdown-escaped files (Pandoc converts $ to \$)
DOLLAR_PATTERN = re.compile(r'\\{1,2}\$([0-9,]+)')

# Engagement week pattern
WEEK_PATTERN = re.compile(r'Week\s+(\d+)', re.IGNORECASE)

# Valid shift codes for schedule grids
SHIFT_CODE_SET = {
    'D12', 'N12', 'D11', 'N11', 'D10', 'N10', 'D8', 'N8', 'E8', 'S8',
    'd12', 'n12', 'd11', 'n11', 'd10', 'n10', 'd8', 'n8', 'e8', 's8',
    'OFF', 'off', 'D', 'N', 'E', 'O'
}


def _is_shift_code(v) -> bool:
    """Return True if a cell value looks like a shift code"""
    if v is None:
        return False
    sv = str(v).strip()
    if sv.upper() in SHIFT_CODE_SET:
        return True
    if sv == '12':
        return True
    return False


class DocumentIngestor:
    """
    Ingests documents and extracts knowledge permanently to database.

    Supports:
    - Word docs via text/Markdown content (existing path, unchanged)
    - PowerPoint via file_bytes (new: chart XML extraction from zip)
    - Excel via file_bytes (new: multi-sheet structured extraction)
    """

    def __init__(self, db_path=None):
        # CRITICAL: Use environment variable for persistent storage
        if db_path is None:
            db_path = os.environ.get('KNOWLEDGE_DB_PATH', 'swarm_intelligence.db')
        self.db_path = db_path
        print(f"📚 Knowledge Database: {self.db_path}")
        self._ensure_tables()

    def _ensure_tables(self):
        """Create knowledge persistence tables if they don't exist"""
        db = sqlite3.connect(self.db_path)
        cursor = db.cursor()

        cursor.execute('''
            CREATE TABLE IF NOT EXISTS knowledge_extracts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                document_type TEXT NOT NULL,
                document_name TEXT,
                extracted_data TEXT NOT NULL,
                client TEXT,
                industry TEXT,
                project_type TEXT,
                source_hash TEXT UNIQUE,
                extracted_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                file_size INTEGER,
                metadata TEXT
            )
        ''')

        cursor.execute('''
            CREATE TABLE IF NOT EXISTS learned_patterns (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                pattern_type TEXT NOT NULL,
                pattern_name TEXT,
                pattern_data TEXT NOT NULL,
                confidence REAL DEFAULT 0.5,
                supporting_documents INTEGER DEFAULT 1,
                first_seen TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                last_updated TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                metadata TEXT
            )
        ''')

        cursor.execute('''
            CREATE TABLE IF NOT EXISTS ingestion_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                document_name TEXT,
                document_type TEXT,
                status TEXT,
                patterns_extracted INTEGER DEFAULT 0,
                insights_extracted INTEGER DEFAULT 0,
                error_message TEXT,
                ingested_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        db.commit()
        db.close()

    # =========================================================================
    # MAIN INGEST ENTRY POINT
    # =========================================================================

    def ingest_document(self, content: str = None, document_type: str = 'general_word',
                        metadata: Dict = None, file_bytes: bytes = None) -> Dict[str, Any]:
        """
        Ingest a document and extract knowledge.

        Args:
            content:       Full document text (Markdown) — for Word doc path
            document_type: One of: contract, implementation_manual, lessons_learned,
                           scope_of_work, general_word, survey_pptx, oaf, eaf, excel
            metadata:      Optional dict with client, industry, project_type, document_name
            file_bytes:    Raw file bytes — required for PPTX and Excel; takes priority
                           over content string for those types.

        Returns:
            Ingestion result dict with counts, highlights, and status message
        """
        metadata = metadata or {}
        document_name = metadata.get('document_name', 'Untitled')

        # Deduplicate via content/bytes hash
        hash_source = file_bytes if file_bytes is not None else (content or '').encode()
        content_hash = hashlib.md5(hash_source).hexdigest()

        db = sqlite3.connect(self.db_path)
        cursor = db.cursor()
        cursor.execute('SELECT id FROM knowledge_extracts WHERE source_hash = ?', (content_hash,))
        if cursor.fetchone():
            db.close()
            return {
                'success': True,
                'already_ingested': True,
                'message': f'{document_name} already in knowledge base'
            }
        db.close()

        # ---- Route to correct extractor ----

        if file_bytes is not None and document_type in ('survey_pptx', 'eaf'):
            extracted = self._extract_from_survey_pptx(file_bytes, metadata)

        elif file_bytes is not None and document_type == 'oaf':
            extracted = self._extract_from_oaf_pptx(file_bytes, metadata)

        elif file_bytes is not None and document_type in ('excel', 'schedule_pattern', 'data_file'):
            extracted = self._extract_from_excel(file_bytes, metadata)

        # Auto-detect PPTX subtype when document_type not explicitly set
        elif file_bytes is not None and document_name.lower().endswith(('.pptx', '.ppt')):
            subtype = self._detect_pptx_subtype(file_bytes)
            if subtype == 'oaf':
                extracted = self._extract_from_oaf_pptx(file_bytes, metadata)
            else:
                extracted = self._extract_from_survey_pptx(file_bytes, metadata)

        # Auto-detect Excel
        elif file_bytes is not None and document_name.lower().endswith(('.xlsx', '.xls')):
            extracted = self._extract_from_excel(file_bytes, metadata)

        # Word doc / text paths (original, unchanged)
        elif document_type == 'contract':
            extracted = self._extract_from_contract(content or '', metadata)
        elif document_type == 'implementation_manual':
            extracted = self._extract_from_implementation_manual(content or '', metadata)
        elif document_type == 'lessons_learned':
            extracted = self._extract_from_lessons_learned(content or '', metadata)
        elif document_type == 'scope_of_work':
            extracted = self._extract_from_scope_of_work(content or '', metadata)
        elif document_type == 'implementation_ppt':
            extracted = self._extract_from_implementation_ppt(content or '', metadata)
        else:
            extracted = self._extract_general_word_doc(content or '', metadata)

        # Auto-detect client from content if not in metadata
        if not metadata.get('client') and content:
            detected_client = self._detect_client(content, extracted)
            if detected_client:
                metadata['client'] = detected_client
                extracted['detected_client'] = detected_client

        # Store to database
        self._store_extraction(
            document_type=document_type,
            document_name=document_name,
            extracted_data=extracted,
            source_hash=content_hash,
            metadata=metadata,
            client=metadata.get('client', ''),
            industry=metadata.get('industry', ''),
            file_size=len(file_bytes) if file_bytes else len(content or '')
        )

        patterns_added = self._update_cumulative_patterns(extracted)

        self._log_ingestion(
            document_name=document_name,
            document_type=document_type,
            patterns_extracted=len(extracted.get('patterns', [])),
            insights_extracted=len(extracted.get('insights', []))
        )

        totals = self.get_knowledge_base_stats()

        return {
            'success': True,
            'document_name': document_name,
            'document_type': document_type,
            'patterns_extracted': len(extracted.get('patterns', [])),
            'insights_extracted': len(extracted.get('insights', [])),
            'patterns_added_to_kb': patterns_added,
            'total_kb_size': totals['total_extracts'],
            'total_patterns': totals['total_patterns'],
            'highlights': extracted.get('highlights', []),
            'message': (
                f"✅ Ingested {document_name}. "
                f"Extracted {len(extracted.get('patterns', []))} patterns, "
                f"{len(extracted.get('insights', []))} insights. "
                f"Knowledge base now has {totals['total_patterns']} total patterns."
            )
        }

    # =========================================================================
    # POWERPOINT EXTRACTORS
    # =========================================================================

    def _detect_pptx_subtype(self, file_bytes: bytes) -> str:
        """
        Auto-detect whether a PPTX is a survey results deck or OAF/operations deck.
        Reads first 5 slides for signal words. Returns 'oaf' or 'survey'.
        """
        try:
            with zipfile.ZipFile(io.BytesIO(file_bytes), 'r') as z:
                all_files = z.namelist()
                slide_files = sorted(
                    [f for f in all_files if re.match(r'ppt/slides/slide\d+\.xml$', f)],
                    key=lambda x: int(re.search(r'slide(\d+)', x).group(1))
                )
                text_sample = ''
                for sf in slide_files[:5]:
                    xml = z.read(sf).decode('utf-8', errors='replace')
                    texts = re.findall(r'<a:t[^>]*>([^<]+)</a:t>', xml)
                    text_sample += ' '.join(texts)

            text_lower = text_sample.lower()
            oaf_signals = ['operations assessment', 'overtime', 'absenteeism',
                           'productivity', 'fte', 'headcount', 'staffing', 'turnover']
            survey_signals = ['survey', 'how long', 'shift', 'employee', 'schedule', 'prefer']
            oaf_score = sum(1 for s in oaf_signals if s in text_lower)
            survey_score = sum(1 for s in survey_signals if s in text_lower)
            return 'oaf' if oaf_score > survey_score else 'survey'
        except Exception:
            return 'survey'

    def _get_slide_question(self, slide_xml: str) -> Optional[str]:
        """
        Extract the survey question text from a slide's XML.
        Returns the first long text string that looks like a survey question.
        """
        raw_texts = [
            t.strip()
            for t in re.findall(r'<a:t[^>]*>([^<]+)</a:t>', slide_xml)
            if t.strip()
        ]
        for t in raw_texts:
            if (len(t) > 25
                    and not re.match(r'^[\d\.\:\t ]+$', t)
                    and 'Shiftwork' not in t
                    and 'Solutions' not in t
                    and '&amp;' not in t
                    and not t.startswith('N=')
                    and not t.startswith('n=')):
                return t
        return None

    def _get_chart_series(self, chart_xml: str) -> List[Dict]:
        """
        Parse chart XML and return list of series dicts.
        Each dict has keys: series (name), categories (list), values (list).
        """
        results = []
        for ser in re.findall(r'<c:ser>(.*?)</c:ser>', chart_xml, re.DOTALL):
            title_m = re.search(r'<c:tx>.*?<c:v>([^<]+)</c:v>', ser, re.DOTALL)
            ser_name = title_m.group(1).strip() if title_m else ''
            cat_block = re.search(r'<c:cat>(.*?)</c:cat>', ser, re.DOTALL)
            val_block = re.search(r'<c:val>(.*?)</c:val>', ser, re.DOTALL)
            cats = re.findall(r'<c:v>([^<]+)</c:v>', cat_block.group(1)) if cat_block else []
            vals = re.findall(r'<c:v>([^<]+)</c:v>', val_block.group(1)) if val_block else []
            results.append({'series': ser_name, 'categories': cats, 'values': vals})
        return results

    def _extract_from_survey_pptx(self, file_bytes: bytes, metadata: Dict) -> Dict:
        """
        Extract knowledge from Employee Assessment (EAF / survey results) PowerPoint.

        Reads the .pptx zip directly to access embedded chart XML — python-pptx
        reads slide TEXT only and completely misses chart data. This extractor:
        - Builds a slide-number → chart-file mapping from _rels files
        - For each slide: finds the survey question from slide text
        - Extracts chart series from ppt/charts/*.xml
        - Separates "Average Shiftworker" (normative) series from client series
        - Stores each question as individual patterns:
            survey_norm: normative distribution for that question
            survey_client_result: client-specific distribution
        """
        extracted = {
            'patterns': [],
            'insights': [],
            'highlights': [],
            'document_category': 'survey_pptx'
        }

        client = metadata.get('client', 'unknown')
        questions_processed = 0
        norms_found = 0
        errors_encountered = 0

        try:
            with zipfile.ZipFile(io.BytesIO(file_bytes), 'r') as z:
                all_files = z.namelist()

                slide_files = sorted(
                    [f for f in all_files if re.match(r'ppt/slides/slide\d+\.xml$', f)],
                    key=lambda x: int(re.search(r'slide(\d+)', x).group(1))
                )

                # Build slide-number -> chart-file list from .rels
                slide_chart_map: Dict[int, List[str]] = {}
                for rf in [f for f in all_files
                           if re.match(r'ppt/slides/_rels/slide\d+\.xml\.rels$', f)]:
                    sn = int(re.search(r'slide(\d+)', rf).group(1))
                    rel_xml = z.read(rf).decode('utf-8', errors='replace')
                    charts = re.findall(r'Target="\.\./charts/(chart\d+\.xml)"', rel_xml)
                    if charts:
                        slide_chart_map[sn] = charts

                for slide_num, sf in enumerate(slide_files, 1):
                    try:
                        slide_xml = z.read(sf).decode('utf-8', errors='replace')
                        question = self._get_slide_question(slide_xml)
                        if not question:
                            continue

                        # Gather all chart series for this slide
                        all_series: List[Dict] = []
                        for cf in slide_chart_map.get(slide_num, []):
                            chart_path = f'ppt/charts/{cf}'
                            if chart_path in all_files:
                                try:
                                    chart_xml = z.read(chart_path).decode('utf-8', errors='replace')
                                    all_series.extend(self._get_chart_series(chart_xml))
                                except Exception:
                                    pass

                        # Separate norm series from client series
                        norm_dist: Dict[str, float] = {}
                        client_dist: Dict[str, float] = {}

                        for ser in all_series:
                            dist: Dict[str, float] = {}
                            for cat, val in zip(ser['categories'], ser['values']):
                                try:
                                    dist[cat] = round(float(val), 2)
                                except (ValueError, TypeError):
                                    pass

                            if 'Average Shiftworker' in ser['series']:
                                norm_dist = dist
                            elif ser['series']:
                                client_dist.update(dist)

                        if not norm_dist and not client_dist:
                            continue

                        questions_processed += 1
                        q_slug = re.sub(r'[^a-z0-9]', '_', question.lower())[:60].strip('_')

                        if client_dist:
                            extracted['patterns'].append({
                                'type': 'survey_client_result',
                                'name': f'client_{q_slug}',
                                'data': {
                                    'question': question,
                                    'client': client,
                                    'distribution': client_dist,
                                    'slide': slide_num
                                },
                                'confidence': 0.95
                            })

                        if norm_dist:
                            norms_found += 1
                            extracted['patterns'].append({
                                'type': 'survey_norm',
                                'name': f'norm_{q_slug}',
                                'data': {
                                    'question': question,
                                    'distribution': norm_dist,
                                    'source': 'pptx_average_shiftworker_series'
                                },
                                'confidence': 0.98
                            })

                    except Exception:
                        errors_encountered += 1
                        continue

        except Exception as zip_err:
            extracted['highlights'].append(f"Error reading PPTX: {str(zip_err)}")
            return extracted

        extracted['insights'].append({
            'type': 'survey_summary',
            'client': client,
            'questions_processed': questions_processed,
            'normative_questions': norms_found,
            'slide_errors': errors_encountered
        })

        extracted['highlights'].append(f"Survey Questions Extracted: {questions_processed}")
        extracted['highlights'].append(f"Normative Benchmarks: {norms_found}")
        if client and client != 'unknown':
            extracted['highlights'].append(f"Client: {client}")

        return extracted

    def _extract_from_oaf_pptx(self, file_bytes: bytes, metadata: Dict) -> Dict:
        """
        Extract knowledge from Operations Assessment (OAF) PowerPoint.

        Reads slide tables (a:tbl XML) and text to extract:
        - Operational metrics: overtime %, absenteeism %, turnover %
        - Productivity tables (department x metric grids)
        - FTE / headcount scenarios
        - Cost comparison data
        """
        extracted = {
            'patterns': [],
            'insights': [],
            'highlights': [],
            'document_category': 'oaf'
        }

        client = metadata.get('client', 'unknown')
        metrics: Dict[str, Any] = {}
        tables_found = 0

        try:
            with zipfile.ZipFile(io.BytesIO(file_bytes), 'r') as z:
                all_files = z.namelist()
                slide_files = sorted(
                    [f for f in all_files if re.match(r'ppt/slides/slide\d+\.xml$', f)],
                    key=lambda x: int(re.search(r'slide(\d+)', x).group(1))
                )

                for slide_num, sf in enumerate(slide_files, 1):
                    try:
                        slide_xml = z.read(sf).decode('utf-8', errors='replace')
                        raw_texts = [
                            t.strip()
                            for t in re.findall(r'<a:t[^>]*>([^<]+)</a:t>', slide_xml)
                            if t.strip()
                        ]
                        slide_text = ' '.join(raw_texts)

                        # Extract tables
                        for tbl in re.findall(r'<a:tbl>(.*?)</a:tbl>', slide_xml, re.DOTALL):
                            rows_xml = re.findall(r'<a:tr>(.*?)</a:tr>', tbl, re.DOTALL)
                            table_data = []
                            for row_xml in rows_xml:
                                cells = re.findall(r'<a:tc>(.*?)</a:tc>', row_xml, re.DOTALL)
                                row_values = []
                                for cell in cells:
                                    cell_texts = re.findall(r'<a:t[^>]*>([^<]+)</a:t>', cell)
                                    row_values.append(
                                        ' '.join(t.strip() for t in cell_texts if t.strip())
                                    )
                                if any(row_values):
                                    table_data.append(row_values)
                            if table_data:
                                tables_found += 1
                                extracted['insights'].append({
                                    'type': 'oaf_table',
                                    'slide': slide_num,
                                    'data': table_data,
                                    'client': client
                                })

                        # Key metrics
                        ot_m = re.search(
                            r'[Oo]vertime[:\s]+(\d+(?:\.\d+)?)\s*%', slide_text)
                        if ot_m and 'overtime_pct' not in metrics:
                            metrics['overtime_pct'] = float(ot_m.group(1))

                        ab_m = re.search(
                            r'[Aa]bsenteeism[:\s]+(\d+(?:\.\d+)?)\s*%', slide_text)
                        if ab_m and 'absenteeism_pct' not in metrics:
                            metrics['absenteeism_pct'] = float(ab_m.group(1))

                        tn_m = re.search(
                            r'[Tt]urnover[:\s]+(\d+(?:\.\d+)?)\s*%', slide_text)
                        if tn_m and 'turnover_pct' not in metrics:
                            metrics['turnover_pct'] = float(tn_m.group(1))

                        hc_m = re.search(
                            r'(\d+)\s*(?:employees?|operators?|workers?|headcount)',
                            slide_text, re.IGNORECASE)
                        if hc_m and 'headcount' not in metrics:
                            hc = int(hc_m.group(1))
                            if 10 <= hc <= 5000:
                                metrics['headcount'] = hc

                    except Exception:
                        continue

        except Exception as zip_err:
            extracted['highlights'].append(f"Error reading OAF PPTX: {str(zip_err)}")
            return extracted

        if metrics:
            extracted['patterns'].append({
                'type': 'operational_metrics',
                'name': f'oaf_metrics_{re.sub(r"[^a-z0-9]", "_", client.lower())}',
                'data': {**metrics, 'client': client},
                'confidence': 0.85
            })

        extracted['insights'].append({
            'type': 'oaf_summary',
            'client': client,
            'metrics': metrics,
            'tables_found': tables_found
        })

        for k, v in metrics.items():
            label = k.replace('_', ' ').title()
            extracted['highlights'].append(f"{label}: {v}{'%' if 'pct' in k else ''}")
        extracted['highlights'].append(f"Tables Extracted: {tables_found}")

        return extracted

    # =========================================================================
    # EXCEL EXTRACTOR
    # =========================================================================

    def _detect_sheet_type(self, ws, sheet_name: str) -> str:
        """
        Detect what type of data a worksheet contains.
        Returns: cost_model, schedule_pattern, schedule_staffing,
                 historical_data, or general.
        """
        # Check col H/I area for cost model keywords (labels live in col I = index 8)
        for row in ws.iter_rows(min_row=1, max_row=30, max_col=12, values_only=True):
            for cell in row:
                if cell and isinstance(cell, str):
                    cl = cell.lower()
                    if any(kw in cl for kw in [
                        'cost of scheduled', 'overtime burden', 'benefits burden',
                        'cost of time', 'marginal cost'
                    ]):
                        return 'cost_model'

        # Schedule staffing: crews and staffing rows
        rows_sample = []
        for row in ws.iter_rows(min_row=1, max_row=10, max_col=3, values_only=True):
            rows_sample.append([str(v).strip().lower() if v else '' for v in row])
        flat = ' '.join(' '.join(r) for r in rows_sample)
        if any(kw in flat for kw in ['staffing', 'day shift staffing', 'total staffing']):
            return 'schedule_staffing'

        # Schedule pattern: "Week" headers followed by shift codes
        week_count = 0
        shift_count = 0
        for row in ws.iter_rows(min_row=1, max_row=20, values_only=True):
            vals = [str(v).strip() if v is not None else '' for v in row]
            if 'Week' in vals:
                week_count += 1
            shift_count += sum(1 for v in vals if _is_shift_code(v))

        if week_count >= 1 and shift_count >= 5:
            return 'schedule_pattern'

        # Historical data: date column
        for row in ws.iter_rows(min_row=1, max_row=5, max_col=3, values_only=True):
            for cell in row:
                if cell and re.match(r'\d{4}-\d{2}-\d{2}', str(cell)):
                    return 'historical_data'

        return 'general'

    def _extract_cost_model_sheet(self, ws, sheet_name: str, metadata: Dict) -> Dict:
        """
        Extract cost model scenarios from a worksheet.

        Layout (Cost_of_time_Best.xlsx):
        - Row 1: scenario names in columns J-O (0-indexed 9-14)
        - Column I (index 8): row labels
        - Values in scenario columns for each label row
        - Column C/D area: base input parameters
        """
        result = {
            'sheet': sheet_name,
            'type': 'cost_model',
            'scenarios': {},
            'inputs': {}
        }

        # Scenario names from row 1
        row1 = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        scenario_cols: Dict[int, str] = {}
        for i, v in enumerate(row1):
            if v and isinstance(v, str) and len(str(v).strip()) > 2:
                scenario_cols[i] = str(v).strip().replace('\n', ' ')

        LABEL_COL = 8  # Column I (0-indexed)

        for row in ws.iter_rows(values_only=True):
            row = list(row)
            label = (
                str(row[LABEL_COL]).strip()
                if LABEL_COL < len(row) and row[LABEL_COL] not in (None, '', ' ')
                else None
            )
            if not label or label == 'None':
                continue

            for col_idx, sc_name in scenario_cols.items():
                if col_idx < len(row) and isinstance(row[col_idx], (int, float)):
                    if sc_name not in result['scenarios']:
                        result['scenarios'][sc_name] = {}
                    result['scenarios'][sc_name][label] = round(float(row[col_idx]), 4)

        # Base inputs from col C (index 2) / col D (index 3)
        for row in ws.iter_rows(min_row=2, max_row=30, values_only=True):
            row = list(row)
            lbl = (
                str(row[2]).strip()
                if len(row) > 2 and row[2] not in (None, '', ' ')
                else None
            )
            val = row[3] if len(row) > 3 else None
            if lbl and isinstance(val, (int, float)):
                result['inputs'][lbl] = round(float(val), 4)

        return result

    def _extract_schedule_pattern_sheet(self, ws, sheet_name: str) -> Dict:
        """
        Extract shift rotation patterns from a worksheet.

        Each pattern block starts with a "Week" header row in col B.
        Data rows: week number in col B, shift codes in cols C-I.
        Patterns are separated by empty rows.
        """
        result = {
            'sheet': sheet_name,
            'type': 'schedule_patterns',
            'patterns': []
        }

        rows_list = list(ws.iter_rows(values_only=True))
        current_weeks: List[List[str]] = []
        current_header_row: Optional[int] = None
        current_day_headers: List[str] = []

        def save_pattern(start_row, day_hdrs, weeks):
            if not weeks:
                return
            all_codes = [c for wk in weeks for c in wk if c and c != '']
            shift_types = set()
            for c in all_codes:
                cu = c.upper()
                if '12' in c:
                    shift_types.add('12-hour')
                elif '10' in c:
                    shift_types.add('10-hour')
                elif '8' in c:
                    shift_types.add('8-hour')
                if 'D' in cu and cu not in ('OFF', ''):
                    shift_types.add('Day')
                if 'N' in cu and cu not in ('OFF', ''):
                    shift_types.add('Night')
                if 'E' in cu and cu not in ('OFF', ''):
                    shift_types.add('Evening')
            result['patterns'].append({
                'start_row': start_row,
                'cycle_weeks': len(weeks),
                'cycle_days': len(weeks) * 7,
                'day_headers': day_hdrs,
                'weeks': [
                    {f'day_{j+1}': s for j, s in enumerate(wk)}
                    for wk in weeks
                ],
                'shift_types': sorted(shift_types)
            })

        for i, row in enumerate(rows_list):
            row = list(row)
            col_b = row[1] if len(row) > 1 else None

            if col_b == 'Week':
                save_pattern(current_header_row, current_day_headers, current_weeks)
                current_header_row = i + 1
                current_day_headers = [
                    str(row[j]).strip() for j in range(2, 9)
                    if len(row) > j and row[j]
                ]
                current_weeks = []

            elif isinstance(col_b, int) and 1 <= col_b <= 20:
                week_shifts = [
                    str(row[j]).strip() if (len(row) > j and row[j] is not None) else ''
                    for j in range(2, 9)
                ]
                if any(_is_shift_code(row[j]) for j in range(2, 9) if len(row) > j):
                    current_weeks.append(week_shifts)

        save_pattern(current_header_row, current_day_headers, current_weeks)
        return result

    def _extract_schedule_staffing_sheet(self, ws, sheet_name: str) -> Dict:
        """
        Extract crew and staffing data from a client implementation sheet.
        Looks for crew rows (shift codes), staffing summary rows (headcount).
        """
        result = {
            'sheet': sheet_name,
            'type': 'schedule_staffing',
            'crews': [],
            'staffing': {}
        }

        for row in ws.iter_rows(values_only=True):
            row = list(row)
            label = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ''
            if not label:
                continue

            label_lower = label.lower()

            if 'staffing' in label_lower or ('total' in label_lower and 'staffing' in label_lower):
                day_vals = [
                    row[j] for j in range(2, 9)
                    if len(row) > j and isinstance(row[j], (int, float))
                ]
                if day_vals:
                    result['staffing'][label] = {f'day_{j+1}': v for j, v in enumerate(day_vals)}

            elif any(_is_shift_code(row[j]) for j in range(2, 9) if len(row) > j):
                shifts = [
                    str(row[j]).strip() if (len(row) > j and row[j] is not None) else ''
                    for j in range(2, 9)
                ]
                headcount = (
                    row[9] if len(row) > 9 and isinstance(row[9], (int, float)) else None
                )
                result['crews'].append({
                    'name': label,
                    'shifts': shifts,
                    'headcount': headcount
                })

        return result

    def _extract_from_excel(self, file_bytes: bytes, metadata: Dict) -> Dict:
        """
        Extract knowledge from an Excel file.

        Processes each sheet individually, auto-detects sheet type:
        - cost_model       → scenario labor cost comparisons
        - schedule_pattern → shift rotation grids
        - schedule_staffing→ crew assignments and headcount
        - historical_data  → summary only (skip raw rows)
        - general          → basic dimensions summary
        """
        extracted = {
            'patterns': [],
            'insights': [],
            'highlights': [],
            'document_category': 'excel'
        }

        client = metadata.get('client', 'unknown')

        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        except Exception as e:
            extracted['highlights'].append(f"Error reading Excel file: {str(e)}")
            return extracted

        sheets_processed = 0
        cost_models_found = 0
        schedule_patterns_found = 0
        staffing_sheets_found = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_type = self._detect_sheet_type(ws, sheet_name)

            if sheet_type == 'cost_model':
                cost_data = self._extract_cost_model_sheet(ws, sheet_name, metadata)
                if cost_data['scenarios']:
                    cost_models_found += 1
                    extracted['patterns'].append({
                        'type': 'cost_model',
                        'name': f'cost_model_{re.sub(r"[^a-z0-9]", "_", sheet_name.lower())}',
                        'data': cost_data,
                        'confidence': 0.9
                    })
                    for sc_name, sc_data in cost_data['scenarios'].items():
                        if 'Cost of Scheduled Time' in sc_data:
                            cost = sc_data['Cost of Scheduled Time']
                            extracted['highlights'].append(
                                f"{sc_name}: ${cost:.2f}/hr scheduled labor cost"
                            )
                    extracted['insights'].append({
                        'type': 'cost_model_summary',
                        'sheet': sheet_name,
                        'scenarios': list(cost_data['scenarios'].keys()),
                        'client': client
                    })

            elif sheet_type == 'schedule_pattern':
                pattern_data = self._extract_schedule_pattern_sheet(ws, sheet_name)
                if pattern_data['patterns']:
                    schedule_patterns_found += len(pattern_data['patterns'])
                    extracted['patterns'].append({
                        'type': 'schedule_rotation_library',
                        'name': f'rotations_{re.sub(r"[^a-z0-9]", "_", sheet_name.lower())}',
                        'data': pattern_data,
                        'confidence': 0.9
                    })
                    extracted['highlights'].append(
                        f"Sheet '{sheet_name}': {len(pattern_data['patterns'])} rotation patterns"
                    )
                    extracted['insights'].append({
                        'type': 'schedule_patterns_found',
                        'sheet': sheet_name,
                        'count': len(pattern_data['patterns']),
                        'cycle_lengths': list(set(
                            p['cycle_weeks'] for p in pattern_data['patterns']
                        ))
                    })

            elif sheet_type == 'schedule_staffing':
                staffing_data = self._extract_schedule_staffing_sheet(ws, sheet_name)
                if staffing_data['crews'] or staffing_data['staffing']:
                    staffing_sheets_found += 1
                    extracted['patterns'].append({
                        'type': 'schedule_staffing',
                        'name': f'staffing_{re.sub(r"[^a-z0-9]", "_", sheet_name.lower())}',
                        'data': staffing_data,
                        'confidence': 0.85
                    })
                    extracted['highlights'].append(
                        f"Sheet '{sheet_name}': {len(staffing_data['crews'])} crew assignments"
                    )
                    extracted['insights'].append({
                        'type': 'staffing_summary',
                        'sheet': sheet_name,
                        'crews': len(staffing_data['crews']),
                        'client': client
                    })

            elif sheet_type == 'historical_data':
                extracted['insights'].append({
                    'type': 'historical_data_present',
                    'sheet': sheet_name,
                    'row_count': ws.max_row,
                    'note': 'Raw historical data — summary only stored'
                })

            else:
                extracted['insights'].append({
                    'type': 'general_sheet',
                    'sheet': sheet_name,
                    'rows': ws.max_row,
                    'cols': ws.max_column
                })

            sheets_processed += 1

        extracted['insights'].append({
            'type': 'excel_file_summary',
            'document_name': metadata.get('document_name'),
            'client': client,
            'sheets_processed': sheets_processed,
            'cost_models': cost_models_found,
            'schedule_patterns': schedule_patterns_found,
            'staffing_sheets': staffing_sheets_found
        })

        extracted['highlights'].insert(0,
            f"Sheets: {sheets_processed} | Cost Models: {cost_models_found} | "
            f"Rotation Patterns: {schedule_patterns_found} | "
            f"Staffing: {staffing_sheets_found}"
        )

        return extracted

    # =========================================================================
    # WORD DOCUMENT EXTRACTORS (original, unchanged)
    # =========================================================================

    def _extract_from_contract(self, content: str, metadata: Dict) -> Dict:
        """
        Extract knowledge from contract/engagement agreement.
        Extracts: client, total fee, payment milestones, engagement weeks,
        interest rate, termination formula.
        """
        extracted = {
            'patterns': [],
            'insights': [],
            'highlights': [],
            'document_category': 'contract'
        }

        client_match = re.search(
            r'\*\*([A-Z][A-Z\s&,\.\-]+)\*\*\s*\(the\s+["\']?Client["\']?\)',
            content
        )
        client = client_match.group(1).strip() if client_match else metadata.get('client', '')

        fee_match = re.search(r'fixed sum of\\+\$([0-9,]+)', content)
        total_fee = fee_match.group(1).replace(',', '') if fee_match else None

        payments = re.findall(r'Due upon\s+([^\:]+):\s*\\+\$([0-9,]+)', content)
        payment_schedule = [
            {'milestone': m.strip(), 'amount': int(a.replace(',', ''))}
            for m, a in payments
        ]

        weeks = sorted(set(int(w) for w in WEEK_PATTERN.findall(content)))
        max_week = max(weeks) if weeks else None

        interest_match = re.search(r'(\d+)%\s+per year on balances unpaid', content)
        interest_rate = int(interest_match.group(1)) if interest_match else 11

        formula_match = re.search(r'Consulting Fees Due\s*=\s*([^\n]+)', content)
        termination_formula = formula_match.group(1).strip() if formula_match else None

        expense_match = re.search(r'receipts for expenses greater than\\+\$(\d+)', content)
        expense_threshold = int(expense_match.group(1)) if expense_match else 40

        contract_data = {
            'client': client,
            'total_fee': int(total_fee) if total_fee else None,
            'payment_schedule': payment_schedule,
            'engagement_weeks': max_week,
            'interest_rate_pct': interest_rate,
            'termination_formula': termination_formula,
            'expense_receipt_threshold': expense_threshold,
            'ip_ownership': (
                'consultant' if 'provided or developed by Consultant' in content else 'unknown'
            )
        }

        extracted['insights'].append({'type': 'contract_terms', 'data': contract_data})

        if total_fee:
            extracted['patterns'].append({
                'type': 'engagement_fee',
                'name': f'fee_{total_fee}',
                'data': {'fee': int(total_fee), 'client': client, 'weeks': max_week},
                'confidence': 0.95
            })

        if payment_schedule:
            extracted['patterns'].append({
                'type': 'payment_structure',
                'name': f'payment_milestones_{len(payment_schedule)}',
                'data': payment_schedule,
                'confidence': 0.95
            })

        extracted['patterns'].append({
            'type': 'contract_terms',
            'name': 'interest_rate',
            'data': {'rate_pct': interest_rate, 'days_before_interest': 30},
            'confidence': 1.0
        })

        if client:
            extracted['highlights'].append(f"Client: {client}")
        if total_fee:
            extracted['highlights'].append(f"Total Fee: ${int(total_fee):,}")
        if payment_schedule:
            for p in payment_schedule:
                extracted['highlights'].append(
                    f"Payment: ${p['amount']:,} due upon {p['milestone']}"
                )
        if max_week:
            extracted['highlights'].append(f"Engagement Duration: {max_week} weeks")

        return extracted

    def _extract_from_implementation_manual(self, content: str, metadata: Dict) -> Dict:
        """
        Extract knowledge from implementation manual.
        Extracts: schedule patterns, shift lengths, start times, facility context,
        appendix structure, employee selection process.
        """
        extracted = {
            'patterns': [],
            'insights': [],
            'highlights': [],
            'document_category': 'implementation_manual'
        }

        found_patterns = [p for p in KNOWN_SCHEDULE_PATTERNS if p in content]
        if found_patterns:
            extracted['patterns'].append({
                'type': 'schedule_patterns_presented',
                'name': 'schedules_in_manual',
                'data': found_patterns,
                'confidence': 0.9
            })
            extracted['highlights'].append(f"Schedule Patterns: {', '.join(found_patterns)}")

        shift_lengths = list(set(re.findall(r'(\d+)-hour shifts?', content, re.IGNORECASE)))
        if shift_lengths:
            extracted['patterns'].append({
                'type': 'shift_lengths',
                'name': 'hours_per_shift',
                'data': [int(h) for h in shift_lengths],
                'confidence': 0.95
            })
            extracted['highlights'].append(f"Shift Lengths: {', '.join(shift_lengths)}-hour")

        times = list(set(re.findall(r'(\d{1,2}:\d{2}\s*(?:am|pm))', content, re.IGNORECASE)))
        if times:
            extracted['insights'].append({'type': 'shift_start_times', 'times': times})

        appendices = re.findall(r'Appendix\s+[A-Z][:\s\-]+([A-Za-z][^\n\*\[\\]+)', content)
        appendices = [a.strip() for a in appendices if len(a.strip()) > 3]
        seen = set()
        unique_appendices = []
        for a in appendices:
            if a not in seen:
                seen.add(a)
                unique_appendices.append(a)
        if unique_appendices:
            extracted['insights'].append({
                'type': 'financial_analyses_included',
                'appendices': unique_appendices
            })
            extracted['highlights'].append(f"Appendices: {', '.join(unique_appendices[:4])}")

        if 'Preference Form' in content or 'preference form' in content:
            extracted['patterns'].append({
                'type': 'employee_involvement',
                'name': 'preference_form_used',
                'data': {'method': 'written_preference_form'},
                'confidence': 0.9
            })

        if '24/7' in content:
            extracted['patterns'].append({
                'type': 'operation_type',
                'name': '24_7_coverage',
                'data': {'is_24_7': True},
                'confidence': 0.95
            })
            extracted['highlights'].append("Operation Type: 24/7 coverage")

        headings = re.findall(r'^#+ (.+)$', content, re.MULTILINE)
        headings = [h.strip().replace('*', '') for h in headings if len(h.strip()) > 3]
        if headings:
            extracted['insights'].append({'type': 'manual_structure', 'sections': headings[:20]})

        if metadata.get('client'):
            extracted['insights'].append({
                'type': 'client_context',
                'client': metadata['client'],
                'industry': metadata.get('industry', 'unknown')
            })

        return extracted

    def _extract_from_lessons_learned(self, content: str, metadata: Dict) -> Dict:
        """
        Extract knowledge from lessons learned document.
        Parses each lesson individually: title, category, situation, principle, examples.
        Each lesson becomes its own searchable pattern.
        """
        extracted = {
            'patterns': [],
            'insights': [],
            'highlights': [],
            'document_category': 'lessons_learned'
        }

        lesson_blocks = re.split(r'(?=### Lesson #\d+)', content)
        lesson_blocks = [b.strip() for b in lesson_blocks if b.strip().startswith('### Lesson')]

        category_map = {}
        current_category = 'General'
        for line in content.split('\n'):
            line = line.strip()
            if (line.startswith('## ')
                    and 'Categories' not in line
                    and 'How to Add' not in line
                    and 'Notes for' not in line):
                current_category = line.replace('## ', '').strip()
            elif line.startswith('### Lesson'):
                title_match = re.match(r'### (Lesson #\d+)', line)
                if title_match:
                    category_map[title_match.group(1)] = current_category

        lessons_extracted = []

        for block in lesson_blocks:
            title_match = re.match(r'### (Lesson #(\d+)[^\n]+)', block)
            if not title_match:
                continue

            full_title = title_match.group(1).strip()
            lesson_num = int(title_match.group(2))
            lesson_name = re.sub(r'Lesson #\d+:\s*', '', full_title).strip()
            category = category_map.get(f"Lesson #{lesson_num}", 'General')

            situation_match = re.search(
                r'\*\*Situation[^*]*\*\*\s*\n+(.*?)(?=\n\*\*|\Z)', block, re.DOTALL
            )
            situation = situation_match.group(1).strip()[:300] if situation_match else ''

            principle_match = re.search(
                r'\*\*Key Principle[:\*\*\s]+(.+?)(?=\n---|\n###|\Z)', block, re.DOTALL
            )
            principle = principle_match.group(1).strip()[:400] if principle_match else ''

            examples = re.findall(
                r'\*\*Real Example[^\n]*\*\*\s*\n+(.*?)(?=\n\*\*|\n---|\Z)', block, re.DOTALL
            )
            examples = [e.strip()[:200] for e in examples]

            lesson_data = {
                'lesson_number': lesson_num,
                'title': full_title,
                'lesson_name': lesson_name,
                'category': category,
                'situation': situation,
                'key_principle': principle,
                'real_examples': examples,
                'full_text': block[:1000]
            }

            lessons_extracted.append(lesson_data)
            extracted['patterns'].append({
                'type': 'consulting_lesson',
                'name': (
                    f'lesson_{lesson_num}_'
                    f'{re.sub(r"[^a-z0-9]", "_", lesson_name.lower())[:40]}'
                ),
                'data': lesson_data,
                'confidence': 1.0
            })

        categories_covered = list(set(l['category'] for l in lessons_extracted))
        extracted['insights'].append({
            'type': 'lessons_learned_summary',
            'total_lessons': len(lessons_extracted),
            'categories': categories_covered,
            'lessons': [
                {
                    'number': l['lesson_number'],
                    'name': l['lesson_name'],
                    'category': l['category']
                }
                for l in lessons_extracted
            ]
        })

        extracted['highlights'].append(f"Total Lessons: {len(lessons_extracted)}")
        extracted['highlights'].append(f"Categories: {', '.join(categories_covered)}")

        return extracted

    def _extract_from_scope_of_work(self, content: str, metadata: Dict) -> Dict:
        """
        Extract knowledge from scope of work document.
        Extracts: client, total cost, weekly deliverables, pre-project requirements.
        """
        extracted = {
            'patterns': [],
            'insights': [],
            'highlights': [],
            'document_category': 'scope_of_work'
        }

        bold_items = re.findall(r'\*\*([^*\n]+)\*\*', content)
        client = None
        for item in bold_items:
            item = item.strip()
            if item in ('Attachment A: Scope of Work', 'Summary:', 'Detailed Description:'):
                continue
            if re.search(r'[A-Z]{3,}', item) and len(item) < 60:
                client = item
                break

        cost_match = re.search(r'\\+\$([0-9,]+)\s*\+?\s*travel', content, re.IGNORECASE)
        if not cost_match:
            cost_match = re.search(r'Project Cost[^$]*\\+\$([0-9,]+)', content, re.IGNORECASE)
        total_cost = int(cost_match.group(1).replace(',', '')) if cost_match else None

        weekly_deliverables = {}
        week_sections = re.finditer(
            r'\*\*Week\s+(\d+)[:\*\s]+(.*?)(?=\*\*Week\s+\d+|\*\*Follow|\*\*Project Cost|'
            r'\*\*Detailed|\Z)',
            content, re.DOTALL
        )
        for match in week_sections:
            week_num = int(match.group(1))
            week_content = match.group(2).strip()
            bullets = [b.strip() for b in re.findall(r'-\s+(.+)', week_content) if b.strip()]
            if bullets:
                weekly_deliverables[f'week_{week_num}'] = bullets

        pre_project = []
        pre_match = re.search(
            r'Pre-Project[^\n]*\n+(.*?)(?=\*\*Project Week 1|\Z)',
            content, re.DOTALL | re.IGNORECASE
        )
        if pre_match:
            pre_project = [
                b.strip() for b in re.findall(r'-\s+(.+)', pre_match.group(1)) if b.strip()
            ]

        followup_match = re.search(
            r'\*\*Follow[- ]?up[:\*\s]+(.*?)(?=\*\*Project Cost|\*\*Detailed|\Z)',
            content, re.DOTALL | re.IGNORECASE
        )
        followup = []
        if followup_match:
            followup = [
                b.strip()
                for b in re.findall(r'-\s+(.+)', followup_match.group(1))
                if b.strip()
            ]

        sow_data = {
            'client': client,
            'total_cost': total_cost,
            'weekly_deliverables': weekly_deliverables,
            'pre_project_data_requirements': pre_project,
            'follow_up_commitments': followup,
            'total_weeks': len(weekly_deliverables)
        }

        extracted['insights'].append({'type': 'scope_of_work', 'data': sow_data})

        if total_cost:
            extracted['patterns'].append({
                'type': 'engagement_cost',
                'name': f'sow_cost_{total_cost}',
                'data': {'cost': total_cost, 'client': client},
                'confidence': 0.95
            })

        if weekly_deliverables:
            extracted['patterns'].append({
                'type': 'project_structure',
                'name': f'week_by_week_{len(weekly_deliverables)}_weeks',
                'data': weekly_deliverables,
                'confidence': 0.9
            })

        if client:
            extracted['highlights'].append(f"Client: {client}")
        if total_cost:
            extracted['highlights'].append(f"Project Cost: ${total_cost:,} + travel")
        if weekly_deliverables:
            extracted['highlights'].append(f"Project Duration: {len(weekly_deliverables)} weeks")
        if pre_project:
            extracted['highlights'].append(f"Pre-project data items: {len(pre_project)}")

        return extracted

    def _extract_general_word_doc(self, content: str, metadata: Dict) -> Dict:
        """
        Improved generic extractor for any Word document not matched by specific extractor.
        Extracts: bold headings, dollar amounts, schedule patterns, percentages.
        """
        extracted = {
            'patterns': [],
            'insights': [],
            'highlights': [],
            'document_category': 'general_word'
        }

        headings = re.findall(r'\*\*([^*\n]{5,80})\*\*', content)
        headings = [h.strip() for h in headings if not h.strip().endswith(':')]
        if headings:
            extracted['insights'].append({'type': 'document_structure', 'headings': headings[:30]})

        amounts = [int(a.replace(',', '')) for a in DOLLAR_PATTERN.findall(content)]
        if amounts:
            extracted['insights'].append({'type': 'financial_figures', 'amounts': amounts})
            extracted['highlights'].append(
                f"Dollar amounts found: {', '.join(f'${a:,}' for a in sorted(set(amounts))[:5])}"
            )

        found_patterns = [p for p in KNOWN_SCHEDULE_PATTERNS if p in content]
        if found_patterns:
            extracted['patterns'].append({
                'type': 'schedule_patterns_mentioned',
                'name': 'schedules_referenced',
                'data': found_patterns,
                'confidence': 0.8
            })
            extracted['highlights'].append(f"Schedule patterns: {', '.join(found_patterns)}")

        percentages = [float(p) for p in re.findall(r'(\d+(?:\.\d+)?)\s*%', content)]
        if percentages:
            extracted['insights'].append({
                'type': 'percentage_figures',
                'values': sorted(set(percentages))[:10]
            })

        extracted['insights'].append({
            'type': 'document_summary',
            'content_length': len(content),
            'word_count': len(content.split()),
            'document_name': metadata.get('document_name', 'unknown')
        })

        return extracted

    def _extract_from_implementation_ppt(self, content: str, metadata: Dict) -> Dict:
        """
        Extract from implementation presentation (text content path).
        For .pptx files, prefer the file_bytes path (_extract_from_survey_pptx /
        _extract_from_oaf_pptx) which provides chart data access.
        """
        extracted = self._extract_general_word_doc(content, metadata)
        extracted['document_category'] = 'implementation_ppt'

        found_patterns = [p for p in KNOWN_SCHEDULE_PATTERNS if p in content]
        if found_patterns:
            extracted['insights'].append({
                'type': 'schedules_implemented',
                'patterns': found_patterns,
                'client': metadata.get('client')
            })

        return extracted

    # =========================================================================
    # HELPER METHODS (original, unchanged)
    # =========================================================================

    def _detect_client(self, content: str, extracted: Dict) -> Optional[str]:
        """Auto-detect client name from content if not provided in metadata."""
        for insight in extracted.get('insights', []):
            if insight.get('type') == 'contract_terms':
                return insight['data'].get('client')
            if insight.get('type') == 'scope_of_work':
                return insight['data'].get('client')

        candidates = re.findall(r'\*\*([A-Z][A-Z\s&,\.\-]{3,40})\*\*', content)
        for candidate in candidates:
            candidate = candidate.strip()
            skip = {'SAMPLE', 'TABLE OF CONTENTS', 'APPENDIX', 'INTRODUCTION',
                    'SUMMARY', 'SCOPE OF WORK', 'NOTE', 'IMPORTANT', 'WARNING'}
            if candidate not in skip and len(candidate) >= 4:
                return candidate

        return None

    def _store_extraction(self, document_type: str, document_name: str,
                          extracted_data: Dict, source_hash: str, metadata: Dict,
                          client: str = '', industry: str = '', file_size: int = 0):
        """Store extracted knowledge in database"""
        db = sqlite3.connect(self.db_path)
        cursor = db.cursor()

        cursor.execute('''
            INSERT INTO knowledge_extracts (
                document_type, document_name, extracted_data,
                client, industry, project_type, source_hash, file_size, metadata
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            document_type,
            document_name,
            json.dumps(extracted_data),
            client,
            industry,
            metadata.get('project_type'),
            source_hash,
            file_size,
            json.dumps(metadata)
        ))

        db.commit()
        db.close()

    def _update_cumulative_patterns(self, extracted: Dict) -> int:
        """
        Update cumulative learned patterns.
        Patterns grow stronger (higher confidence) with each supporting document.
        """
        patterns_added = 0
        db = sqlite3.connect(self.db_path)
        cursor = db.cursor()

        for pattern in extracted.get('patterns', []):
            pattern_key_type = pattern.get('type', 'unknown')
            pattern_key_name = pattern.get('name', 'unknown')

            cursor.execute('''
                SELECT id, supporting_documents, confidence FROM learned_patterns
                WHERE pattern_type = ? AND pattern_name = ?
            ''', (pattern_key_type, pattern_key_name))

            existing = cursor.fetchone()

            if existing:
                pattern_id, support_count, current_confidence = existing
                new_support = support_count + 1
                new_confidence = min(0.99, current_confidence + (0.05 / new_support))
                cursor.execute('''
                    UPDATE learned_patterns
                    SET supporting_documents = ?,
                        confidence = ?,
                        last_updated = ?,
                        pattern_data = ?
                    WHERE id = ?
                ''', (
                    new_support,
                    new_confidence,
                    datetime.now(),
                    json.dumps(pattern.get('data', {})),
                    pattern_id
                ))
            else:
                cursor.execute('''
                    INSERT INTO learned_patterns (
                        pattern_type, pattern_name, pattern_data, confidence
                    ) VALUES (?, ?, ?, ?)
                ''', (
                    pattern_key_type,
                    pattern_key_name,
                    json.dumps(pattern.get('data', {})),
                    pattern.get('confidence', 0.5)
                ))
                patterns_added += 1

        db.commit()
        db.close()
        return patterns_added

    def _log_ingestion(self, document_name: str, document_type: str,
                       patterns_extracted: int, insights_extracted: int,
                       error_message: str = None):
        """Log ingestion for tracking"""
        db = sqlite3.connect(self.db_path)
        cursor = db.cursor()
        cursor.execute('''
            INSERT INTO ingestion_log (
                document_name, document_type, status,
                patterns_extracted, insights_extracted, error_message
            ) VALUES (?, ?, ?, ?, ?, ?)
        ''', (
            document_name,
            document_type,
            'success' if not error_message else 'error',
            patterns_extracted,
            insights_extracted,
            error_message
        ))
        db.commit()
        db.close()

    def get_knowledge_base_stats(self) -> Dict[str, Any]:
        """Get knowledge base statistics"""
        db = sqlite3.connect(self.db_path)
        db.row_factory = sqlite3.Row
        cursor = db.cursor()

        cursor.execute('SELECT COUNT(*) as count FROM knowledge_extracts')
        total_extracts = cursor.fetchone()['count']

        cursor.execute('SELECT COUNT(*) as count FROM learned_patterns')
        total_patterns = cursor.fetchone()['count']

        cursor.execute('''
            SELECT document_type, COUNT(*) as count
            FROM knowledge_extracts
            GROUP BY document_type
        ''')
        by_type = {row['document_type']: row['count'] for row in cursor.fetchall()}

        cursor.execute('''
            SELECT document_name, document_type, status, patterns_extracted,
                   insights_extracted, ingested_at
            FROM ingestion_log
            WHERE status = 'success'
            ORDER BY ingested_at DESC
            LIMIT 10
        ''')
        recent = [dict(row) for row in cursor.fetchall()]

        db.close()
        return {
            'total_extracts': total_extracts,
            'total_patterns': total_patterns,
            'by_document_type': by_type,
            'recent_ingestions': recent
        }

    def query_knowledge(self, query: str, document_type: str = None,
                        client: str = None, limit: int = 10) -> List[Dict]:
        """
        Query the knowledge base for relevant extracts.

        Args:
            query: Search terms
            document_type: Optional filter by doc type
            client: Optional filter by client
            limit: Max results

        Returns:
            List of matching knowledge extracts
        """
        db = sqlite3.connect(self.db_path)
        db.row_factory = sqlite3.Row

        sql = 'SELECT * FROM knowledge_extracts WHERE 1=1'
        params = []

        if document_type:
            sql += ' AND document_type = ?'
            params.append(document_type)

        if client:
            sql += ' AND client LIKE ?'
            params.append(f'%{client}%')

        if query:
            sql += ' AND (document_name LIKE ? OR extracted_data LIKE ?)'
            params.extend([f'%{query}%', f'%{query}%'])

        sql += ' ORDER BY extracted_at DESC LIMIT ?'
        params.append(limit)

        rows = db.execute(sql, params).fetchall()
        db.close()

        results = []
        for row in rows:
            item = dict(row)
            if item.get('extracted_data'):
                try:
                    item['extracted_data'] = json.loads(item['extracted_data'])
                except Exception:
                    pass
            results.append(item)

        return results


# =========================================================================
# SINGLETON
# =========================================================================

_ingestor = None


def get_document_ingestor() -> DocumentIngestor:
    """Get singleton document ingestor"""
    global _ingestor
    if _ingestor is None:
        _ingestor = DocumentIngestor()
    return _ingestor


# I did no harm and this file is not truncated
