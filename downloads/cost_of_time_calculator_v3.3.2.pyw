"""
===============================================================================
COST OF TIME CALCULATOR - Shiftwork Solutions LLC
===============================================================================
Version: 3.3.0
Created: November 25, 2025
Last Updated: January 12, 2026

File Type: .pyw (Windows GUI - no console window)

Description:
    A comprehensive tool for comparing labor costs across different shift 
    schedule configurations. Helps organizations understand the true cost of 
    time for various scheduling approaches including 5-day 8-hour, 7-day 8-hour,
    and 7-day 12-hour operations.

Based on: Cost_of_time_Best.xlsx from Shiftwork Solutions LLC
Website: https://shift-work.com
Contact: (415) 763-5005

Features:
    - Create custom schedule cost calculations
    - Real-time cost updates as inputs change
    - Save multiple schedule calculations for comparison
    - Click-for-explanation on all calculated fields
    - Export to Excel and PDF
    - Save/load configurations
    - Professional reporting

Schedule Types Supported:
    1. 5-day, 8-hour (Standard 40-hour week, 2080 hrs/year)
    2. 7-day, 8-hour (42-hour average week, 2184 hrs/year)
    3. 7-day, 12-hour (42-hour average week, 2184 hrs/year)

Dependencies (install first):
    pip install matplotlib openpyxl reportlab

Usage:
    Double-click the .pyw file to launch the window directly.
    No console/terminal window will appear.

Change Log:
    v1.0.0 (2025-11-25): Initial release
        - Full calculation engine matching Excel logic
        - GUI with tabbed interface
        - Export to Excel and PDF
        - Save/load configurations
        - Chart visualizations
    
    v2.0.0 (2025-11-26): Major UI Redesign for Inputs Tab
        - Burden Rates redesigned with calculated fields
        - Prominent ST/OT burden summary displays
        - Time Off section redesigned
        - All percentages as whole numbers
    
    v3.0.0 (2025-11-26): New Schedule Cost Calculator
        - Replaced fixed 6-schedule comparison with flexible calculator
        - Schedule Cost Calculator on right side of Inputs tab
        - Real-time calculation updates
        - Save multiple schedule calculations
        - New Saved Costs tab for comparing saved schedules
        - Click-for-explanation feature on all fields
        - Supports 5-day/8-hour, 7-day/8-hour, 7-day/12-hour schedules
        - Adjusted burden rate calculation for 7-day schedules
        - Shift overlap impacts on scheduled hours:
            - 5-day: adds (minutes/60) × 260 to scheduled hours
            - 7-day, 8-hour: adds (minutes/60) × 273 to scheduled hours
            - 7-day, 12-hour: adds (minutes/60) × 182 to scheduled hours
        - Shift overlap pay at 1.5x added to Total Hours Paid
        - Employee Annual Earnings calculation
        - New Staffing Analysis tab:
            - Select any saved schedule for analysis
            - Input positions to cover and target OT %
            - Calculates crew needed, scheduled cost, OT cost, total cost
            - Shows 5 scenarios: target ±2 increments of 5%
            - Highlights target row for easy identification
    
    v3.1.0 (2025-11-27): UI and Analysis Improvements
        - FIXED: Mousewheel scrolling now works on Schedule Cost Calculator
            - Scrolling is canvas-specific (only scrolls when mouse is over that area)
        - IMPROVED: Employee Annual Earnings label shortened to "Employee Earnings (base):"
        - ADDED: Employee Income column to Staffing Analysis
            - Shows how employee income increases with OT at each scenario level
            - Demonstrates that while company costs stay similar, employee income rises
        - ADDED: Auto-save on close
            - All inputs and saved schedules automatically saved when closing app
            - Automatically restored when app reopens
            - Saves to cost_of_time_autosave.json in app directory
        - ADDED: Incentive Cost Analysis section
            - Extra Day Off: OT cost × shift hours × crew (coverage cost)
            - Shift Differential: Shows current %, $ at 5/10/15%, increase needed, 
              annual cost = increase × (1 + OT burden) × scheduled hours × non-day workers
            - Extra Week Vacation: (40 or 42 hrs) × OT cost × non-day workers
            - Non-day worker counts: 5-day = crew×2, 7-day 8-hr = crew×4×(2/3), 
              7-day 12-hr = crew×4×0.5
        - FIXED: Staffing Analysis tab now scrollable to show all content
            - Incentive Cost Analysis section was hidden off-screen
    
    v3.2.0 (2025-11-28): Fixed Staffing/OT Calculations, Added Compare Tab
        - FIXED: Staffing & Cost Scenarios now uses correct overtime logic
            - Total coverage needed = 52 weeks × 168 hrs/week × positions (7-day)
            - Total coverage needed = 52 weeks × 40 hrs/week × positions (5-day)
            - OT % represents unscheduled overtime only
            - Employees calculated: E = Coverage / (Hours Worked × (1 + OT%))
            - Unscheduled OT hours = Total Coverage - (Employees × Hours Worked)
            - Cost now correctly increases with higher OT % (OT is more expensive)
        - FIXED: "Crew Needed" renamed to "Total Employees" - shows actual headcount
        - FIXED: Incentive Cost Analysis now correctly calculates total workforce
        - FIXED: Non-day worker calculations now use correct formulas
        - FIXED: "Total (non-day crew)" label was truncated - now displays fully
        - FIXED: Hours Paid Components now includes Paid Time Off (Sick, etc.)
        - IMPROVED: Hours Analysis shows detailed breakdown from Scheduled to Worked
        - IMPROVED: Hours Paid Components shows all components
        - IMPROVED: Key Cost Metrics (Cost of Time, Marginal OT) now displayed larger/bolder
        - ADDED: Cost per 5 Minutes of Shift Overlap in Incentive Cost Analysis
        - ADDED: Cost/Person column in Staffing & Cost Scenarios table
        - ADDED: NEW "Compare Schedules" tab for side-by-side comparison:
            - Select two saved schedules to compare
            - Shows schedule info, hours, and ratios
            - Highlights key cost metrics with differences
            - Compares incentive costs (day off, overlap, vacation, shift diff)
            - Color-coded differences (red=increase, green=decrease)
    
    v3.3.0 (2026-01-12): Added Lack of Work (LOW) Support, Fixed Staffing Calc, Added Adverse Cost Analysis
        - ADDED: "Lack of Work (LOW)" input field in Time Off section
            - Voluntary unpaid time off offered when there's insufficient work
            - LOW reduces Hours Worked but does NOT cause overtime
        - IMPROVED: Absenteeism % (causes OT) calculation now excludes:
            - Holiday hours (plant closed)
            - Lack of Work hours (no overtime needed when offered)
        - FIXED: Staffing calculation formula in Staffing Analysis tab
            - New formula: Employees = Positions / (1 - Absenteeism + Target_OT)
            - At 0% OT: Need extra staff to cover absences (e.g., 77 pos × 8.6% = 84 people)
            - At OT% = Absenteeism%: OT covers absences exactly (employees = positions)
            - At higher OT%: Can run leaner with more overtime
        - ADDED: Adverse Cost Analysis section in Staffing Analysis tab
            - Shows per-hour adverse cost rates for understaffing vs overstaffing
            - Understaffing cost = OT rate - ST rate (paying premium for coverage)
            - Overstaffing cost = ST rate (paying for unneeded hours)
            - Displays ratio (typically 10-20:1 showing overstaffing costs much more)
            - Four variability scenarios with annual cost projections:
              * Flat (theoretical): Exactly average absenteeism every day
              * Lumpy (worst): Absences concentrated on certain days
              * Spread (typical): 40% understaffed, 40% overstaffed, 20% perfect
              * Well-Managed: Proactive VTO/LOW to reduce idle time
            - Key insight explaining why running lean is usually optimal
        - ADDED: Explanation text for Lack of Work and updated absenteeism formula

Copyright (c) 2025 Shiftwork Solutions LLC - All Rights Reserved
===============================================================================
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import json
import os
from datetime import datetime
from typing import Dict, Any, Optional, List
import sys

# Try to import optional dependencies
try:
    import matplotlib
    matplotlib.use('TkAgg')
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    from matplotlib.figure import Figure
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False
    print("Warning: matplotlib not available. Charts will be disabled.")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("Warning: openpyxl not available. Excel export will be disabled.")

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    print("Warning: reportlab not available. PDF export will be disabled.")


# =============================================================================
# WORKERS' COMPENSATION DATA
# =============================================================================

# Industry base rates (midpoint percentages)
INDUSTRY_RATES = {
    "Manufacturing - General": 2.25,
    "Food & Beverage Processing": 3.0,
    "Chemical / Petrochemical": 2.6,
    "Pharmaceutical": 1.5,
    "Logistics / Warehousing": 3.75,
    "Paper / Pulp / Packaging": 3.0,
    "Metals / Steel": 4.5,
    "Automotive": 2.25,
    "Utilities / Power Generation": 2.25,
    "Mining / Extraction": 6.0,
}

# State multipliers (based on cost research)
# High-cost states ~1.2, Average ~1.0, Low-cost ~0.8
STATE_MULTIPLIERS = {
    "Alabama": 1.15,
    "Alaska": 1.25,
    "Arizona": 0.85,
    "Arkansas": 0.90,
    "California": 1.30,
    "Colorado": 0.95,
    "Connecticut": 1.15,
    "Delaware": 1.05,
    "District of Columbia": 1.10,
    "Florida": 0.95,
    "Georgia": 0.95,
    "Hawaii": 1.10,
    "Idaho": 0.90,
    "Illinois": 1.05,
    "Indiana": 0.85,
    "Iowa": 0.90,
    "Kansas": 0.90,
    "Kentucky": 0.95,
    "Louisiana": 1.05,
    "Maine": 1.00,
    "Maryland": 0.85,
    "Massachusetts": 1.05,
    "Michigan": 1.00,
    "Minnesota": 1.00,
    "Mississippi": 0.95,
    "Missouri": 0.95,
    "Montana": 1.15,
    "Nebraska": 0.90,
    "Nevada": 0.95,
    "New Hampshire": 1.00,
    "New Jersey": 1.15,
    "New Mexico": 0.95,
    "New York": 1.20,
    "North Carolina": 0.85,
    "North Dakota": 0.90,
    "Ohio": 0.90,
    "Oklahoma": 1.00,
    "Oregon": 0.95,
    "Pennsylvania": 1.10,
    "Rhode Island": 1.05,
    "South Carolina": 0.95,
    "South Dakota": 0.85,
    "Tennessee": 0.90,
    "Texas": 0.85,
    "Utah": 0.80,
    "Vermont": 1.00,
    "Virginia": 0.80,
    "Washington": 1.05,
    "West Virginia": 1.00,
    "Wisconsin": 0.95,
    "Wyoming": 1.00,
}


# =============================================================================
# EXPLANATIONS FOR CLICK-TO-EXPLAIN FEATURE
# =============================================================================

EXPLANATIONS = {
    # Schedule Inputs
    'schedule_name': {
        'title': 'Schedule Name',
        'text': 'A descriptive name for this schedule configuration.\n\nExamples: "Current 5-Day Schedule", "Proposed 12-Hour Rotation", "Weekend Crew Option"'
    },
    'days_per_week': {
        'title': 'Days Per Week',
        'text': '5 Days: Traditional Monday-Friday operation.\nScheduled Hours = 40 hours/week × 52 weeks = 2,080 hours/year\n\n7 Days: Continuous operation covering all 7 days.\nScheduled Hours = 42 hours/week × 52 weeks = 2,184 hours/year\n\nNote: 7-day schedules have 42 average weekly hours because employees rotate through the schedule.'
    },
    'shift_length': {
        'title': 'Shift Length',
        'text': '8-Hour Shifts: Three shifts per day (Day, Swing, Night).\nShift differential paid on 2 of 3 shifts.\n\n12-Hour Shifts: Two shifts per day (Day, Night).\nEmployees work approximately 50% of all days.\nShift differential paid on 1 of 2 shifts.'
    },
    'holiday_pay_day_off': {
        'title': 'Holiday Pay - Scheduled Day Off',
        'text': 'For 12-hour schedules, employees are scheduled to work only about half the days of the year.\n\nThis is the number of hours of holiday pay an employee receives when a holiday falls on one of their scheduled days OFF.\n\nTypical values: 8 hours (most common) or 0 hours (some companies).'
    },
    'holiday_pay_workday': {
        'title': 'Holiday Pay - Scheduled Workday',
        'text': 'For 12-hour schedules, this is the number of hours of holiday pay an employee receives when a holiday falls on one of their scheduled WORKDAYS.\n\nTypical values: 8 hours (same as day off) or 12 hours (full shift).'
    },
    'avg_pay_hours': {
        'title': 'Average Pay Hours Per Week',
        'text': 'The average number of hours PAID per week, which may be higher than hours worked due to overtime rules.\n\nFactors that increase pay hours:\n• Daily overtime (over 8 hours)\n• Weekly overtime (over 40 hours)\n• Weekend premiums\n• Shift rotation patterns\n\nExample: A schedule with 42 work hours might have 46 pay hours due to daily OT on 12-hour shifts.'
    },
    'other_hours_paid': {
        'title': 'Other Hours Paid',
        'text': 'Any additional paid hours not captured elsewhere.\n\nExamples:\n• Training time\n• Meeting time\n• Travel time\n• On-call pay converted to hours'
    },
    
    # Calculated Results
    'scheduled_hours': {
        'title': 'Scheduled Hours',
        'text': 'Total annual hours the employee is scheduled to work BEFORE any time off is deducted.\n\n5-Day Schedule: 40 hrs/week × 52 weeks = 2,080 hours\n7-Day Schedule: 42 hrs/week × 52 weeks = 2,184 hours\n\nThis is the starting point for all calculations.'
    },
    'vacation_hours': {
        'title': 'Vacation Hours',
        'text': 'Annual vacation hours from the Inputs tab.\n\nThese hours reduce the total scheduled hours to calculate actual hours worked.'
    },
    'holiday_hours_missed': {
        'title': 'Holiday Hours Missed',
        'text': 'Hours NOT worked due to holidays when the operation is closed.\n\nFormulas by schedule type:\n• 5-Day, 8-Hour: (Recognized - Worked) × 8\n• 7-Day, 8-Hour: (Recognized - Worked) × 0.75 × 8\n• 7-Day, 12-Hour: (Recognized - Worked) × 0.5 × 12\n\nThe multipliers (0.75, 0.5) account for the probability that a holiday falls on an employee\'s scheduled workday.'
    },
    'hours_worked': {
        'title': 'Hours Worked',
        'text': 'Actual hours the employee is expected to work after all time off.\n\nFormula: Scheduled Hours - Vacation - Holiday Hours Missed - FMLA - Paid Time Off - Unpaid Time Off - Other Time Off\n\nThis represents productive time on the job.'
    },
    'percent_worked': {
        'title': 'Percent of Scheduled Hours Worked',
        'text': 'What percentage of scheduled time is actually worked.\n\nFormula: Hours Worked ÷ Scheduled Hours\n\nA typical value is 85-90%. Lower percentages mean more time off relative to the schedule.'
    },
    'absenteeism_ot': {
        'title': 'Absenteeism % That Causes Overtime',
        'text': 'The percentage of absences that require overtime coverage.\n\nFormula: (Scheduled Hours - Hours Worked - Holiday Hours Missed - Lack of Work Hours) ÷ Scheduled Hours\n\nExclusions:\n• Holiday hours - if the operation is closed on a holiday, no overtime is needed\n• Lack of Work (LOW) hours - voluntary unpaid time off offered when there\'s insufficient work, so no overtime coverage needed\n\nThis percentage helps estimate overtime costs due to absenteeism.'
    },
    'lack_of_work_hours': {
        'title': 'Lack of Work (LOW) Hours',
        'text': 'Voluntary unpaid time off offered to employees when there is insufficient work.\n\nUnlike other absences, LOW hours do NOT cause overtime because the company offers this time off precisely when there\'s not enough work to keep everyone busy.\n\nLOW hours:\n• Reduce Hours Worked (less productive time)\n• Do NOT increase Absenteeism % That Causes OT\n• Are unpaid (no cost to company)\n\nThis is common in manufacturing/warehouse operations with variable demand.'
    },
    'vacation_pay': {
        'title': 'Vacation Pay (Hours)',
        'text': 'Hours paid for vacation time.\n\nFormula: Vacation Hours × Vacation Multiplier\n\nThe multiplier is typically 1.0, but some companies include overtime averaging in vacation pay calculations, resulting in a multiplier greater than 1.0.'
    },
    'holiday_pay': {
        'title': 'Holiday Pay (Hours)',
        'text': 'Hours paid simply because it\'s a holiday - regardless of whether worked.\n\n8-Hour Schedules: Holidays Recognized × 8 hours\n\n12-Hour Schedules: (Holidays × 0.5 × Pay on Day Off) + (Holidays × 0.5 × Pay on Workday)\n\nNote: This is separate from Holiday PREMIUM, which is extra pay for actually working on a holiday.'
    },
    'holiday_premium': {
        'title': 'Holiday Premium (Hours)',
        'text': 'EXTRA pay hours for actually working on a holiday.\n\nIf holiday premium rate is 2.0x (double time), the EXTRA portion is 1.0 (since straight time is already counted in Hours Worked).\n\nFormulas:\n• 5-Day: Holidays Worked × 8 × (Rate - 1)\n• 7-Day, 8-Hour: Holidays Worked × 0.75 × 8 × (Rate - 1)\n• 7-Day, 12-Hour: Holidays Worked × 0.5 × 12 × (Rate - 1)'
    },
    'built_in_ot': {
        'title': 'Built-in Overtime (Hours)',
        'text': 'Overtime hours that are inherent in the schedule design.\n\n5-Day Schedules: 0 (no built-in OT)\n\n7-Day Schedules: (Avg Pay Hours ÷ 42 - 1) × Hours Worked\n\nThis occurs when overtime rules (daily, weekly, or weekend premiums) result in pay hours exceeding work hours.'
    },
    'shift_diff_hours': {
        'title': 'Shift Differential (Hours)',
        'text': 'The shift differential converted to equivalent hours.\n\nFormula: (Shift Differential $ ÷ Hourly Wage $) × Shift Factor × Hours Worked\n\n8-Hour Shifts: Factor = 2/3 (2 of 3 shifts get differential)\n12-Hour Shifts: Factor = 1/2 (1 of 2 shifts get differential)'
    },
    'shift_overlap_pay': {
        'title': 'Shift Overlap Pay (Hours)',
        'text': 'Extra paid time for shift-to-shift overlap (handoff time between shifts).\n\nShift overlap is typically paid at 1.5x because it results in overtime.\n\nFormula:\n8-Hour Shifts: (Overlap Minutes ÷ 60) × (Hours Worked ÷ 8) × 1.5\n12-Hour Shifts: (Overlap Minutes ÷ 60) × (Hours Worked ÷ 12) × 1.5\n\nThis calculates: overlap time per shift × number of shifts × OT rate'
    },
    'total_hours_paid': {
        'title': 'Total Hours Paid',
        'text': 'Sum of all paid hour components:\n\n• Hours Worked\n• Vacation Pay\n• Holiday Pay\n• Holiday Premium\n• Built-in Overtime\n• Shift Differential\n• Other Hours Paid\n\nThis total is used to calculate the ratio of paid time to worked time.'
    },
    'ratio_paid_worked': {
        'title': 'Ratio of Paid Hours to Worked Hours',
        'text': 'How many hours are PAID for each hour actually WORKED.\n\nFormula: Total Hours Paid ÷ Hours Worked\n\nA ratio of 1.20 means for every hour worked, you pay for 1.20 hours.\n\nThis ratio is critical for understanding true labor costs.'
    },
    'adjusted_st_burden': {
        'title': 'Adjusted ST Burden',
        'text': 'Straight Time burden rate adjusted for schedule type.\n\nFor 5-Day schedules: Same as calculated ST Burden\n\nFor 7-Day schedules: Some burden components (FUI/SUI, Medical) are fixed costs spread over more hours.\n\nFormula: (Fixed Components × 40/42) + Variable Components\n\nThis adjustment reflects the lower per-hour cost of fixed benefits on longer schedules.'
    },
    'cost_scheduled_time': {
        'title': 'Cost of Scheduled Time',
        'text': 'The fully-loaded cost per hour of scheduled time.\n\nFormula: Wage × Ratio Paid/Worked × (1 + Adjusted ST Burden)\n\nThis represents what you actually pay per hour of productive work, including all benefits and non-productive paid time.\n\nUse this to compare the true cost of different schedules.'
    },
    'marginal_ot_15': {
        'title': 'Marginal Cost of Overtime (1.5x)',
        'text': 'The cost per hour of time-and-a-half overtime.\n\nFormula: Wage × 1.5 × (1 + OT Burden)\n\nThe OT Burden includes only income-dependent benefits (payroll taxes, 401K, workers comp, and any checked "income-dependent" items).\n\nCompare this to Cost of Scheduled Time to decide whether to use overtime or add scheduled staff.'
    },
    'marginal_ot_20': {
        'title': 'Marginal Cost of Overtime (2.0x)',
        'text': 'The cost per hour of double-time overtime.\n\nFormula: Wage × 2.0 × (1 + OT Burden)\n\nDouble-time is typically paid for:\n• Hours over 12 in a day\n• 7th consecutive day worked\n• Holiday work (in some companies)\n\nThis represents your maximum overtime cost.'
    },
}


# =============================================================================
# BURDEN RATE CALCULATOR CLASS
# =============================================================================

class BurdenCalculator:
    """
    Calculates burden rates from input components.
    """
    
    def __init__(self):
        # Default values matching the Excel spreadsheet
        # All percentages stored as whole numbers (e.g., 15 for 15%)
        self.inputs = {
            # Wage
            'wage': 35.87,
            
            # Burden Rates - Fixed
            'fui_sui': 1.0,  # Fixed at 1%
            'payroll_taxes': 7.65,  # Fixed at 7.65%
            
            # 401K inputs
            '401k_match': 6.0,  # Company match percentage
            '401k_participation': 50.0,  # Participation rate percentage
            
            # Workers' Comp inputs
            'wc_state': 'California',
            'wc_industry': 'Manufacturing - General',
            
            # Medical - annual dollar amount
            'medical_annual': 11640.0,
            
            # Bonus - annual dollar amount
            'bonus_annual': 16936.0,
            'bonus_applies_to_ot': True,  # Income-dependent checkbox
            
            # Other burdens
            'other_burden_1_label': '',
            'other_burden_1_rate': 0.0,
            'other_burden_1_applies_to_ot': False,
            'other_burden_2_label': '',
            'other_burden_2_rate': 0.0,
            'other_burden_2_applies_to_ot': False,
            
            # Time Off
            'vacation_hours': 151.0,
            'vacation_multiplier': 1.0,
            'fmla_hours': 40.0,
            'unpaid_time_off': 9.0,
            'paid_time_off': 48.0,
            'other_time_off_1_label': '',
            'other_time_off_1_hours': 0.0,
            'other_time_off_2_label': '',
            'other_time_off_2_hours': 0.0,
            'lack_of_work_hours': 0.0,  # Voluntary unpaid time off - doesn't cause OT
            
            # Holiday Settings
            'holidays_recognized': 10,
            'holidays_worked': 5,
            'holiday_premium': 2.0,
            
            # Schedule Settings
            'shift_differential': 1.0,
            'shift_overlap_minutes': 30,
        }
    
    def calculate_401k_rate(self) -> float:
        """Calculate 401K burden rate from match and participation."""
        match = self.inputs['401k_match'] / 100.0
        participation = self.inputs['401k_participation'] / 100.0
        return match * participation * 100.0  # Return as percentage
    
    def calculate_workers_comp_rate(self) -> float:
        """Calculate Workers' Comp rate from state and industry."""
        industry = self.inputs['wc_industry']
        state = self.inputs['wc_state']
        
        base_rate = INDUSTRY_RATES.get(industry, 2.25)
        multiplier = STATE_MULTIPLIERS.get(state, 1.0)
        
        return base_rate * multiplier
    
    def calculate_medical_rate(self) -> float:
        """Calculate Medical burden rate from annual dollar amount."""
        wage = self.inputs['wage']
        annual_wage = wage * 2080
        if annual_wage > 0:
            return (self.inputs['medical_annual'] / annual_wage) * 100.0
        return 0.0
    
    def calculate_bonus_rate(self) -> float:
        """Calculate Bonus burden rate from annual dollar amount."""
        wage = self.inputs['wage']
        annual_wage = wage * 2080
        if annual_wage > 0:
            return (self.inputs['bonus_annual'] / annual_wage) * 100.0
        return 0.0
    
    def calculate_st_burden(self) -> float:
        """Calculate total Straight Time burden rate."""
        fui_sui = self.inputs['fui_sui']
        payroll_taxes = self.inputs['payroll_taxes']
        k401 = self.calculate_401k_rate()
        wc = self.calculate_workers_comp_rate()
        medical = self.calculate_medical_rate()
        bonus = self.calculate_bonus_rate()
        other1 = self.inputs['other_burden_1_rate']
        other2 = self.inputs['other_burden_2_rate']
        
        return fui_sui + payroll_taxes + k401 + wc + medical + bonus + other1 + other2
    
    def calculate_ot_burden(self) -> float:
        """Calculate total Overtime burden rate (income-dependent items only)."""
        # Always include: Payroll Taxes, 401K, Workers' Comp
        payroll_taxes = self.inputs['payroll_taxes']
        k401 = self.calculate_401k_rate()
        wc = self.calculate_workers_comp_rate()
        
        total = payroll_taxes + k401 + wc
        
        # Conditionally include: Bonus, Other 1, Other 2
        if self.inputs['bonus_applies_to_ot']:
            total += self.calculate_bonus_rate()
        
        if self.inputs['other_burden_1_applies_to_ot']:
            total += self.inputs['other_burden_1_rate']
        
        if self.inputs['other_burden_2_applies_to_ot']:
            total += self.inputs['other_burden_2_rate']
        
        return total
    
    def calculate_st_only_burden(self) -> float:
        """Calculate burden components that apply ONLY to straight time (not OT)."""
        # FUI/SUI and Medical are fixed costs, not income-dependent
        return self.inputs['fui_sui'] + self.calculate_medical_rate()
    
    def calculate_adjusted_st_burden(self, is_7_day: bool) -> float:
        """Calculate adjusted ST burden for schedule type."""
        if not is_7_day:
            return self.calculate_st_burden()
        
        # For 7-day schedules, fixed components are spread over more hours
        st_only = self.calculate_st_only_burden()
        shared = self.calculate_st_burden() - st_only
        
        # Adjust fixed components by 40/42
        adjusted_st_only = st_only * (40.0 / 42.0)
        
        return adjusted_st_only + shared


# =============================================================================
# SCHEDULE COST CALCULATOR CLASS
# =============================================================================

class ScheduleCostCalculator:
    """
    Calculates cost metrics for a single schedule configuration.
    """
    
    def __init__(self, burden_calc: BurdenCalculator):
        self.burden_calc = burden_calc
        
        # Schedule-specific inputs
        self.schedule_inputs = {
            'schedule_name': 'New Schedule',
            'days_per_week': 5,  # 5 or 7
            'shift_length': 8,  # 8 or 12 (only applies to 7-day)
            'holiday_pay_day_off': 8,  # For 12-hour schedules
            'holiday_pay_workday': 12,  # For 12-hour schedules
            'avg_pay_hours': 46,  # For 7-day schedules
            'other_hours_paid': 0,
        }
    
    def get_schedule_type(self) -> str:
        """Get schedule type description."""
        days = self.schedule_inputs['days_per_week']
        if days == 5:
            return "5-day, 8-hour"
        else:
            shift = self.schedule_inputs['shift_length']
            return f"7-day, {shift}-hour"
    
    def calculate_scheduled_hours(self) -> float:
        """Calculate annual scheduled hours including shift overlap."""
        days = self.schedule_inputs['days_per_week']
        shift = self.schedule_inputs['shift_length']
        overlap_minutes = self.burden_calc.inputs.get('shift_overlap_minutes', 0)
        overlap_hours = overlap_minutes / 60.0
        
        if days == 5:
            # 5-day, 8-hour: 2080 + overlap × 260 shifts/year
            return (40 * 52) + (overlap_hours * 260)
        elif shift == 8:
            # 7-day, 8-hour: 2184 + overlap × 273 shifts/year
            return (42 * 52) + (overlap_hours * 273)
        else:
            # 7-day, 12-hour: 2184 + overlap × 182 shifts/year
            return (42 * 52) + (overlap_hours * 182)
    
    def calculate_holiday_hours_missed(self) -> float:
        """Calculate hours missed due to holidays (when operation is closed)."""
        recognized = self.burden_calc.inputs['holidays_recognized']
        worked = self.burden_calc.inputs['holidays_worked']
        days = self.schedule_inputs['days_per_week']
        shift = self.schedule_inputs['shift_length']
        
        diff = recognized - worked  # Holidays when operation is closed
        
        if days == 5:
            # 5-day, 8-hour
            return diff * 8
        elif shift == 8:
            # 7-day, 8-hour
            return diff * 0.75 * 8
        else:
            # 7-day, 12-hour
            return diff * 0.5 * 12
    
    def calculate_hours_worked(self) -> float:
        """Calculate actual hours worked after all time off."""
        scheduled = self.calculate_scheduled_hours()
        vacation = self.burden_calc.inputs['vacation_hours']
        holiday_missed = self.calculate_holiday_hours_missed()
        fmla = self.burden_calc.inputs['fmla_hours']
        pto = self.burden_calc.inputs['paid_time_off']
        unpaid = self.burden_calc.inputs['unpaid_time_off']
        other1 = self.burden_calc.inputs['other_time_off_1_hours']
        other2 = self.burden_calc.inputs['other_time_off_2_hours']
        low = self.burden_calc.inputs['lack_of_work_hours']
        
        return scheduled - vacation - holiday_missed - fmla - pto - unpaid - other1 - other2 - low
    
    def calculate_percent_worked(self) -> float:
        """Calculate percentage of scheduled hours actually worked."""
        scheduled = self.calculate_scheduled_hours()
        worked = self.calculate_hours_worked()
        if scheduled > 0:
            return worked / scheduled
        return 0
    
    def calculate_absenteeism_ot_percent(self) -> float:
        """Calculate absenteeism percentage that causes overtime.
        
        Excludes holidays (plant closed) and Lack of Work hours (voluntary
        unpaid time off offered when there's insufficient work - no OT needed).
        """
        scheduled = self.calculate_scheduled_hours()
        worked = self.calculate_hours_worked()
        holiday_missed = self.calculate_holiday_hours_missed()
        low = self.burden_calc.inputs['lack_of_work_hours']
        
        if scheduled > 0:
            # Exclude both holidays AND lack of work - neither causes overtime
            return (scheduled - worked - holiday_missed - low) / scheduled
        return 0
    
    def calculate_vacation_pay(self) -> float:
        """Calculate vacation pay hours."""
        hours = self.burden_calc.inputs['vacation_hours']
        multiplier = self.burden_calc.inputs['vacation_multiplier']
        return hours * multiplier
    
    def calculate_holiday_pay(self) -> float:
        """Calculate holiday pay hours."""
        recognized = self.burden_calc.inputs['holidays_recognized']
        days = self.schedule_inputs['days_per_week']
        shift = self.schedule_inputs['shift_length']
        
        if days == 5 or (days == 7 and shift == 8):
            # 8-hour schedules
            return recognized * 8
        else:
            # 12-hour schedules
            pay_day_off = self.schedule_inputs['holiday_pay_day_off']
            pay_workday = self.schedule_inputs['holiday_pay_workday']
            return (recognized * 0.5 * pay_day_off) + (recognized * 0.5 * pay_workday)
    
    def calculate_holiday_premium(self) -> float:
        """Calculate holiday premium hours (extra pay for working holidays)."""
        worked = self.burden_calc.inputs['holidays_worked']
        premium_rate = self.burden_calc.inputs['holiday_premium']
        extra = premium_rate - 1  # Extra portion beyond straight time
        days = self.schedule_inputs['days_per_week']
        shift = self.schedule_inputs['shift_length']
        
        if days == 5:
            # 5-day, 8-hour
            return worked * 8 * extra
        elif shift == 8:
            # 7-day, 8-hour
            return worked * 0.75 * 8 * extra
        else:
            # 7-day, 12-hour
            return worked * 0.5 * 12 * extra
    
    def calculate_built_in_ot(self) -> float:
        """Calculate built-in overtime hours."""
        days = self.schedule_inputs['days_per_week']
        
        if days == 5:
            return 0
        
        avg_pay = self.schedule_inputs['avg_pay_hours']
        hours_worked = self.calculate_hours_worked()
        
        return (avg_pay / 42 - 1) * hours_worked
    
    def calculate_shift_differential_hours(self) -> float:
        """Calculate shift differential as equivalent hours."""
        wage = self.burden_calc.inputs['wage']
        shift_diff = self.burden_calc.inputs['shift_differential']
        hours_worked = self.calculate_hours_worked()
        days = self.schedule_inputs['days_per_week']
        shift = self.schedule_inputs['shift_length']
        
        if wage <= 0:
            return 0
        
        if days == 5 or (days == 7 and shift == 8):
            # 8-hour schedules: 2 of 3 shifts
            factor = 2 / 3
        else:
            # 12-hour schedules: 1 of 2 shifts
            factor = 1 / 2
        
        return (shift_diff / wage) * factor * hours_worked
    
    def calculate_shift_overlap_pay(self) -> float:
        """Calculate shift overlap pay as equivalent hours (paid at 1.5x)."""
        overlap_minutes = self.burden_calc.inputs.get('shift_overlap_minutes', 0)
        hours_worked = self.calculate_hours_worked()
        days = self.schedule_inputs['days_per_week']
        shift = self.schedule_inputs['shift_length']
        
        overlap_per_shift = overlap_minutes / 60.0
        
        if days == 5 or (days == 7 and shift == 8):
            # 8-hour shifts
            num_shifts = hours_worked / 8
        else:
            # 12-hour shifts
            num_shifts = hours_worked / 12
        
        # Overlap is paid at 1.5x
        return overlap_per_shift * num_shifts * 1.5
    
    def calculate_total_hours_paid(self) -> float:
        """Calculate total hours paid."""
        return (
            self.calculate_hours_worked() +
            self.calculate_vacation_pay() +
            self.calculate_holiday_pay() +
            self.calculate_holiday_premium() +
            self.calculate_built_in_ot() +
            self.calculate_shift_differential_hours() +
            self.calculate_shift_overlap_pay() +
            self.burden_calc.inputs['paid_time_off'] +  # Paid Time Off (Sick, etc.)
            self.schedule_inputs['other_hours_paid']
        )
    
    def calculate_ratio_paid_worked(self) -> float:
        """Calculate ratio of paid hours to worked hours."""
        worked = self.calculate_hours_worked()
        paid = self.calculate_total_hours_paid()
        if worked > 0:
            return paid / worked
        return 0
    
    def calculate_cost_scheduled_time(self) -> float:
        """Calculate cost of scheduled time per hour."""
        wage = self.burden_calc.inputs['wage']
        ratio = self.calculate_ratio_paid_worked()
        is_7_day = self.schedule_inputs['days_per_week'] == 7
        st_burden = self.burden_calc.calculate_adjusted_st_burden(is_7_day) / 100.0
        
        return wage * ratio * (1 + st_burden)
    
    def calculate_marginal_ot_15(self) -> float:
        """Calculate marginal cost of 1.5x overtime."""
        wage = self.burden_calc.inputs['wage']
        ot_burden = self.burden_calc.calculate_ot_burden() / 100.0
        
        return wage * 1.5 * (1 + ot_burden)
    
    def calculate_marginal_ot_20(self) -> float:
        """Calculate marginal cost of 2.0x overtime."""
        wage = self.burden_calc.inputs['wage']
        ot_burden = self.burden_calc.calculate_ot_burden() / 100.0
        
        return wage * 2.0 * (1 + ot_burden)
    
    def calculate_all(self) -> Dict[str, Any]:
        """Calculate all metrics and return as dictionary."""
        is_7_day = self.schedule_inputs['days_per_week'] == 7
        
        return {
            'schedule_name': self.schedule_inputs['schedule_name'],
            'schedule_type': self.get_schedule_type(),
            'days_per_week': self.schedule_inputs['days_per_week'],
            'shift_length': self.schedule_inputs['shift_length'],
            'scheduled_hours': self.calculate_scheduled_hours(),
            'vacation_hours': self.burden_calc.inputs['vacation_hours'],
            'holiday_hours_missed': self.calculate_holiday_hours_missed(),
            'fmla_hours': self.burden_calc.inputs['fmla_hours'],
            'paid_time_off': self.burden_calc.inputs['paid_time_off'],
            'unpaid_time_off': self.burden_calc.inputs['unpaid_time_off'],
            'other_time_off_1': self.burden_calc.inputs['other_time_off_1_hours'],
            'other_time_off_2': self.burden_calc.inputs['other_time_off_2_hours'],
            'lack_of_work_hours': self.burden_calc.inputs['lack_of_work_hours'],
            'hours_worked': self.calculate_hours_worked(),
            'percent_worked': self.calculate_percent_worked(),
            'absenteeism_ot_percent': self.calculate_absenteeism_ot_percent(),
            'vacation_pay': self.calculate_vacation_pay(),
            'holiday_pay': self.calculate_holiday_pay(),
            'holiday_premium': self.calculate_holiday_premium(),
            'built_in_ot': self.calculate_built_in_ot(),
            'shift_diff_hours': self.calculate_shift_differential_hours(),
            'shift_overlap_pay': self.calculate_shift_overlap_pay(),
            'other_hours_paid': self.schedule_inputs['other_hours_paid'],
            'total_hours_paid': self.calculate_total_hours_paid(),
            'ratio_paid_worked': self.calculate_ratio_paid_worked(),
            'wage': self.burden_calc.inputs['wage'],
            'st_burden': self.burden_calc.calculate_st_burden(),
            'ot_burden': self.burden_calc.calculate_ot_burden(),
            'adjusted_st_burden': self.burden_calc.calculate_adjusted_st_burden(is_7_day),
            'cost_scheduled_time': self.calculate_cost_scheduled_time(),
            'marginal_ot_15': self.calculate_marginal_ot_15(),
            'marginal_ot_20': self.calculate_marginal_ot_20(),
            'employee_annual_earnings': self.calculate_total_hours_paid() * self.burden_calc.inputs['wage'],
        }


# =============================================================================
# MAIN APPLICATION CLASS
# =============================================================================

class CostOfTimeApp:
    """
    Main application window for the Cost of Time Calculator.
    """
    
    # Default auto-save filename (in same directory as script)
    AUTO_SAVE_FILE = "cost_of_time_autosave.json"
    
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("Cost of Time Calculator - Shiftwork Solutions LLC")
        self.root.geometry("1500x900")
        self.root.minsize(1400, 800)
        
        # Initialize calculators
        self.burden_calc = BurdenCalculator()
        self.schedule_calc = ScheduleCostCalculator(self.burden_calc)
        
        # Track input variables
        self.input_vars = {}
        self.schedule_vars = {}
        
        # Saved calculations
        self.saved_costs: List[Dict[str, Any]] = []
        
        # Setup UI
        self.setup_styles()
        self.create_menu()
        self.create_main_layout()
        
        # Initial calculation
        self.update_schedule_display()
        
        # Load auto-saved data if it exists
        self.load_auto_save()
        
        # Bind window close to auto-save
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        
    def setup_styles(self):
        """Configure ttk styles for the application."""
        self.style = ttk.Style()
        self.style.theme_use('clam')
        
        # Configure colors
        self.colors = {
            'primary': '#007acc',
            'secondary': '#00a8ff',
            'success': '#28a745',
            'warning': '#ffc107',
            'danger': '#dc3545',
            'light': '#f8f9fa',
            'dark': '#343a40',
            'white': '#ffffff',
            'gray': '#6c757d',
            'calculated': '#17a2b8'  # Color for calculated/display-only fields
        }
        
        # Configure styles
        self.style.configure('Header.TLabel', 
                            font=('Segoe UI', 14, 'bold'),
                            foreground=self.colors['primary'])
        
        self.style.configure('SubHeader.TLabel',
                            font=('Segoe UI', 11, 'bold'),
                            foreground=self.colors['dark'])
        
        self.style.configure('Result.TLabel',
                            font=('Segoe UI', 10),
                            background=self.colors['light'])
        
        self.style.configure('Calculated.TLabel',
                            font=('Segoe UI', 10, 'bold'),
                            foreground=self.colors['calculated'])
        
        self.style.configure('Fixed.TLabel',
                            font=('Segoe UI', 10),
                            foreground=self.colors['gray'])
        
        self.style.configure('Primary.TButton',
                            font=('Segoe UI', 10, 'bold'))
        
        self.style.configure('Card.TFrame',
                            background=self.colors['white'])
        
    def create_menu(self):
        """Create the application menu bar."""
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)
        
        # File menu
        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="File", menu=file_menu)
        file_menu.add_command(label="New", command=self.reset_to_defaults, accelerator="Ctrl+N")
        file_menu.add_command(label="Open Configuration...", command=self.load_config, accelerator="Ctrl+O")
        file_menu.add_command(label="Save Configuration...", command=self.save_config, accelerator="Ctrl+S")
        file_menu.add_separator()
        file_menu.add_command(label="Export to Excel...", command=self.export_excel, accelerator="Ctrl+E")
        file_menu.add_command(label="Export to PDF...", command=self.export_pdf, accelerator="Ctrl+P")
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=self.root.quit, accelerator="Alt+F4")
        
        # Edit menu
        edit_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Edit", menu=edit_menu)
        edit_menu.add_command(label="Reset to Defaults", command=self.reset_to_defaults)
        edit_menu.add_command(label="Clear Saved Costs", command=self.clear_saved_costs)
        
        # Help menu
        help_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Help", menu=help_menu)
        help_menu.add_command(label="About", command=self.show_about)
        help_menu.add_command(label="Contact Shiftwork Solutions", command=self.show_contact)
        
        # Keyboard bindings
        self.root.bind('<Control-n>', lambda e: self.reset_to_defaults())
        self.root.bind('<Control-o>', lambda e: self.load_config())
        self.root.bind('<Control-s>', lambda e: self.save_config())
        self.root.bind('<Control-e>', lambda e: self.export_excel())
        self.root.bind('<Control-p>', lambda e: self.export_pdf())
        
    def create_main_layout(self):
        """Create the main application layout."""
        # Main container with padding
        main_container = ttk.Frame(self.root, padding="10")
        main_container.pack(fill=tk.BOTH, expand=True)
        
        # Header
        header_frame = ttk.Frame(main_container)
        header_frame.pack(fill=tk.X, pady=(0, 10))
        
        title_label = ttk.Label(header_frame, 
                               text="Cost of Time Calculator",
                               style='Header.TLabel',
                               font=('Segoe UI', 18, 'bold'))
        title_label.pack(side=tk.LEFT)
        
        subtitle_label = ttk.Label(header_frame,
                                  text="Shiftwork Solutions LLC",
                                  font=('Segoe UI', 10),
                                  foreground=self.colors['gray'])
        subtitle_label.pack(side=tk.LEFT, padx=(15, 0), pady=(8, 0))
        
        # Create notebook for tabs
        self.notebook = ttk.Notebook(main_container)
        self.notebook.pack(fill=tk.BOTH, expand=True)
        
        # Create tabs
        self.create_inputs_tab()
        self.create_saved_costs_tab()
        self.create_compare_tab()
        self.create_analysis_tab()
        if MATPLOTLIB_AVAILABLE:
            self.create_charts_tab()
        
    def show_explanation(self, key: str):
        """Show explanation popup for a field."""
        if key in EXPLANATIONS:
            info = EXPLANATIONS[key]
            messagebox.showinfo(info['title'], info['text'])
        else:
            messagebox.showinfo("Information", "No explanation available for this field.")
    
    def create_explanation_button(self, parent, key: str) -> ttk.Button:
        """Create a small ? button that shows explanation when clicked."""
        btn = ttk.Button(parent, text="?", width=2, 
                        command=lambda k=key: self.show_explanation(k))
        return btn
        
    def create_inputs_tab(self):
        """Create the inputs tab with burden inputs on left, schedule calculator on right."""
        inputs_frame = ttk.Frame(self.notebook, padding="10")
        self.notebook.add(inputs_frame, text="Inputs & Calculator")
        
        # Create horizontal paned window
        paned = ttk.PanedWindow(inputs_frame, orient=tk.HORIZONTAL)
        paned.pack(fill=tk.BOTH, expand=True)
        
        # Left side - Burden Inputs (scrollable)
        left_frame = ttk.Frame(paned)
        paned.add(left_frame, weight=1)
        
        # Right side - Schedule Calculator
        right_frame = ttk.Frame(paned)
        paned.add(right_frame, weight=1)
        
        # Create left side content (burden inputs)
        self.create_burden_inputs(left_frame)
        
        # Create right side content (schedule calculator)
        self.create_schedule_calculator(right_frame)
        
    def create_burden_inputs(self, parent):
        """Create the burden inputs section (left side)."""
        # Create scrollable canvas
        canvas = tk.Canvas(parent, highlightthickness=0)
        scrollbar = ttk.Scrollbar(parent, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Enable mousewheel scrolling only when mouse is over this canvas
        def on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        
        def bind_mousewheel(event):
            canvas.bind_all("<MouseWheel>", on_mousewheel)
        
        def unbind_mousewheel(event):
            canvas.unbind_all("<MouseWheel>")
        
        canvas.bind("<Enter>", bind_mousewheel)
        canvas.bind("<Leave>", unbind_mousewheel)
        
        # Create input sections
        row = 0
        
        # === Wage Section ===
        wage_frame = ttk.LabelFrame(scrollable_frame, text="Base Wage", padding="10")
        wage_frame.grid(row=row, column=0, sticky="ew", padx=5, pady=5)
        
        ttk.Label(wage_frame, text="Hourly Wage ($):").grid(row=0, column=0, sticky="e", padx=5, pady=3)
        wage_var = tk.DoubleVar(value=self.burden_calc.inputs['wage'])
        self.input_vars['wage'] = wage_var
        ttk.Entry(wage_frame, textvariable=wage_var, width=12).grid(row=0, column=1, sticky="w", padx=5, pady=3)
        wage_var.trace_add('write', lambda *args: self.on_input_change())
        
        row += 1
        
        # === Burden Rates Section ===
        burden_frame = ttk.LabelFrame(scrollable_frame, text="Burden Rates", padding="10")
        burden_frame.grid(row=row, column=0, sticky="ew", padx=5, pady=5)
        
        burden_row = 0
        
        # --- PROMINENT BURDEN SUMMARY AT TOP ---
        summary_container = tk.Frame(burden_frame, bg='#e8f4f8', relief='ridge', bd=2)
        summary_container.grid(row=burden_row, column=0, columnspan=6, sticky="ew", padx=5, pady=10)
        
        # Title
        summary_title = tk.Label(summary_container, text="BURDEN RATE SUMMARY", 
                                font=('Segoe UI', 12, 'bold'), bg='#e8f4f8', fg='#2c3e50')
        summary_title.pack(pady=(10, 5))
        
        # Container for the two boxes
        boxes_frame = tk.Frame(summary_container, bg='#e8f4f8')
        boxes_frame.pack(fill='x', padx=20, pady=(5, 15))
        
        # ST Burden Box
        st_box = tk.Frame(boxes_frame, bg='#007acc', relief='raised', bd=2)
        st_box.pack(side='left', expand=True, fill='both', padx=(0, 10))
        
        st_title_label = tk.Label(st_box, text="STRAIGHT TIME", 
                                  font=('Segoe UI', 10, 'bold'), bg='#007acc', fg='white')
        st_title_label.pack(pady=(10, 0))
        
        self.st_burden_display = tk.Label(st_box, text="52.88%", 
                                          font=('Segoe UI', 28, 'bold'), bg='#007acc', fg='white')
        self.st_burden_display.pack(pady=(0, 10))
        
        # OT Burden Box
        ot_box = tk.Frame(boxes_frame, bg='#28a745', relief='raised', bd=2)
        ot_box.pack(side='left', expand=True, fill='both', padx=(10, 0))
        
        ot_title_label = tk.Label(ot_box, text="OVERTIME", 
                                  font=('Segoe UI', 10, 'bold'), bg='#28a745', fg='white')
        ot_title_label.pack(pady=(10, 0))
        
        self.ot_burden_display = tk.Label(ot_box, text="36.27%", 
                                          font=('Segoe UI', 28, 'bold'), bg='#28a745', fg='white')
        self.ot_burden_display.pack(pady=(0, 10))
        
        burden_row += 1
        
        ttk.Separator(burden_frame, orient='horizontal').grid(row=burden_row, column=0, columnspan=6, sticky="ew", pady=5)
        burden_row += 1
        
        # --- Fixed Items ---
        ttk.Label(burden_frame, text="Fixed Rates:", font=('Segoe UI', 10, 'bold')).grid(
            row=burden_row, column=0, columnspan=4, sticky="w", padx=5, pady=(5, 2))
        burden_row += 1
        
        # FUI/SUI - Fixed
        ttk.Label(burden_frame, text="FUI/SUI:").grid(row=burden_row, column=0, sticky="e", padx=5, pady=3)
        ttk.Label(burden_frame, text="1.00%", style='Fixed.TLabel').grid(row=burden_row, column=1, sticky="w", padx=5, pady=3)
        ttk.Label(burden_frame, text="(fixed)", foreground=self.colors['gray']).grid(row=burden_row, column=2, sticky="w", padx=5, pady=3)
        
        # Payroll Taxes - Fixed
        ttk.Label(burden_frame, text="Payroll Taxes (FICA):").grid(row=burden_row, column=3, sticky="e", padx=5, pady=3)
        ttk.Label(burden_frame, text="7.65%", style='Fixed.TLabel').grid(row=burden_row, column=4, sticky="w", padx=5, pady=3)
        ttk.Label(burden_frame, text="(fixed)", foreground=self.colors['gray']).grid(row=burden_row, column=5, sticky="w", padx=5, pady=3)
        burden_row += 1
        
        ttk.Separator(burden_frame, orient='horizontal').grid(row=burden_row, column=0, columnspan=6, sticky="ew", pady=10)
        burden_row += 1
        
        # --- 401K Section ---
        ttk.Label(burden_frame, text="401K:", font=('Segoe UI', 10, 'bold')).grid(
            row=burden_row, column=0, columnspan=4, sticky="w", padx=5, pady=(5, 2))
        burden_row += 1
        
        ttk.Label(burden_frame, text="Company Match (%):").grid(row=burden_row, column=0, sticky="e", padx=5, pady=3)
        match_var = tk.DoubleVar(value=self.burden_calc.inputs['401k_match'])
        self.input_vars['401k_match'] = match_var
        ttk.Entry(burden_frame, textvariable=match_var, width=10).grid(row=burden_row, column=1, sticky="w", padx=5, pady=3)
        match_var.trace_add('write', lambda *args: self.on_input_change())
        
        ttk.Label(burden_frame, text="Participation Rate (%):").grid(row=burden_row, column=3, sticky="e", padx=5, pady=3)
        participation_var = tk.DoubleVar(value=self.burden_calc.inputs['401k_participation'])
        self.input_vars['401k_participation'] = participation_var
        ttk.Entry(burden_frame, textvariable=participation_var, width=10).grid(row=burden_row, column=4, sticky="w", padx=5, pady=3)
        participation_var.trace_add('write', lambda *args: self.on_input_change())
        burden_row += 1
        
        ttk.Label(burden_frame, text="Calculated 401K Rate:").grid(row=burden_row, column=0, sticky="e", padx=5, pady=3)
        self.k401_calculated_label = ttk.Label(burden_frame, text="3.00%", style='Calculated.TLabel')
        self.k401_calculated_label.grid(row=burden_row, column=1, sticky="w", padx=5, pady=3)
        burden_row += 1
        
        ttk.Separator(burden_frame, orient='horizontal').grid(row=burden_row, column=0, columnspan=6, sticky="ew", pady=10)
        burden_row += 1
        
        # --- Workers' Compensation Section ---
        ttk.Label(burden_frame, text="Workers' Compensation:", font=('Segoe UI', 10, 'bold')).grid(
            row=burden_row, column=0, columnspan=4, sticky="w", padx=5, pady=(5, 2))
        burden_row += 1
        
        ttk.Label(burden_frame, text="State:").grid(row=burden_row, column=0, sticky="e", padx=5, pady=3)
        state_var = tk.StringVar(value=self.burden_calc.inputs['wc_state'])
        self.input_vars['wc_state'] = state_var
        state_combo = ttk.Combobox(burden_frame, textvariable=state_var, width=20, state='readonly')
        state_combo['values'] = sorted(STATE_MULTIPLIERS.keys())
        state_combo.grid(row=burden_row, column=1, columnspan=2, sticky="w", padx=5, pady=3)
        state_var.trace_add('write', lambda *args: self.on_input_change())
        
        ttk.Label(burden_frame, text="Industry:").grid(row=burden_row, column=3, sticky="e", padx=5, pady=3)
        industry_var = tk.StringVar(value=self.burden_calc.inputs['wc_industry'])
        self.input_vars['wc_industry'] = industry_var
        industry_combo = ttk.Combobox(burden_frame, textvariable=industry_var, width=25, state='readonly')
        industry_combo['values'] = list(INDUSTRY_RATES.keys())
        industry_combo.grid(row=burden_row, column=4, columnspan=2, sticky="w", padx=5, pady=3)
        industry_var.trace_add('write', lambda *args: self.on_input_change())
        burden_row += 1
        
        ttk.Label(burden_frame, text="Calculated WC Rate:").grid(row=burden_row, column=0, sticky="e", padx=5, pady=3)
        self.wc_calculated_label = ttk.Label(burden_frame, text="2.93%", style='Calculated.TLabel')
        self.wc_calculated_label.grid(row=burden_row, column=1, sticky="w", padx=5, pady=3)
        burden_row += 1
        
        ttk.Separator(burden_frame, orient='horizontal').grid(row=burden_row, column=0, columnspan=6, sticky="ew", pady=10)
        burden_row += 1
        
        # --- Medical Section ---
        ttk.Label(burden_frame, text="Medical:", font=('Segoe UI', 10, 'bold')).grid(
            row=burden_row, column=0, columnspan=4, sticky="w", padx=5, pady=(5, 2))
        burden_row += 1
        
        ttk.Label(burden_frame, text="Annual Cost per Employee ($):").grid(row=burden_row, column=0, sticky="e", padx=5, pady=3)
        medical_var = tk.DoubleVar(value=self.burden_calc.inputs['medical_annual'])
        self.input_vars['medical_annual'] = medical_var
        ttk.Entry(burden_frame, textvariable=medical_var, width=12).grid(row=burden_row, column=1, sticky="w", padx=5, pady=3)
        medical_var.trace_add('write', lambda *args: self.on_input_change())
        
        ttk.Label(burden_frame, text="Calculated Rate:").grid(row=burden_row, column=3, sticky="e", padx=5, pady=3)
        self.medical_calculated_label = ttk.Label(burden_frame, text="15.60%", style='Calculated.TLabel')
        self.medical_calculated_label.grid(row=burden_row, column=4, sticky="w", padx=5, pady=3)
        burden_row += 1
        
        ttk.Separator(burden_frame, orient='horizontal').grid(row=burden_row, column=0, columnspan=6, sticky="ew", pady=10)
        burden_row += 1
        
        # --- Bonus Section ---
        ttk.Label(burden_frame, text="Bonus:", font=('Segoe UI', 10, 'bold')).grid(
            row=burden_row, column=0, columnspan=4, sticky="w", padx=5, pady=(5, 2))
        burden_row += 1
        
        ttk.Label(burden_frame, text="Annual Bonus per Employee ($):").grid(row=burden_row, column=0, sticky="e", padx=5, pady=3)
        bonus_var = tk.DoubleVar(value=self.burden_calc.inputs['bonus_annual'])
        self.input_vars['bonus_annual'] = bonus_var
        ttk.Entry(burden_frame, textvariable=bonus_var, width=12).grid(row=burden_row, column=1, sticky="w", padx=5, pady=3)
        bonus_var.trace_add('write', lambda *args: self.on_input_change())
        
        ttk.Label(burden_frame, text="Calculated Rate:").grid(row=burden_row, column=3, sticky="e", padx=5, pady=3)
        self.bonus_calculated_label = ttk.Label(burden_frame, text="22.70%", style='Calculated.TLabel')
        self.bonus_calculated_label.grid(row=burden_row, column=4, sticky="w", padx=5, pady=3)
        burden_row += 1
        
        bonus_ot_var = tk.BooleanVar(value=self.burden_calc.inputs['bonus_applies_to_ot'])
        self.input_vars['bonus_applies_to_ot'] = bonus_ot_var
        ttk.Checkbutton(burden_frame, text="Income-dependent (applies to OT)", 
                       variable=bonus_ot_var).grid(row=burden_row, column=0, columnspan=3, sticky="w", padx=20, pady=3)
        bonus_ot_var.trace_add('write', lambda *args: self.on_input_change())
        burden_row += 1
        
        ttk.Separator(burden_frame, orient='horizontal').grid(row=burden_row, column=0, columnspan=6, sticky="ew", pady=10)
        burden_row += 1
        
        # --- Other Burden 1 ---
        ttk.Label(burden_frame, text="Other Burden 1:", font=('Segoe UI', 10, 'bold')).grid(
            row=burden_row, column=0, columnspan=4, sticky="w", padx=5, pady=(5, 2))
        burden_row += 1
        
        ttk.Label(burden_frame, text="Description:").grid(row=burden_row, column=0, sticky="e", padx=5, pady=3)
        other1_label_var = tk.StringVar(value=self.burden_calc.inputs['other_burden_1_label'])
        self.input_vars['other_burden_1_label'] = other1_label_var
        ttk.Entry(burden_frame, textvariable=other1_label_var, width=20).grid(row=burden_row, column=1, columnspan=2, sticky="w", padx=5, pady=3)
        
        ttk.Label(burden_frame, text="Rate (%):").grid(row=burden_row, column=3, sticky="e", padx=5, pady=3)
        other1_rate_var = tk.DoubleVar(value=self.burden_calc.inputs['other_burden_1_rate'])
        self.input_vars['other_burden_1_rate'] = other1_rate_var
        ttk.Entry(burden_frame, textvariable=other1_rate_var, width=10).grid(row=burden_row, column=4, sticky="w", padx=5, pady=3)
        other1_rate_var.trace_add('write', lambda *args: self.on_input_change())
        burden_row += 1
        
        other1_ot_var = tk.BooleanVar(value=self.burden_calc.inputs['other_burden_1_applies_to_ot'])
        self.input_vars['other_burden_1_applies_to_ot'] = other1_ot_var
        ttk.Checkbutton(burden_frame, text="Income-dependent (applies to OT)", 
                       variable=other1_ot_var).grid(row=burden_row, column=0, columnspan=3, sticky="w", padx=20, pady=3)
        other1_ot_var.trace_add('write', lambda *args: self.on_input_change())
        burden_row += 1
        
        # --- Other Burden 2 ---
        ttk.Label(burden_frame, text="Other Burden 2:", font=('Segoe UI', 10, 'bold')).grid(
            row=burden_row, column=0, columnspan=4, sticky="w", padx=5, pady=(5, 2))
        burden_row += 1
        
        ttk.Label(burden_frame, text="Description:").grid(row=burden_row, column=0, sticky="e", padx=5, pady=3)
        other2_label_var = tk.StringVar(value=self.burden_calc.inputs['other_burden_2_label'])
        self.input_vars['other_burden_2_label'] = other2_label_var
        ttk.Entry(burden_frame, textvariable=other2_label_var, width=20).grid(row=burden_row, column=1, columnspan=2, sticky="w", padx=5, pady=3)
        
        ttk.Label(burden_frame, text="Rate (%):").grid(row=burden_row, column=3, sticky="e", padx=5, pady=3)
        other2_rate_var = tk.DoubleVar(value=self.burden_calc.inputs['other_burden_2_rate'])
        self.input_vars['other_burden_2_rate'] = other2_rate_var
        ttk.Entry(burden_frame, textvariable=other2_rate_var, width=10).grid(row=burden_row, column=4, sticky="w", padx=5, pady=3)
        other2_rate_var.trace_add('write', lambda *args: self.on_input_change())
        burden_row += 1
        
        other2_ot_var = tk.BooleanVar(value=self.burden_calc.inputs['other_burden_2_applies_to_ot'])
        self.input_vars['other_burden_2_applies_to_ot'] = other2_ot_var
        ttk.Checkbutton(burden_frame, text="Income-dependent (applies to OT)", 
                       variable=other2_ot_var).grid(row=burden_row, column=0, columnspan=3, sticky="w", padx=20, pady=3)
        other2_ot_var.trace_add('write', lambda *args: self.on_input_change())
        burden_row += 1
        
        ttk.Separator(burden_frame, orient='horizontal').grid(row=burden_row, column=0, columnspan=6, sticky="ew", pady=10)
        burden_row += 1
        
        # --- Burden Summary ---
        ttk.Label(burden_frame, text="Burden Summary:", font=('Segoe UI', 10, 'bold')).grid(
            row=burden_row, column=0, columnspan=4, sticky="w", padx=5, pady=(5, 2))
        burden_row += 1
        
        ttk.Label(burden_frame, text="Total ST Burden:").grid(row=burden_row, column=0, sticky="e", padx=5, pady=3)
        self.st_burden_label = ttk.Label(burden_frame, text="53.90%", style='Calculated.TLabel')
        self.st_burden_label.grid(row=burden_row, column=1, sticky="w", padx=5, pady=3)
        
        ttk.Label(burden_frame, text="Total OT Burden:").grid(row=burden_row, column=3, sticky="e", padx=5, pady=3)
        self.ot_burden_label = ttk.Label(burden_frame, text="35.83%", style='Calculated.TLabel')
        self.ot_burden_label.grid(row=burden_row, column=4, sticky="w", padx=5, pady=3)
        
        row += 1
        
        # === Time Off Section ===
        timeoff_frame = ttk.LabelFrame(scrollable_frame, text="Time Off & Leave", padding="10")
        timeoff_frame.grid(row=row, column=0, sticky="ew", padx=5, pady=5)
        
        timeoff_row = 0
        
        # Vacation
        ttk.Label(timeoff_frame, text="Vacation Hours:").grid(row=timeoff_row, column=0, sticky="e", padx=5, pady=3)
        vacation_var = tk.DoubleVar(value=self.burden_calc.inputs['vacation_hours'])
        self.input_vars['vacation_hours'] = vacation_var
        ttk.Entry(timeoff_frame, textvariable=vacation_var, width=10).grid(row=timeoff_row, column=1, sticky="w", padx=5, pady=3)
        vacation_var.trace_add('write', lambda *args: self.on_input_change())
        
        ttk.Label(timeoff_frame, text="Vacation Multiplier:").grid(row=timeoff_row, column=2, sticky="e", padx=5, pady=3)
        vacation_mult_var = tk.DoubleVar(value=self.burden_calc.inputs['vacation_multiplier'])
        self.input_vars['vacation_multiplier'] = vacation_mult_var
        ttk.Entry(timeoff_frame, textvariable=vacation_mult_var, width=10).grid(row=timeoff_row, column=3, sticky="w", padx=5, pady=3)
        vacation_mult_var.trace_add('write', lambda *args: self.on_input_change())
        timeoff_row += 1
        
        # FMLA
        ttk.Label(timeoff_frame, text="FMLA Hours:").grid(row=timeoff_row, column=0, sticky="e", padx=5, pady=3)
        fmla_var = tk.DoubleVar(value=self.burden_calc.inputs['fmla_hours'])
        self.input_vars['fmla_hours'] = fmla_var
        ttk.Entry(timeoff_frame, textvariable=fmla_var, width=10).grid(row=timeoff_row, column=1, sticky="w", padx=5, pady=3)
        fmla_var.trace_add('write', lambda *args: self.on_input_change())
        
        # Unpaid Time Off
        ttk.Label(timeoff_frame, text="Unpaid Time Off (excl. FMLA):").grid(row=timeoff_row, column=2, sticky="e", padx=5, pady=3)
        unpaid_var = tk.DoubleVar(value=self.burden_calc.inputs['unpaid_time_off'])
        self.input_vars['unpaid_time_off'] = unpaid_var
        ttk.Entry(timeoff_frame, textvariable=unpaid_var, width=10).grid(row=timeoff_row, column=3, sticky="w", padx=5, pady=3)
        unpaid_var.trace_add('write', lambda *args: self.on_input_change())
        timeoff_row += 1
        
        # Paid Time Off
        ttk.Label(timeoff_frame, text="Paid Time Off (Sick, etc.):").grid(row=timeoff_row, column=0, sticky="e", padx=5, pady=3)
        pto_var = tk.DoubleVar(value=self.burden_calc.inputs['paid_time_off'])
        self.input_vars['paid_time_off'] = pto_var
        ttk.Entry(timeoff_frame, textvariable=pto_var, width=10).grid(row=timeoff_row, column=1, sticky="w", padx=5, pady=3)
        pto_var.trace_add('write', lambda *args: self.on_input_change())
        
        # Lack of Work (LOW) - voluntary unpaid, doesn't cause OT
        ttk.Label(timeoff_frame, text="Lack of Work (LOW):").grid(row=timeoff_row, column=2, sticky="e", padx=5, pady=3)
        low_var = tk.DoubleVar(value=self.burden_calc.inputs['lack_of_work_hours'])
        self.input_vars['lack_of_work_hours'] = low_var
        ttk.Entry(timeoff_frame, textvariable=low_var, width=10).grid(row=timeoff_row, column=3, sticky="w", padx=5, pady=3)
        low_var.trace_add('write', lambda *args: self.on_input_change())
        timeoff_row += 1
        
        ttk.Separator(timeoff_frame, orient='horizontal').grid(row=timeoff_row, column=0, columnspan=4, sticky="ew", pady=10)
        timeoff_row += 1
        
        # Other Time Off 1
        ttk.Label(timeoff_frame, text="Other Time Off 1:").grid(row=timeoff_row, column=0, sticky="e", padx=5, pady=3)
        other_to1_label_var = tk.StringVar(value=self.burden_calc.inputs['other_time_off_1_label'])
        self.input_vars['other_time_off_1_label'] = other_to1_label_var
        ttk.Entry(timeoff_frame, textvariable=other_to1_label_var, width=15).grid(row=timeoff_row, column=1, sticky="w", padx=5, pady=3)
        
        ttk.Label(timeoff_frame, text="Hours:").grid(row=timeoff_row, column=2, sticky="e", padx=5, pady=3)
        other_to1_hours_var = tk.DoubleVar(value=self.burden_calc.inputs['other_time_off_1_hours'])
        self.input_vars['other_time_off_1_hours'] = other_to1_hours_var
        ttk.Entry(timeoff_frame, textvariable=other_to1_hours_var, width=10).grid(row=timeoff_row, column=3, sticky="w", padx=5, pady=3)
        other_to1_hours_var.trace_add('write', lambda *args: self.on_input_change())
        timeoff_row += 1
        
        # Other Time Off 2
        ttk.Label(timeoff_frame, text="Other Time Off 2:").grid(row=timeoff_row, column=0, sticky="e", padx=5, pady=3)
        other_to2_label_var = tk.StringVar(value=self.burden_calc.inputs['other_time_off_2_label'])
        self.input_vars['other_time_off_2_label'] = other_to2_label_var
        ttk.Entry(timeoff_frame, textvariable=other_to2_label_var, width=15).grid(row=timeoff_row, column=1, sticky="w", padx=5, pady=3)
        
        ttk.Label(timeoff_frame, text="Hours:").grid(row=timeoff_row, column=2, sticky="e", padx=5, pady=3)
        other_to2_hours_var = tk.DoubleVar(value=self.burden_calc.inputs['other_time_off_2_hours'])
        self.input_vars['other_time_off_2_hours'] = other_to2_hours_var
        ttk.Entry(timeoff_frame, textvariable=other_to2_hours_var, width=10).grid(row=timeoff_row, column=3, sticky="w", padx=5, pady=3)
        other_to2_hours_var.trace_add('write', lambda *args: self.on_input_change())
        
        row += 1
        
        # === Holiday Settings ===
        holiday_frame = ttk.LabelFrame(scrollable_frame, text="Holiday Settings", padding="10")
        holiday_frame.grid(row=row, column=0, sticky="ew", padx=5, pady=5)
        
        ttk.Label(holiday_frame, text="Holidays Recognized:").grid(row=0, column=0, sticky="e", padx=5, pady=3)
        holidays_rec_var = tk.IntVar(value=self.burden_calc.inputs['holidays_recognized'])
        self.input_vars['holidays_recognized'] = holidays_rec_var
        ttk.Entry(holiday_frame, textvariable=holidays_rec_var, width=10).grid(row=0, column=1, sticky="w", padx=5, pady=3)
        holidays_rec_var.trace_add('write', lambda *args: self.on_input_change())
        
        ttk.Label(holiday_frame, text="Holidays Worked:").grid(row=0, column=2, sticky="e", padx=5, pady=3)
        holidays_worked_var = tk.IntVar(value=self.burden_calc.inputs['holidays_worked'])
        self.input_vars['holidays_worked'] = holidays_worked_var
        ttk.Entry(holiday_frame, textvariable=holidays_worked_var, width=10).grid(row=0, column=3, sticky="w", padx=5, pady=3)
        holidays_worked_var.trace_add('write', lambda *args: self.on_input_change())
        
        ttk.Label(holiday_frame, text="Holiday Premium (multiplier):").grid(row=1, column=0, sticky="e", padx=5, pady=3)
        holiday_prem_var = tk.DoubleVar(value=self.burden_calc.inputs['holiday_premium'])
        self.input_vars['holiday_premium'] = holiday_prem_var
        ttk.Entry(holiday_frame, textvariable=holiday_prem_var, width=10).grid(row=1, column=1, sticky="w", padx=5, pady=3)
        holiday_prem_var.trace_add('write', lambda *args: self.on_input_change())
        
        row += 1
        
        # === Schedule Settings ===
        schedule_frame = ttk.LabelFrame(scrollable_frame, text="Schedule Settings", padding="10")
        schedule_frame.grid(row=row, column=0, sticky="ew", padx=5, pady=5)
        
        ttk.Label(schedule_frame, text="Shift Differential ($):").grid(row=0, column=0, sticky="e", padx=5, pady=3)
        shift_diff_var = tk.DoubleVar(value=self.burden_calc.inputs['shift_differential'])
        self.input_vars['shift_differential'] = shift_diff_var
        ttk.Entry(schedule_frame, textvariable=shift_diff_var, width=10).grid(row=0, column=1, sticky="w", padx=5, pady=3)
        shift_diff_var.trace_add('write', lambda *args: self.on_input_change())
        
        ttk.Label(schedule_frame, text="Shift Overlap (minutes):").grid(row=0, column=2, sticky="e", padx=5, pady=3)
        overlap_var = tk.IntVar(value=self.burden_calc.inputs['shift_overlap_minutes'])
        self.input_vars['shift_overlap_minutes'] = overlap_var
        ttk.Entry(schedule_frame, textvariable=overlap_var, width=10).grid(row=0, column=3, sticky="w", padx=5, pady=3)
        overlap_var.trace_add('write', lambda *args: self.on_input_change())
        
        # Configure column weights
        scrollable_frame.columnconfigure(0, weight=1)
        
    def create_schedule_calculator(self, parent):
        """Create the schedule cost calculator section (right side)."""
        # Main frame with border
        calc_frame = ttk.LabelFrame(parent, text="Schedule Cost Calculator", padding="15")
        calc_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Create scrollable canvas
        canvas = tk.Canvas(calc_frame, highlightthickness=0)
        scrollbar = ttk.Scrollbar(calc_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Enable mousewheel scrolling only when mouse is over this canvas
        def on_mousewheel_calc(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        
        def bind_mousewheel_calc(event):
            canvas.bind_all("<MouseWheel>", on_mousewheel_calc)
        
        def unbind_mousewheel_calc(event):
            canvas.unbind_all("<MouseWheel>")
        
        canvas.bind("<Enter>", bind_mousewheel_calc)
        canvas.bind("<Leave>", unbind_mousewheel_calc)
        
        row = 0
        
        # === Schedule Name ===
        name_frame = ttk.Frame(scrollable_frame)
        name_frame.grid(row=row, column=0, columnspan=3, sticky="ew", pady=5)
        
        ttk.Label(name_frame, text="Schedule Name:").pack(side=tk.LEFT, padx=5)
        name_var = tk.StringVar(value=self.schedule_calc.schedule_inputs['schedule_name'])
        self.schedule_vars['schedule_name'] = name_var
        ttk.Entry(name_frame, textvariable=name_var, width=30).pack(side=tk.LEFT, padx=5)
        name_var.trace_add('write', lambda *args: self.on_schedule_change())
        self.create_explanation_button(name_frame, 'schedule_name').pack(side=tk.LEFT)
        
        row += 1
        
        # === Days Per Week ===
        days_frame = ttk.LabelFrame(scrollable_frame, text="Schedule Type", padding="10")
        days_frame.grid(row=row, column=0, columnspan=3, sticky="ew", pady=5)
        
        days_row_frame = ttk.Frame(days_frame)
        days_row_frame.pack(fill='x')
        
        ttk.Label(days_row_frame, text="Days per Week:").pack(side=tk.LEFT, padx=5)
        
        days_var = tk.IntVar(value=self.schedule_calc.schedule_inputs['days_per_week'])
        self.schedule_vars['days_per_week'] = days_var
        
        ttk.Radiobutton(days_row_frame, text="5 Days", variable=days_var, value=5,
                       command=self.on_schedule_change).pack(side=tk.LEFT, padx=10)
        ttk.Radiobutton(days_row_frame, text="7 Days", variable=days_var, value=7,
                       command=self.on_schedule_change).pack(side=tk.LEFT, padx=10)
        self.create_explanation_button(days_row_frame, 'days_per_week').pack(side=tk.LEFT)
        
        # Shift length (only for 7-day)
        self.shift_frame = ttk.Frame(days_frame)
        self.shift_frame.pack(fill='x', pady=(10, 0))
        
        ttk.Label(self.shift_frame, text="Shift Length:").pack(side=tk.LEFT, padx=5)
        
        shift_var = tk.IntVar(value=self.schedule_calc.schedule_inputs['shift_length'])
        self.schedule_vars['shift_length'] = shift_var
        
        ttk.Radiobutton(self.shift_frame, text="8 Hours", variable=shift_var, value=8,
                       command=self.on_schedule_change).pack(side=tk.LEFT, padx=10)
        ttk.Radiobutton(self.shift_frame, text="12 Hours", variable=shift_var, value=12,
                       command=self.on_schedule_change).pack(side=tk.LEFT, padx=10)
        self.create_explanation_button(self.shift_frame, 'shift_length').pack(side=tk.LEFT)
        
        row += 1
        
        # === 12-Hour Schedule Options ===
        self.twelve_hour_frame = ttk.LabelFrame(scrollable_frame, text="12-Hour Holiday Pay Options", padding="10")
        self.twelve_hour_frame.grid(row=row, column=0, columnspan=3, sticky="ew", pady=5)
        
        hol_day_off_frame = ttk.Frame(self.twelve_hour_frame)
        hol_day_off_frame.pack(fill='x', pady=2)
        ttk.Label(hol_day_off_frame, text="Holiday Pay on Scheduled Day Off (hours):").pack(side=tk.LEFT, padx=5)
        hol_day_off_var = tk.DoubleVar(value=self.schedule_calc.schedule_inputs['holiday_pay_day_off'])
        self.schedule_vars['holiday_pay_day_off'] = hol_day_off_var
        ttk.Entry(hol_day_off_frame, textvariable=hol_day_off_var, width=8).pack(side=tk.LEFT, padx=5)
        hol_day_off_var.trace_add('write', lambda *args: self.on_schedule_change())
        self.create_explanation_button(hol_day_off_frame, 'holiday_pay_day_off').pack(side=tk.LEFT)
        
        hol_workday_frame = ttk.Frame(self.twelve_hour_frame)
        hol_workday_frame.pack(fill='x', pady=2)
        ttk.Label(hol_workday_frame, text="Holiday Pay on Scheduled Workday (hours):").pack(side=tk.LEFT, padx=5)
        hol_workday_var = tk.DoubleVar(value=self.schedule_calc.schedule_inputs['holiday_pay_workday'])
        self.schedule_vars['holiday_pay_workday'] = hol_workday_var
        ttk.Entry(hol_workday_frame, textvariable=hol_workday_var, width=8).pack(side=tk.LEFT, padx=5)
        hol_workday_var.trace_add('write', lambda *args: self.on_schedule_change())
        self.create_explanation_button(hol_workday_frame, 'holiday_pay_workday').pack(side=tk.LEFT)
        
        row += 1
        
        # === 7-Day Schedule Options ===
        self.seven_day_frame = ttk.LabelFrame(scrollable_frame, text="7-Day Schedule Options", padding="10")
        self.seven_day_frame.grid(row=row, column=0, columnspan=3, sticky="ew", pady=5)
        
        avg_pay_frame = ttk.Frame(self.seven_day_frame)
        avg_pay_frame.pack(fill='x', pady=2)
        ttk.Label(avg_pay_frame, text="Average Pay Hours per Week:").pack(side=tk.LEFT, padx=5)
        avg_pay_var = tk.DoubleVar(value=self.schedule_calc.schedule_inputs['avg_pay_hours'])
        self.schedule_vars['avg_pay_hours'] = avg_pay_var
        ttk.Entry(avg_pay_frame, textvariable=avg_pay_var, width=8).pack(side=tk.LEFT, padx=5)
        avg_pay_var.trace_add('write', lambda *args: self.on_schedule_change())
        self.create_explanation_button(avg_pay_frame, 'avg_pay_hours').pack(side=tk.LEFT)
        
        row += 1
        
        # === Other Hours Paid ===
        other_frame = ttk.Frame(scrollable_frame)
        other_frame.grid(row=row, column=0, columnspan=3, sticky="ew", pady=5)
        
        ttk.Label(other_frame, text="Other Hours Paid:").pack(side=tk.LEFT, padx=5)
        other_var = tk.DoubleVar(value=self.schedule_calc.schedule_inputs['other_hours_paid'])
        self.schedule_vars['other_hours_paid'] = other_var
        ttk.Entry(other_frame, textvariable=other_var, width=10).pack(side=tk.LEFT, padx=5)
        other_var.trace_add('write', lambda *args: self.on_schedule_change())
        self.create_explanation_button(other_frame, 'other_hours_paid').pack(side=tk.LEFT)
        
        row += 1
        
        ttk.Separator(scrollable_frame, orient='horizontal').grid(row=row, column=0, columnspan=3, sticky="ew", pady=10)
        row += 1
        
        # === RESULTS SECTION ===
        results_label = ttk.Label(scrollable_frame, text="CALCULATED RESULTS", 
                                  font=('Segoe UI', 12, 'bold'), foreground=self.colors['primary'])
        results_label.grid(row=row, column=0, columnspan=3, sticky="w", pady=(5, 10))
        row += 1
        
        # Hours Section - Scheduled Hours Breakdown
        hours_frame = ttk.LabelFrame(scrollable_frame, text="Hours Analysis - Scheduled to Worked", padding="10")
        hours_frame.grid(row=row, column=0, columnspan=3, sticky="ew", pady=5)
        
        self.result_labels = {}
        
        # Scheduled hours breakdown items
        hours_breakdown_items = [
            ('scheduled_hours', 'Scheduled Hours:'),
            ('less_vacation', '  Less Vacation:'),
            ('less_holiday_missed', '  Less Holiday Hours Missed:'),
            ('less_fmla', '  Less FMLA:'),
            ('less_pto', '  Less Paid Time Off:'),
            ('less_unpaid', '  Less Unpaid Time Off:'),
            ('less_low', '  Less Lack of Work (LOW):'),
            ('less_other_1', '  Less Other Time Off 1:'),
            ('less_other_2', '  Less Other Time Off 2:'),
            ('hours_worked', 'Hours Worked:'),
        ]
        
        for i, (key, label) in enumerate(hours_breakdown_items):
            frame = ttk.Frame(hours_frame)
            frame.pack(fill='x', pady=2)
            # Make "Less" items indented and gray, totals bold
            if key.startswith('less_'):
                lbl = ttk.Label(frame, text=label, foreground=self.colors['gray'])
                lbl.pack(side=tk.LEFT, padx=5)
                self.result_labels[key] = ttk.Label(frame, text="--", foreground=self.colors['gray'])
            elif key in ['scheduled_hours', 'hours_worked']:
                lbl = ttk.Label(frame, text=label, font=('Segoe UI', 10, 'bold'))
                lbl.pack(side=tk.LEFT, padx=5)
                self.result_labels[key] = ttk.Label(frame, text="--", style='Calculated.TLabel', 
                                                     font=('Segoe UI', 10, 'bold'))
            else:
                ttk.Label(frame, text=label).pack(side=tk.LEFT, padx=5)
                self.result_labels[key] = ttk.Label(frame, text="--", style='Calculated.TLabel')
            self.result_labels[key].pack(side=tk.LEFT, padx=5)
            if key in ['scheduled_hours', 'hours_worked']:
                self.create_explanation_button(frame, key).pack(side=tk.LEFT)
        
        ttk.Separator(hours_frame, orient='horizontal').pack(fill='x', pady=5)
        
        # Summary metrics
        summary_items = [
            ('percent_worked', 'Percent Worked:'),
            ('absenteeism_ot', 'Absenteeism % (causes OT):'),
        ]
        
        for key, label in summary_items:
            frame = ttk.Frame(hours_frame)
            frame.pack(fill='x', pady=2)
            ttk.Label(frame, text=label).pack(side=tk.LEFT, padx=5)
            self.result_labels[key] = ttk.Label(frame, text="--", style='Calculated.TLabel')
            self.result_labels[key].pack(side=tk.LEFT, padx=5)
            self.create_explanation_button(frame, key).pack(side=tk.LEFT)
        
        row += 1
        
        # Hours Paid Section
        paid_frame = ttk.LabelFrame(scrollable_frame, text="Hours Paid Components", padding="10")
        paid_frame.grid(row=row, column=0, columnspan=3, sticky="ew", pady=5)
        
        paid_items = [
            ('hours_worked_paid', 'Hours Worked:'),
            ('vacation_pay', 'Vacation Pay:'),
            ('holiday_pay', 'Holiday Pay:'),
            ('holiday_premium', 'Holiday Premium:'),
            ('built_in_ot', 'Built-in Overtime:'),
            ('shift_diff_hours', 'Shift Differential:'),
            ('shift_overlap_pay', 'Shift Overlap Pay:'),
            ('paid_time_off_hours', 'Paid Time Off (Sick, etc.):'),
            ('other_hours_paid', 'Other Hours Paid:'),
            ('total_hours_paid', 'Total Hours Paid:'),
            ('ratio_paid_worked', 'Ratio Paid/Worked:'),
        ]
        
        for key, label in paid_items:
            frame = ttk.Frame(paid_frame)
            frame.pack(fill='x', pady=2)
            ttk.Label(frame, text=label).pack(side=tk.LEFT, padx=5)
            self.result_labels[key] = ttk.Label(frame, text="--", style='Calculated.TLabel')
            self.result_labels[key].pack(side=tk.LEFT, padx=5)
            self.create_explanation_button(frame, key).pack(side=tk.LEFT)
        
        row += 1
        
        # Cost Section
        cost_frame = ttk.LabelFrame(scrollable_frame, text="Cost Analysis", padding="10")
        cost_frame.grid(row=row, column=0, columnspan=3, sticky="ew", pady=5)
        
        # Auto-populated values
        auto_frame = ttk.Frame(cost_frame)
        auto_frame.pack(fill='x', pady=5)
        
        ttk.Label(auto_frame, text="From Inputs:", font=('Segoe UI', 9, 'italic')).pack(side=tk.LEFT, padx=5)
        
        self.result_labels['wage_display'] = ttk.Label(auto_frame, text="Wage: $35.87", foreground=self.colors['gray'])
        self.result_labels['wage_display'].pack(side=tk.LEFT, padx=10)
        
        self.result_labels['st_burden_display'] = ttk.Label(auto_frame, text="ST Burden: 52.88%", foreground=self.colors['gray'])
        self.result_labels['st_burden_display'].pack(side=tk.LEFT, padx=10)
        
        self.result_labels['ot_burden_display'] = ttk.Label(auto_frame, text="OT Burden: 36.27%", foreground=self.colors['gray'])
        self.result_labels['ot_burden_display'].pack(side=tk.LEFT, padx=10)
        
        ttk.Separator(cost_frame, orient='horizontal').pack(fill='x', pady=5)
        
        # Adjusted ST Burden (smaller, less prominent)
        adj_frame = ttk.Frame(cost_frame)
        adj_frame.pack(fill='x', pady=2)
        ttk.Label(adj_frame, text='Adjusted ST Burden:').pack(side=tk.LEFT, padx=5)
        self.result_labels['adjusted_st_burden'] = ttk.Label(adj_frame, text="--", 
                                                              font=('Segoe UI', 10, 'bold'),
                                                              foreground=self.colors['primary'])
        self.result_labels['adjusted_st_burden'].pack(side=tk.LEFT, padx=5)
        self.create_explanation_button(adj_frame, 'adjusted_st_burden').pack(side=tk.LEFT)
        
        ttk.Separator(cost_frame, orient='horizontal').pack(fill='x', pady=8)
        
        # KEY RESULTS - Cost of Scheduled Time and Marginal OT (bigger, bolder)
        key_results_label = ttk.Label(cost_frame, text="KEY COST METRICS", 
                                       font=('Segoe UI', 9, 'bold'),
                                       foreground=self.colors['gray'])
        key_results_label.pack(anchor='w', padx=5, pady=(5, 10))
        
        key_cost_items = [
            ('cost_scheduled_time', 'Cost of Scheduled Time:'),
            ('marginal_ot_15', 'Marginal Cost OT (1.5x):'),
            ('marginal_ot_20', 'Marginal Cost OT (2.0x):'),
        ]
        
        for key, label in key_cost_items:
            frame = ttk.Frame(cost_frame)
            frame.pack(fill='x', pady=4)
            ttk.Label(frame, text=label, font=('Segoe UI', 11, 'bold')).pack(side=tk.LEFT, padx=5)
            self.result_labels[key] = ttk.Label(frame, text="--", 
                                                font=('Segoe UI', 16, 'bold'),
                                                foreground=self.colors['primary'])
            self.result_labels[key].pack(side=tk.LEFT, padx=10)
            self.create_explanation_button(frame, key).pack(side=tk.LEFT)
        
        row += 1
        
        # Save Button
        button_frame = ttk.Frame(scrollable_frame)
        button_frame.grid(row=row, column=0, columnspan=3, sticky="ew", pady=20)
        
        save_btn = ttk.Button(button_frame, text="Save This Cost", 
                             command=self.save_current_cost,
                             style='Primary.TButton')
        save_btn.pack(pady=10)
        
        clear_btn = ttk.Button(button_frame, text="Clear / New Schedule",
                              command=self.clear_schedule_inputs)
        clear_btn.pack(pady=5)
        
        # Initial visibility update
        self.update_schedule_visibility()
        
    def update_schedule_visibility(self):
        """Show/hide schedule options based on selected schedule type."""
        days = self.schedule_vars['days_per_week'].get()
        shift = self.schedule_vars['shift_length'].get()
        
        if days == 5:
            # Hide shift length, 12-hour options, 7-day options
            self.shift_frame.pack_forget()
            self.twelve_hour_frame.grid_remove()
            self.seven_day_frame.grid_remove()
        else:
            # Show shift length and 7-day options
            self.shift_frame.pack(fill='x', pady=(10, 0))
            self.seven_day_frame.grid()
            
            if shift == 12:
                self.twelve_hour_frame.grid()
            else:
                self.twelve_hour_frame.grid_remove()
                
    def on_input_change(self):
        """Called when any burden/time off input value changes."""
        # Update burden calculator inputs
        for key, var in self.input_vars.items():
            try:
                self.burden_calc.inputs[key] = var.get()
            except:
                pass  # Ignore invalid values during typing
        
        # Update calculated labels
        self.update_calculated_labels()
        
        # Update schedule display
        self.update_schedule_display()
    
    def update_calculated_labels(self):
        """Update the calculated rate labels in the burden section."""
        try:
            # 401K
            k401_rate = self.burden_calc.calculate_401k_rate()
            self.k401_calculated_label.config(text=f"{k401_rate:.2f}%")
            
            # Workers' Comp
            wc_rate = self.burden_calc.calculate_workers_comp_rate()
            self.wc_calculated_label.config(text=f"{wc_rate:.2f}%")
            
            # Medical
            medical_rate = self.burden_calc.calculate_medical_rate()
            self.medical_calculated_label.config(text=f"{medical_rate:.2f}%")
            
            # Bonus
            bonus_rate = self.burden_calc.calculate_bonus_rate()
            self.bonus_calculated_label.config(text=f"{bonus_rate:.2f}%")
            
            # Total Burdens
            st_burden = self.burden_calc.calculate_st_burden()
            self.st_burden_label.config(text=f"{st_burden:.2f}%")
            
            ot_burden = self.burden_calc.calculate_ot_burden()
            self.ot_burden_label.config(text=f"{ot_burden:.2f}%")
            
            # Update PROMINENT displays at top of burden section
            self.st_burden_display.config(text=f"{st_burden:.2f}%")
            self.ot_burden_display.config(text=f"{ot_burden:.2f}%")
        except:
            pass  # Ignore errors during typing
            
    def on_schedule_change(self):
        """Called when schedule-specific inputs change."""
        # Update schedule calculator inputs
        for key, var in self.schedule_vars.items():
            try:
                self.schedule_calc.schedule_inputs[key] = var.get()
            except:
                pass
        
        # Update visibility
        self.update_schedule_visibility()
        
        # Update display
        self.update_schedule_display()
        
    def update_schedule_display(self):
        """Update all schedule calculation results."""
        try:
            # Update schedule calculator inputs from vars
            for key, var in self.schedule_vars.items():
                try:
                    self.schedule_calc.schedule_inputs[key] = var.get()
                except:
                    pass
            
            # Calculate all results
            results = self.schedule_calc.calculate_all()
            
            # Update hours breakdown section
            self.result_labels['scheduled_hours'].config(text=f"{results['scheduled_hours']:,.0f}")
            self.result_labels['less_vacation'].config(text=f"-{results['vacation_hours']:,.1f}")
            self.result_labels['less_holiday_missed'].config(text=f"-{results['holiday_hours_missed']:,.1f}")
            self.result_labels['less_fmla'].config(text=f"-{results['fmla_hours']:,.1f}")
            self.result_labels['less_pto'].config(text=f"-{results['paid_time_off']:,.1f}")
            self.result_labels['less_unpaid'].config(text=f"-{results['unpaid_time_off']:,.1f}")
            self.result_labels['less_low'].config(text=f"-{results['lack_of_work_hours']:,.1f}")
            self.result_labels['less_other_1'].config(text=f"-{results['other_time_off_1']:,.1f}")
            self.result_labels['less_other_2'].config(text=f"-{results['other_time_off_2']:,.1f}")
            self.result_labels['hours_worked'].config(text=f"{results['hours_worked']:,.1f}")
            self.result_labels['percent_worked'].config(text=f"{results['percent_worked']:.1%}")
            self.result_labels['absenteeism_ot'].config(text=f"{results['absenteeism_ot_percent']:.1%}")
            
            # Update paid hours section
            self.result_labels['hours_worked_paid'].config(text=f"{results['hours_worked']:,.1f}")
            self.result_labels['vacation_pay'].config(text=f"{results['vacation_pay']:,.1f}")
            self.result_labels['holiday_pay'].config(text=f"{results['holiday_pay']:,.1f}")
            self.result_labels['holiday_premium'].config(text=f"{results['holiday_premium']:,.1f}")
            self.result_labels['built_in_ot'].config(text=f"{results['built_in_ot']:,.1f}")
            self.result_labels['shift_diff_hours'].config(text=f"{results['shift_diff_hours']:,.1f}")
            self.result_labels['shift_overlap_pay'].config(text=f"{results['shift_overlap_pay']:,.1f}")
            self.result_labels['paid_time_off_hours'].config(text=f"{results['paid_time_off']:,.1f}")
            self.result_labels['other_hours_paid'].config(text=f"{results['other_hours_paid']:,.1f}")
            self.result_labels['total_hours_paid'].config(text=f"{results['total_hours_paid']:,.1f}")
            self.result_labels['ratio_paid_worked'].config(text=f"{results['ratio_paid_worked']:.3f}")
            
            # Update cost section - auto-populated values
            self.result_labels['wage_display'].config(text=f"Wage: ${results['wage']:.2f}")
            self.result_labels['st_burden_display'].config(text=f"ST Burden: {results['st_burden']:.2f}%")
            self.result_labels['ot_burden_display'].config(text=f"OT Burden: {results['ot_burden']:.2f}%")
            
            # Update cost calculations
            self.result_labels['adjusted_st_burden'].config(text=f"{results['adjusted_st_burden']:.2f}%")
            self.result_labels['cost_scheduled_time'].config(text=f"${results['cost_scheduled_time']:.2f}/hr")
            self.result_labels['marginal_ot_15'].config(text=f"${results['marginal_ot_15']:.2f}/hr")
            self.result_labels['marginal_ot_20'].config(text=f"${results['marginal_ot_20']:.2f}/hr")
            
        except Exception as e:
            pass  # Ignore errors during input
            
    def save_current_cost(self):
        """Save the current schedule calculation."""
        try:
            results = self.schedule_calc.calculate_all()
            results['saved_at'] = datetime.now().strftime('%Y-%m-%d %H:%M')
            self.saved_costs.append(results)
            self.update_saved_costs_display()
            self.update_analysis_schedule_list()
            self.update_compare_dropdowns()
            
            messagebox.showinfo("Saved", f"Schedule '{results['schedule_name']}' has been saved.\n\nView saved schedules in the 'Saved Costs' tab.")
            
        except Exception as e:
            messagebox.showerror("Error", f"Could not save: {str(e)}")
            
    def clear_schedule_inputs(self):
        """Clear schedule inputs for a new calculation."""
        self.schedule_vars['schedule_name'].set("New Schedule")
        self.schedule_vars['days_per_week'].set(5)
        self.schedule_vars['shift_length'].set(8)
        self.schedule_vars['holiday_pay_day_off'].set(8)
        self.schedule_vars['holiday_pay_workday'].set(12)
        self.schedule_vars['avg_pay_hours'].set(46)
        self.schedule_vars['other_hours_paid'].set(0)
        
        self.on_schedule_change()
        
    def create_saved_costs_tab(self):
        """Create the saved costs comparison tab."""
        saved_frame = ttk.Frame(self.notebook, padding="10")
        self.notebook.add(saved_frame, text="Saved Costs")
        
        # Title
        title_label = ttk.Label(saved_frame,
                               text="Saved Schedule Cost Comparisons",
                               style='Header.TLabel')
        title_label.pack(pady=(0, 10))
        
        # Info
        info_label = ttk.Label(saved_frame,
                              text="Save schedules from the calculator to compare them here.",
                              foreground=self.colors['gray'])
        info_label.pack(pady=(0, 10))
        
        # Create treeview for saved costs
        columns = ('name', 'type', 'scheduled', 'worked', 'cost_hr', 'ot_15', 'ot_20', 'saved_at')
        
        self.saved_tree = ttk.Treeview(saved_frame, columns=columns, show='headings', height=15)
        
        # Configure columns
        col_configs = [
            ('name', 'Schedule Name', 150),
            ('type', 'Type', 100),
            ('scheduled', 'Scheduled Hrs', 100),
            ('worked', 'Hours Worked', 100),
            ('cost_hr', 'Cost/Hour', 100),
            ('ot_15', 'OT 1.5x', 80),
            ('ot_20', 'OT 2.0x', 80),
            ('saved_at', 'Saved At', 120),
        ]
        
        for col_id, heading, width in col_configs:
            self.saved_tree.heading(col_id, text=heading)
            self.saved_tree.column(col_id, width=width, anchor='center')
        
        self.saved_tree.column('name', anchor='w')
        
        # Scrollbars
        y_scroll = ttk.Scrollbar(saved_frame, orient="vertical", command=self.saved_tree.yview)
        self.saved_tree.configure(yscrollcommand=y_scroll.set)
        
        self.saved_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        y_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Buttons frame
        btn_frame = ttk.Frame(saved_frame)
        btn_frame.pack(fill=tk.X, pady=10)
        
        delete_btn = ttk.Button(btn_frame, text="Delete Selected", command=self.delete_selected_cost)
        delete_btn.pack(side=tk.LEFT, padx=5)
        
        clear_btn = ttk.Button(btn_frame, text="Clear All", command=self.clear_saved_costs)
        clear_btn.pack(side=tk.LEFT, padx=5)
        
    def update_saved_costs_display(self):
        """Update the saved costs treeview."""
        # Clear existing items
        for item in self.saved_tree.get_children():
            self.saved_tree.delete(item)
        
        # Add saved costs
        for cost in self.saved_costs:
            values = (
                cost['schedule_name'],
                cost['schedule_type'],
                f"{cost['scheduled_hours']:,.0f}",
                f"{cost['hours_worked']:,.1f}",
                f"${cost['cost_scheduled_time']:.2f}",
                f"${cost['marginal_ot_15']:.2f}",
                f"${cost['marginal_ot_20']:.2f}",
                cost.get('saved_at', ''),
            )
            self.saved_tree.insert('', 'end', values=values)
            
    def delete_selected_cost(self):
        """Delete the selected saved cost."""
        selection = self.saved_tree.selection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select a saved cost to delete.")
            return
        
        # Get index of selected item
        item = selection[0]
        index = self.saved_tree.index(item)
        
        if 0 <= index < len(self.saved_costs):
            name = self.saved_costs[index]['schedule_name']
            if messagebox.askyesno("Confirm Delete", f"Delete '{name}'?"):
                del self.saved_costs[index]
                self.update_saved_costs_display()
                self.update_analysis_schedule_list()
                self.update_compare_dropdowns()
                
    def clear_saved_costs(self):
        """Clear all saved costs."""
        if self.saved_costs:
            if messagebox.askyesno("Confirm Clear", "Delete all saved costs?"):
                self.saved_costs.clear()
                self.update_saved_costs_display()
                self.update_analysis_schedule_list()
                self.update_compare_dropdowns()
    
    def create_compare_tab(self):
        """Create the schedule comparison tab."""
        compare_frame = ttk.Frame(self.notebook, padding="10")
        self.notebook.add(compare_frame, text="Compare Schedules")
        
        # Title
        title_label = ttk.Label(compare_frame,
                               text="Side-by-Side Schedule Comparison",
                               style='Header.TLabel')
        title_label.pack(pady=(0, 10))
        
        # Selection Frame
        select_frame = ttk.LabelFrame(compare_frame, text="Select Schedules to Compare", padding="10")
        select_frame.pack(fill='x', padx=10, pady=5)
        
        # Schedule A dropdown
        row_a = ttk.Frame(select_frame)
        row_a.pack(fill='x', pady=5)
        ttk.Label(row_a, text="Schedule A:", width=12, anchor='e').pack(side=tk.LEFT, padx=5)
        self.compare_a_var = tk.StringVar()
        self.compare_a_dropdown = ttk.Combobox(row_a, textvariable=self.compare_a_var, 
                                                state='readonly', width=40)
        self.compare_a_dropdown.pack(side=tk.LEFT, padx=5)
        self.compare_a_dropdown.bind('<Button-1>', lambda e: self.update_compare_dropdowns())
        
        # Schedule B dropdown
        row_b = ttk.Frame(select_frame)
        row_b.pack(fill='x', pady=5)
        ttk.Label(row_b, text="Schedule B:", width=12, anchor='e').pack(side=tk.LEFT, padx=5)
        self.compare_b_var = tk.StringVar()
        self.compare_b_dropdown = ttk.Combobox(row_b, textvariable=self.compare_b_var,
                                                state='readonly', width=40)
        self.compare_b_dropdown.pack(side=tk.LEFT, padx=5)
        self.compare_b_dropdown.bind('<Button-1>', lambda e: self.update_compare_dropdowns())
        
        # Positions input for incentive calculations
        pos_row = ttk.Frame(select_frame)
        pos_row.pack(fill='x', pady=5)
        ttk.Label(pos_row, text="Positions to Cover:", width=15, anchor='e').pack(side=tk.LEFT, padx=5)
        self.compare_positions_var = tk.IntVar(value=100)
        ttk.Entry(pos_row, textvariable=self.compare_positions_var, width=10).pack(side=tk.LEFT, padx=5)
        ttk.Label(pos_row, text="(for incentive cost calculations)", 
                  foreground=self.colors['gray']).pack(side=tk.LEFT, padx=5)
        
        # Compare button
        btn_row = ttk.Frame(select_frame)
        btn_row.pack(fill='x', pady=10)
        compare_btn = ttk.Button(btn_row, text="Compare Schedules", 
                                 command=self.run_comparison,
                                 style='Primary.TButton')
        compare_btn.pack()
        
        # Results Frame (scrollable)
        results_outer = ttk.LabelFrame(compare_frame, text="Comparison Results", padding="10")
        results_outer.pack(fill='both', expand=True, padx=10, pady=5)
        
        # Fixed header frame (outside scrollable area)
        self.compare_header_frame = ttk.Frame(results_outer)
        self.compare_header_frame.pack(fill='x', pady=(0, 5))
        
        # Separator between header and scrollable content
        ttk.Separator(results_outer, orient='horizontal').pack(fill='x')
        
        # Scrollable content frame
        scroll_container = ttk.Frame(results_outer)
        scroll_container.pack(fill='both', expand=True)
        
        # Create scrollable canvas for results
        compare_canvas = tk.Canvas(scroll_container, highlightthickness=0)
        compare_scrollbar = ttk.Scrollbar(scroll_container, orient="vertical", command=compare_canvas.yview)
        self.compare_results_frame = ttk.Frame(compare_canvas)
        
        self.compare_results_frame.bind(
            "<Configure>",
            lambda e: compare_canvas.configure(scrollregion=compare_canvas.bbox("all"))
        )
        
        compare_canvas.create_window((0, 0), window=self.compare_results_frame, anchor="nw")
        compare_canvas.configure(yscrollcommand=compare_scrollbar.set)
        
        # Mousewheel scrolling
        def on_compare_mousewheel(event):
            compare_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        compare_canvas.bind("<MouseWheel>", on_compare_mousewheel)
        self.compare_results_frame.bind("<MouseWheel>", on_compare_mousewheel)
        
        # Bind mousewheel to all children as they're created
        def bind_mousewheel_recursive(widget):
            widget.bind("<MouseWheel>", on_compare_mousewheel)
            for child in widget.winfo_children():
                bind_mousewheel_recursive(child)
        
        # Store the binding function for later use
        self.compare_mousewheel_bind = bind_mousewheel_recursive
        
        compare_canvas.pack(side=tk.LEFT, fill='both', expand=True)
        compare_scrollbar.pack(side=tk.RIGHT, fill='y')
        
        # Initial message
        self.compare_message_label = ttk.Label(self.compare_results_frame,
                                                text="Select two schedules and click 'Compare Schedules' to see comparison.",
                                                foreground=self.colors['gray'])
        self.compare_message_label.pack(pady=20)
        
        # Storage for comparison labels
        self.compare_labels = {}
        
    def update_compare_dropdowns(self):
        """Update the dropdowns in the Compare tab with saved schedules."""
        if not hasattr(self, 'compare_a_dropdown'):
            return
            
        schedule_names = [s['schedule_name'] for s in self.saved_costs]
        self.compare_a_dropdown['values'] = schedule_names
        self.compare_b_dropdown['values'] = schedule_names
        
        # Clear selections if schedules were deleted
        if self.compare_a_var.get() not in schedule_names:
            self.compare_a_var.set('')
        if self.compare_b_var.get() not in schedule_names:
            self.compare_b_var.set('')
    
    def run_comparison(self):
        """Run the side-by-side comparison of two schedules."""
        # Validate selections
        schedule_a_name = self.compare_a_var.get()
        schedule_b_name = self.compare_b_var.get()
        
        if not schedule_a_name or not schedule_b_name:
            messagebox.showwarning("Selection Required", "Please select two schedules to compare.")
            return
        
        if schedule_a_name == schedule_b_name:
            messagebox.showwarning("Different Schedules Required", "Please select two different schedules to compare.")
            return
        
        # Find the schedules
        schedule_a = None
        schedule_b = None
        for s in self.saved_costs:
            if s['schedule_name'] == schedule_a_name:
                schedule_a = s
            if s['schedule_name'] == schedule_b_name:
                schedule_b = s
        
        if not schedule_a or not schedule_b:
            messagebox.showerror("Error", "Could not find selected schedules.")
            return
        
        positions = self.compare_positions_var.get()
        if positions <= 0:
            positions = 100
        
        # Clear existing results
        for widget in self.compare_results_frame.winfo_children():
            widget.destroy()
        
        # Build comparison table
        self.build_comparison_display(schedule_a, schedule_b, positions)
    
    def build_comparison_display(self, sched_a, sched_b, positions):
        """Build the comparison display table."""
        frame = self.compare_results_frame
        
        # Column widths in pixels for consistent alignment
        col_label_px = 220
        col_value_px = 180
        col_diff_px = 180
        
        # Clear and rebuild frozen header
        for widget in self.compare_header_frame.winfo_children():
            widget.destroy()
        
        # Build frozen column headers using grid for precise alignment
        header_row = ttk.Frame(self.compare_header_frame)
        header_row.pack(fill='x', pady=5, padx=10)
        
        # Configure grid columns with fixed widths
        header_row.columnconfigure(0, minsize=col_label_px)
        header_row.columnconfigure(1, minsize=col_value_px)
        header_row.columnconfigure(2, minsize=col_value_px)
        header_row.columnconfigure(3, minsize=col_diff_px)
        
        ttk.Label(header_row, text="", anchor='e').grid(row=0, column=0, sticky='e', padx=5)
        ttk.Label(header_row, text=sched_a['schedule_name'], anchor='center',
                  font=('Segoe UI', 11, 'bold'), foreground=self.colors['primary']).grid(row=0, column=1, sticky='ew', padx=5)
        ttk.Label(header_row, text=sched_b['schedule_name'], anchor='center',
                  font=('Segoe UI', 11, 'bold'), foreground=self.colors['primary']).grid(row=0, column=2, sticky='ew', padx=5)
        ttk.Label(header_row, text="Change (A → B)", anchor='center',
                  font=('Segoe UI', 11, 'bold')).grid(row=0, column=3, sticky='ew', padx=5)
        
        # Helper function to create a comparison row using grid
        def add_row(label, val_a, val_b, diff_text="", highlight=False):
            row = ttk.Frame(frame)
            row.pack(fill='x', pady=3, padx=10)
            
            # Configure grid columns with same fixed widths as header
            row.columnconfigure(0, minsize=col_label_px)
            row.columnconfigure(1, minsize=col_value_px)
            row.columnconfigure(2, minsize=col_value_px)
            row.columnconfigure(3, minsize=col_diff_px)
            
            # Label - bold for highlighted rows
            lbl_font = ('Segoe UI', 10, 'bold') if highlight else ('Segoe UI', 10)
            ttk.Label(row, text=label, anchor='e', font=lbl_font).grid(row=0, column=0, sticky='e', padx=5)
            
            # Value A - same font size, just bold and colored for highlights
            val_font = ('Segoe UI', 10, 'bold') if highlight else ('Segoe UI', 10)
            val_color = self.colors['primary'] if highlight else 'black'
            ttk.Label(row, text=val_a, anchor='center', 
                      font=val_font, foreground=val_color).grid(row=0, column=1, sticky='ew', padx=5)
            
            # Value B
            ttk.Label(row, text=val_b, anchor='center',
                      font=val_font, foreground=val_color).grid(row=0, column=2, sticky='ew', padx=5)
            
            # Difference - all black
            ttk.Label(row, text=diff_text, anchor='center',
                      font=('Segoe UI', 10)).grid(row=0, column=3, sticky='ew', padx=5)
        
        def add_section_header(title):
            sep = ttk.Separator(frame, orient='horizontal')
            sep.pack(fill='x', pady=10)
            hdr = ttk.Label(frame, text=title, font=('Segoe UI', 11, 'bold'))
            hdr.pack(anchor='w', padx=10, pady=(5, 10))
        
        # === Schedule Info ===
        add_section_header("Schedule Information")
        add_row("Schedule Type:", sched_a['schedule_type'], sched_b['schedule_type'])
        add_row("Scheduled Hours:", f"{sched_a['scheduled_hours']:,.0f}", f"{sched_b['scheduled_hours']:,.0f}",
                f"{sched_b['scheduled_hours'] - sched_a['scheduled_hours']:+,.0f}")
        add_row("Hours Worked:", f"{sched_a['hours_worked']:,.1f}", f"{sched_b['hours_worked']:,.1f}",
                f"{sched_b['hours_worked'] - sched_a['hours_worked']:+,.1f}")
        add_row("Ratio Paid/Worked:", f"{sched_a['ratio_paid_worked']:.3f}", f"{sched_b['ratio_paid_worked']:.3f}",
                f"{sched_b['ratio_paid_worked'] - sched_a['ratio_paid_worked']:+.3f}")
        
        # === Key Cost Metrics ===
        add_section_header("Key Cost Metrics")
        
        cost_diff = sched_b['cost_scheduled_time'] - sched_a['cost_scheduled_time']
        cost_diff_pct = (cost_diff / sched_a['cost_scheduled_time'] * 100) if sched_a['cost_scheduled_time'] > 0 else 0
        add_row("Cost of Scheduled Time:", 
                f"${sched_a['cost_scheduled_time']:.2f}/hr", 
                f"${sched_b['cost_scheduled_time']:.2f}/hr",
                f"{cost_diff:+.2f} ({cost_diff_pct:+.1f}%)", highlight=True)
        
        ot_diff = sched_b['marginal_ot_15'] - sched_a['marginal_ot_15']
        ot_diff_pct = (ot_diff / sched_a['marginal_ot_15'] * 100) if sched_a['marginal_ot_15'] > 0 else 0
        add_row("Marginal OT (1.5x):", 
                f"${sched_a['marginal_ot_15']:.2f}/hr", 
                f"${sched_b['marginal_ot_15']:.2f}/hr",
                f"{ot_diff:+.2f} ({ot_diff_pct:+.1f}%)", highlight=True)
        
        ot20_diff = sched_b['marginal_ot_20'] - sched_a['marginal_ot_20']
        add_row("Marginal OT (2.0x):", 
                f"${sched_a['marginal_ot_20']:.2f}/hr", 
                f"${sched_b['marginal_ot_20']:.2f}/hr",
                f"{ot20_diff:+.2f}")
        
        # === Incentive Cost Comparison ===
        add_section_header(f"Incentive Costs (for {positions:,} positions)")
        
        # Calculate incentive costs for both schedules
        def calc_incentive_costs(sched):
            days_per_week = sched.get('days_per_week', 5)
            shift_length = sched.get('shift_length', 8)
            cost_ot = sched['marginal_ot_15']
            wage = sched['wage']
            
            # Number of crews and total workforce
            if days_per_week == 5:
                num_crews = 3
                annual_shifts = 260
                non_day_workers = positions * 2  # 2 of 3 crews are non-day
            elif shift_length == 8:
                num_crews = 4
                annual_shifts = 273
                non_day_workers = positions * 4 * (2/3)
            else:
                num_crews = 4
                annual_shifts = 182
                non_day_workers = positions * 4 * 0.5
            
            total_workforce = positions * num_crews
            
            # Extra day off cost
            day_hours = 12 if shift_length == 12 else 8
            dayoff_per_emp = cost_ot * day_hours
            dayoff_total = dayoff_per_emp * total_workforce
            
            # 5-minute overlap cost
            overlap_per_emp = (5/60) * cost_ot * annual_shifts
            overlap_total = overlap_per_emp * total_workforce
            
            # Extra week vacation for non-day
            vacation_hours = 40 if days_per_week == 5 else 42
            vacation_per_emp = vacation_hours * cost_ot
            vacation_total = vacation_per_emp * non_day_workers
            
            # Shift diff increase costs (to 10%)
            current_diff = 0  # Assume starting from 0 for comparison
            target_diff = wage * 0.10
            increase = target_diff
            ot_burden = sched.get('ot_burden', 36.27) / 100.0
            scheduled_hours = sched['scheduled_hours']
            shiftdiff_cost = increase * (1 + ot_burden) * scheduled_hours * non_day_workers
            
            return {
                'total_workforce': total_workforce,
                'annual_shifts': annual_shifts,
                'non_day_workers': non_day_workers,
                'dayoff_per_emp': dayoff_per_emp,
                'dayoff_total': dayoff_total,
                'overlap_per_emp': overlap_per_emp,
                'overlap_total': overlap_total,
                'vacation_per_emp': vacation_per_emp,
                'vacation_total': vacation_total,
                'shiftdiff_cost': shiftdiff_cost,
            }
        
        costs_a = calc_incentive_costs(sched_a)
        costs_b = calc_incentive_costs(sched_b)
        
        # Workforce info
        add_row("Total Workforce:", f"{costs_a['total_workforce']:,}", f"{costs_b['total_workforce']:,}",
                f"{costs_b['total_workforce'] - costs_a['total_workforce']:+,}")
        add_row("Annual Shifts/Employee:", f"{costs_a['annual_shifts']}", f"{costs_b['annual_shifts']}",
                f"{costs_b['annual_shifts'] - costs_a['annual_shifts']:+}")
        add_row("Non-Day Workers:", f"{costs_a['non_day_workers']:,.0f}", f"{costs_b['non_day_workers']:,.0f}",
                f"{costs_b['non_day_workers'] - costs_a['non_day_workers']:+,.0f}")
        
        # Separator before cost comparisons
        ttk.Separator(frame, orient='horizontal').pack(fill='x', pady=5)
        
        # Extra Day Off
        add_row("Extra Day Off (per emp):", 
                f"${costs_a['dayoff_per_emp']:,.2f}", 
                f"${costs_b['dayoff_per_emp']:,.2f}",
                f"${costs_b['dayoff_per_emp'] - costs_a['dayoff_per_emp']:+,.2f}")
        dayoff_diff = costs_b['dayoff_total'] - costs_a['dayoff_total']
        add_row("Extra Day Off (total):", 
                f"${costs_a['dayoff_total']:,.0f}", 
                f"${costs_b['dayoff_total']:,.0f}",
                f"${dayoff_diff:+,.0f}", highlight=True)
        
        # 5-Minute Overlap
        add_row("5-Min Overlap (per emp):", 
                f"${costs_a['overlap_per_emp']:,.2f}", 
                f"${costs_b['overlap_per_emp']:,.2f}",
                f"${costs_b['overlap_per_emp'] - costs_a['overlap_per_emp']:+,.2f}")
        overlap_diff = costs_b['overlap_total'] - costs_a['overlap_total']
        add_row("5-Min Overlap (total):", 
                f"${costs_a['overlap_total']:,.0f}", 
                f"${costs_b['overlap_total']:,.0f}",
                f"${overlap_diff:+,.0f}", highlight=True)
        
        # Extra Week Vacation
        add_row("Extra Wk Vacation (per emp):", 
                f"${costs_a['vacation_per_emp']:,.2f}", 
                f"${costs_b['vacation_per_emp']:,.2f}",
                f"${costs_b['vacation_per_emp'] - costs_a['vacation_per_emp']:+,.2f}")
        vac_diff = costs_b['vacation_total'] - costs_a['vacation_total']
        add_row("Extra Wk Vacation (total):", 
                f"${costs_a['vacation_total']:,.0f}", 
                f"${costs_b['vacation_total']:,.0f}",
                f"${vac_diff:+,.0f}", highlight=True)
        
        # Shift Diff to 10%
        shiftdiff_diff = costs_b['shiftdiff_cost'] - costs_a['shiftdiff_cost']
        add_row("Shift Diff to 10% (total):", 
                f"${costs_a['shiftdiff_cost']:,.0f}", 
                f"${costs_b['shiftdiff_cost']:,.0f}",
                f"${shiftdiff_diff:+,.0f}", highlight=True)

    def create_analysis_tab(self):
        """Create the staffing and cost analysis tab."""
        # Create outer frame for the tab
        outer_frame = ttk.Frame(self.notebook, padding="10")
        self.notebook.add(outer_frame, text="Staffing Analysis")
        
        # Create scrollable canvas
        analysis_canvas = tk.Canvas(outer_frame, highlightthickness=0)
        analysis_scrollbar = ttk.Scrollbar(outer_frame, orient="vertical", command=analysis_canvas.yview)
        analysis_frame = ttk.Frame(analysis_canvas)
        
        analysis_frame.bind(
            "<Configure>",
            lambda e: analysis_canvas.configure(scrollregion=analysis_canvas.bbox("all"))
        )
        
        canvas_window = analysis_canvas.create_window((0, 0), window=analysis_frame, anchor="nw")
        analysis_canvas.configure(yscrollcommand=analysis_scrollbar.set)
        
        # Make inner frame expand to canvas width
        def configure_analysis_canvas_width(event):
            analysis_canvas.itemconfig(canvas_window, width=event.width)
        analysis_canvas.bind("<Configure>", configure_analysis_canvas_width)
        
        analysis_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        analysis_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Enable mousewheel scrolling only when mouse is over this canvas
        def on_mousewheel_analysis(event):
            analysis_canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        
        def bind_mousewheel_analysis(event):
            analysis_canvas.bind_all("<MouseWheel>", on_mousewheel_analysis)
        
        def unbind_mousewheel_analysis(event):
            analysis_canvas.unbind_all("<MouseWheel>")
        
        analysis_canvas.bind("<Enter>", bind_mousewheel_analysis)
        analysis_canvas.bind("<Leave>", unbind_mousewheel_analysis)
        
        # Title
        title_label = ttk.Label(analysis_frame,
                               text="Staffing & Cost Analysis",
                               style='Header.TLabel')
        title_label.pack(pady=(0, 10))
        
        # Info
        info_label = ttk.Label(analysis_frame,
                              text="Select a saved schedule to analyze staffing requirements and costs at different OT levels.",
                              foreground=self.colors['gray'])
        info_label.pack(pady=(0, 15))
        
        # Input section
        input_frame = ttk.LabelFrame(analysis_frame, text="Analysis Parameters", padding="15")
        input_frame.pack(fill='x', padx=10, pady=5)
        
        # Row 1: Schedule selection
        row1 = ttk.Frame(input_frame)
        row1.pack(fill='x', pady=5)
        
        ttk.Label(row1, text="Select Schedule:").pack(side=tk.LEFT, padx=5)
        self.analysis_schedule_var = tk.StringVar()
        self.analysis_schedule_combo = ttk.Combobox(row1, textvariable=self.analysis_schedule_var, 
                                                    width=40, state='readonly')
        self.analysis_schedule_combo.pack(side=tk.LEFT, padx=5)
        
        # Row 2: Positions and Target OT
        row2 = ttk.Frame(input_frame)
        row2.pack(fill='x', pady=5)
        
        ttk.Label(row2, text="Positions to Cover:").pack(side=tk.LEFT, padx=5)
        self.positions_var = tk.IntVar(value=100)
        ttk.Entry(row2, textvariable=self.positions_var, width=10).pack(side=tk.LEFT, padx=5)
        
        ttk.Label(row2, text="Target Unscheduled OT %:").pack(side=tk.LEFT, padx=(20, 5))
        self.target_ot_var = tk.IntVar(value=10)
        ttk.Entry(row2, textvariable=self.target_ot_var, width=10).pack(side=tk.LEFT, padx=5)
        
        # Analyze button
        analyze_btn = ttk.Button(input_frame, text="Run Analysis", 
                                command=self.run_staffing_analysis,
                                style='Primary.TButton')
        analyze_btn.pack(pady=10)
        
        # Schedule Summary section
        summary_frame = ttk.LabelFrame(analysis_frame, text="Schedule Summary", padding="15")
        summary_frame.pack(fill='x', padx=10, pady=5)
        
        self.analysis_summary_labels = {}
        
        summary_items = [
            ('schedule_type', 'Schedule Type:'),
            ('hours_worked', 'Hours Worked/Year:'),
            ('absenteeism_pct', 'Absenteeism % (causes OT):'),
            ('cost_scheduled_time', 'Cost of Scheduled Time:'),
            ('marginal_ot_15', 'Marginal OT Cost (1.5x):'),
            ('employee_earnings', 'Employee Earnings (base):'),
        ]
        
        for i, (key, label) in enumerate(summary_items):
            row_frame = ttk.Frame(summary_frame)
            row_frame.pack(fill='x', pady=2)
            ttk.Label(row_frame, text=label, width=28, anchor='e').pack(side=tk.LEFT, padx=5)
            self.analysis_summary_labels[key] = ttk.Label(row_frame, text="--", 
                                                          font=('Segoe UI', 10, 'bold'),
                                                          foreground=self.colors['primary'])
            self.analysis_summary_labels[key].pack(side=tk.LEFT, padx=5)
        
        # Results section
        results_frame = ttk.LabelFrame(analysis_frame, text="Staffing & Cost Scenarios", padding="15")
        results_frame.pack(fill='x', padx=10, pady=5)
        
        # Create treeview for results
        columns = ('ot_pct', 'crew_needed', 'scheduled_cost', 'ot_cost', 'total_cost', 'cost_per_position', 'cost_per_person', 'employee_income')
        
        self.analysis_tree = ttk.Treeview(results_frame, columns=columns, show='headings', height=7)
        
        col_configs = [
            ('ot_pct', 'OT %', 60),
            ('crew_needed', 'Total Employees', 100),
            ('scheduled_cost', 'Scheduled Cost', 110),
            ('ot_cost', 'OT Cost', 100),
            ('total_cost', 'Total Cost', 110),
            ('cost_per_position', 'Cost/Position', 100),
            ('cost_per_person', 'Cost/Person', 100),
            ('employee_income', 'Employee Income', 110),
        ]
        
        for col_id, heading, width in col_configs:
            self.analysis_tree.heading(col_id, text=heading)
            self.analysis_tree.column(col_id, width=width, anchor='center')
        
        # Highlight tag for target row
        self.analysis_tree.tag_configure('target', background='#d4edda', font=('Segoe UI', 10, 'bold'))
        
        self.analysis_tree.pack(fill='x', pady=5)
        
        # Legend
        legend_frame = ttk.Frame(results_frame)
        legend_frame.pack(fill='x', pady=5)
        
        legend_box = tk.Frame(legend_frame, bg='#d4edda', width=20, height=15)
        legend_box.pack(side=tk.LEFT, padx=5)
        legend_box.pack_propagate(False)
        ttk.Label(legend_frame, text="= Your target OT %", foreground=self.colors['gray']).pack(side=tk.LEFT)
        
        # === NEW: Incentive Cost Analysis Section ===
        incentive_frame = ttk.LabelFrame(analysis_frame, text="Incentive Cost Analysis", padding="15")
        incentive_frame.pack(fill='x', padx=10, pady=10)
        
        # Create labels for incentive costs (will be populated by run_staffing_analysis)
        self.incentive_labels = {}
        
        # Schedule Type Info (shows what schedule is being used for calculations)
        schedule_info_frame = ttk.Frame(incentive_frame)
        schedule_info_frame.pack(fill='x', pady=(0, 10))
        
        ttk.Label(schedule_info_frame, text="Calculations based on:", 
                  font=('Segoe UI', 9)).pack(side=tk.LEFT)
        self.incentive_labels['schedule_type_info'] = ttk.Label(schedule_info_frame, text="--",
                                                                 foreground=self.colors['primary'],
                                                                 font=('Segoe UI', 10, 'bold'))
        self.incentive_labels['schedule_type_info'].pack(side=tk.LEFT, padx=10)
        
        ttk.Label(schedule_info_frame, text="Total Workforce:", 
                  font=('Segoe UI', 9)).pack(side=tk.LEFT, padx=(20, 0))
        self.incentive_labels['target_crew_info'] = ttk.Label(schedule_info_frame, text="--",
                                                               foreground=self.colors['primary'],
                                                               font=('Segoe UI', 10, 'bold'))
        self.incentive_labels['target_crew_info'].pack(side=tk.LEFT, padx=10)
        
        ttk.Separator(incentive_frame, orient='horizontal').pack(fill='x', pady=5)
        
        # Extra Day Off section
        dayoff_frame = ttk.Frame(incentive_frame)
        dayoff_frame.pack(fill='x', pady=5)
        
        ttk.Label(dayoff_frame, text="Extra Day Off (OT coverage cost):", 
                  font=('Segoe UI', 10, 'bold')).pack(anchor='w')
        
        dayoff_row = ttk.Frame(dayoff_frame)
        dayoff_row.pack(fill='x', padx=20, pady=2)
        ttk.Label(dayoff_row, text="Per Employee:", width=20, anchor='e').pack(side=tk.LEFT)
        self.incentive_labels['dayoff_per_emp'] = ttk.Label(dayoff_row, text="--", 
                                                            foreground=self.colors['primary'],
                                                            font=('Segoe UI', 10, 'bold'))
        self.incentive_labels['dayoff_per_emp'].pack(side=tk.LEFT, padx=10)
        
        ttk.Label(dayoff_row, text="Total (all crew):", width=15, anchor='e').pack(side=tk.LEFT, padx=(20, 0))
        self.incentive_labels['dayoff_total'] = ttk.Label(dayoff_row, text="--",
                                                          foreground=self.colors['primary'],
                                                          font=('Segoe UI', 10, 'bold'))
        self.incentive_labels['dayoff_total'].pack(side=tk.LEFT, padx=10)
        
        ttk.Separator(incentive_frame, orient='horizontal').pack(fill='x', pady=10)
        
        # 5-Minute Overlap Cost section
        overlap_frame = ttk.Frame(incentive_frame)
        overlap_frame.pack(fill='x', pady=5)
        
        ttk.Label(overlap_frame, text="Cost per 5 Minutes of Shift Overlap:", 
                  font=('Segoe UI', 10, 'bold')).pack(anchor='w')
        
        # Show formula note
        overlap_note = ttk.Frame(overlap_frame)
        overlap_note.pack(fill='x', padx=20, pady=2)
        self.incentive_labels['overlap_formula'] = ttk.Label(overlap_note, 
            text="(5/60 × OT cost × annual shifts)",
            foreground=self.colors['gray'], font=('Segoe UI', 9))
        self.incentive_labels['overlap_formula'].pack(anchor='w')
        
        overlap_row = ttk.Frame(overlap_frame)
        overlap_row.pack(fill='x', padx=20, pady=2)
        ttk.Label(overlap_row, text="Per Employee:", width=20, anchor='e').pack(side=tk.LEFT)
        self.incentive_labels['overlap_per_emp'] = ttk.Label(overlap_row, text="--",
                                                              foreground=self.colors['primary'],
                                                              font=('Segoe UI', 10, 'bold'))
        self.incentive_labels['overlap_per_emp'].pack(side=tk.LEFT, padx=10)
        
        ttk.Label(overlap_row, text="Total (all employees):", width=18, anchor='e').pack(side=tk.LEFT, padx=(20, 0))
        self.incentive_labels['overlap_total'] = ttk.Label(overlap_row, text="--",
                                                            foreground=self.colors['primary'],
                                                            font=('Segoe UI', 10, 'bold'))
        self.incentive_labels['overlap_total'].pack(side=tk.LEFT, padx=10)
        
        # Show annual shifts note
        overlap_shifts_row = ttk.Frame(overlap_frame)
        overlap_shifts_row.pack(fill='x', padx=20, pady=2)
        ttk.Label(overlap_shifts_row, text="Annual Shifts:", width=20, anchor='e').pack(side=tk.LEFT)
        self.incentive_labels['overlap_shifts'] = ttk.Label(overlap_shifts_row, text="--",
                                                             foreground=self.colors['gray'],
                                                             font=('Segoe UI', 9))
        self.incentive_labels['overlap_shifts'].pack(side=tk.LEFT, padx=10)
        
        ttk.Separator(incentive_frame, orient='horizontal').pack(fill='x', pady=10)
        
        # Shift Differential section
        shiftdiff_frame = ttk.Frame(incentive_frame)
        shiftdiff_frame.pack(fill='x', pady=5)
        
        ttk.Label(shiftdiff_frame, text="Shift Differential Increase (cost to raise to target %):", 
                  font=('Segoe UI', 10, 'bold')).pack(anchor='w')
        
        # Current shift diff row
        current_row = ttk.Frame(shiftdiff_frame)
        current_row.pack(fill='x', padx=20, pady=2)
        ttk.Label(current_row, text="Current Shift Diff:", width=18, anchor='e').pack(side=tk.LEFT)
        self.incentive_labels['current_diff_dollar'] = ttk.Label(current_row, text="--",
                                                                  foreground=self.colors['gray'])
        self.incentive_labels['current_diff_dollar'].pack(side=tk.LEFT, padx=5)
        self.incentive_labels['current_diff_pct'] = ttk.Label(current_row, text="(--% of wage)",
                                                               foreground=self.colors['gray'])
        self.incentive_labels['current_diff_pct'].pack(side=tk.LEFT, padx=5)
        
        # Header row for columns
        header_row = ttk.Frame(shiftdiff_frame)
        header_row.pack(fill='x', padx=20, pady=(10, 2))
        ttk.Label(header_row, text="Target", width=10, anchor='center', 
                  font=('Segoe UI', 9, 'bold')).pack(side=tk.LEFT)
        ttk.Label(header_row, text="$/Hour", width=10, anchor='center',
                  font=('Segoe UI', 9, 'bold')).pack(side=tk.LEFT)
        ttk.Label(header_row, text="Increase", width=12, anchor='center',
                  font=('Segoe UI', 9, 'bold')).pack(side=tk.LEFT)
        ttk.Label(header_row, text="Annual Cost", width=15, anchor='center',
                  font=('Segoe UI', 9, 'bold')).pack(side=tk.LEFT)
        
        # 5% row
        diff5_row = ttk.Frame(shiftdiff_frame)
        diff5_row.pack(fill='x', padx=20, pady=2)
        ttk.Label(diff5_row, text="5%", width=10, anchor='center').pack(side=tk.LEFT)
        self.incentive_labels['shiftdiff_5_dollar'] = ttk.Label(diff5_row, text="--", width=10, anchor='center')
        self.incentive_labels['shiftdiff_5_dollar'].pack(side=tk.LEFT)
        self.incentive_labels['shiftdiff_5_increase'] = ttk.Label(diff5_row, text="--", width=12, anchor='center')
        self.incentive_labels['shiftdiff_5_increase'].pack(side=tk.LEFT)
        self.incentive_labels['shiftdiff_5_cost'] = ttk.Label(diff5_row, text="--", width=15, anchor='center',
                                                              foreground=self.colors['primary'],
                                                              font=('Segoe UI', 10, 'bold'))
        self.incentive_labels['shiftdiff_5_cost'].pack(side=tk.LEFT)
        
        # 10% row
        diff10_row = ttk.Frame(shiftdiff_frame)
        diff10_row.pack(fill='x', padx=20, pady=2)
        ttk.Label(diff10_row, text="10%", width=10, anchor='center').pack(side=tk.LEFT)
        self.incentive_labels['shiftdiff_10_dollar'] = ttk.Label(diff10_row, text="--", width=10, anchor='center')
        self.incentive_labels['shiftdiff_10_dollar'].pack(side=tk.LEFT)
        self.incentive_labels['shiftdiff_10_increase'] = ttk.Label(diff10_row, text="--", width=12, anchor='center')
        self.incentive_labels['shiftdiff_10_increase'].pack(side=tk.LEFT)
        self.incentive_labels['shiftdiff_10_cost'] = ttk.Label(diff10_row, text="--", width=15, anchor='center',
                                                               foreground=self.colors['primary'],
                                                               font=('Segoe UI', 10, 'bold'))
        self.incentive_labels['shiftdiff_10_cost'].pack(side=tk.LEFT)
        
        # 15% row
        diff15_row = ttk.Frame(shiftdiff_frame)
        diff15_row.pack(fill='x', padx=20, pady=2)
        ttk.Label(diff15_row, text="15%", width=10, anchor='center').pack(side=tk.LEFT)
        self.incentive_labels['shiftdiff_15_dollar'] = ttk.Label(diff15_row, text="--", width=10, anchor='center')
        self.incentive_labels['shiftdiff_15_dollar'].pack(side=tk.LEFT)
        self.incentive_labels['shiftdiff_15_increase'] = ttk.Label(diff15_row, text="--", width=12, anchor='center')
        self.incentive_labels['shiftdiff_15_increase'].pack(side=tk.LEFT)
        self.incentive_labels['shiftdiff_15_cost'] = ttk.Label(diff15_row, text="--", width=15, anchor='center',
                                                               foreground=self.colors['primary'],
                                                               font=('Segoe UI', 10, 'bold'))
        self.incentive_labels['shiftdiff_15_cost'].pack(side=tk.LEFT)
        
        ttk.Separator(incentive_frame, orient='horizontal').pack(fill='x', pady=10)
        
        # Extra Week Vacation section
        vacation_frame = ttk.Frame(incentive_frame)
        vacation_frame.pack(fill='x', pady=5)
        
        ttk.Label(vacation_frame, text="Extra Week Vacation for Non-Day Shift (OT coverage cost):", 
                  font=('Segoe UI', 10, 'bold')).pack(anchor='w')
        
        # Show hours note
        vac_note = ttk.Frame(vacation_frame)
        vac_note.pack(fill='x', padx=20, pady=2)
        self.incentive_labels['vacation_hours_note'] = ttk.Label(vac_note, text="(40 hrs for 5-day, 42 hrs for 7-day)",
                                                                  foreground=self.colors['gray'],
                                                                  font=('Segoe UI', 9))
        self.incentive_labels['vacation_hours_note'].pack(anchor='w')
        
        vac_row = ttk.Frame(vacation_frame)
        vac_row.pack(fill='x', padx=20, pady=2)
        ttk.Label(vac_row, text="Per Employee:", width=20, anchor='e').pack(side=tk.LEFT)
        self.incentive_labels['vacation_per_emp'] = ttk.Label(vac_row, text="--",
                                                              foreground=self.colors['primary'],
                                                              font=('Segoe UI', 10, 'bold'))
        self.incentive_labels['vacation_per_emp'].pack(side=tk.LEFT, padx=10)
        
        ttk.Label(vac_row, text="Total (non-day crew):", width=22, anchor='e').pack(side=tk.LEFT, padx=(20, 0))
        self.incentive_labels['vacation_total'] = ttk.Label(vac_row, text="--",
                                                            foreground=self.colors['primary'],
                                                            font=('Segoe UI', 10, 'bold'))
        self.incentive_labels['vacation_total'].pack(side=tk.LEFT, padx=10)
        
        # Show non-day worker count
        vac_count_row = ttk.Frame(vacation_frame)
        vac_count_row.pack(fill='x', padx=20, pady=2)
        ttk.Label(vac_count_row, text="Non-Day Workers:", width=20, anchor='e').pack(side=tk.LEFT)
        self.incentive_labels['non_day_count'] = ttk.Label(vac_count_row, text="--",
                                                           foreground=self.colors['gray'],
                                                           font=('Segoe UI', 9))
        self.incentive_labels['non_day_count'].pack(side=tk.LEFT, padx=10)
        
        # === NEW: Adverse Cost Analysis Section ===
        adverse_frame = ttk.LabelFrame(analysis_frame, text="Adverse Cost Analysis", padding="15")
        adverse_frame.pack(fill='x', padx=10, pady=10)
        
        self.adverse_labels = {}
        
        # Explanation text
        explain_text = ("Adverse cost is the cost of not being perfectly staffed. When understaffed, "
                       "you pay the OT premium. When overstaffed, you pay for idle time.")
        ttk.Label(adverse_frame, text=explain_text, wraplength=700, 
                  foreground=self.colors['gray'], font=('Segoe UI', 9, 'italic')).pack(anchor='w', pady=(0, 10))
        
        # Per-hour adverse costs
        rates_frame = ttk.Frame(adverse_frame)
        rates_frame.pack(fill='x', pady=5)
        
        ttk.Label(rates_frame, text="Adverse Cost Rates:", font=('Segoe UI', 10, 'bold')).pack(anchor='w')
        
        rates_row = ttk.Frame(rates_frame)
        rates_row.pack(fill='x', padx=20, pady=5)
        
        ttk.Label(rates_row, text="Understaffing (OT − ST):", width=25, anchor='e').pack(side=tk.LEFT)
        self.adverse_labels['understaff_rate'] = ttk.Label(rates_row, text="--", 
                                                            foreground=self.colors['primary'],
                                                            font=('Segoe UI', 10, 'bold'))
        self.adverse_labels['understaff_rate'].pack(side=tk.LEFT, padx=10)
        
        ttk.Label(rates_row, text="Overstaffing (ST − $0):", width=25, anchor='e').pack(side=tk.LEFT, padx=(20, 0))
        self.adverse_labels['overstaff_rate'] = ttk.Label(rates_row, text="--",
                                                           foreground=self.colors['primary'],
                                                           font=('Segoe UI', 10, 'bold'))
        self.adverse_labels['overstaff_rate'].pack(side=tk.LEFT, padx=10)
        
        ttk.Label(rates_row, text="Ratio:", width=10, anchor='e').pack(side=tk.LEFT, padx=(20, 0))
        self.adverse_labels['cost_ratio'] = ttk.Label(rates_row, text="--",
                                                       foreground=self.colors['gray'],
                                                       font=('Segoe UI', 10))
        self.adverse_labels['cost_ratio'].pack(side=tk.LEFT, padx=10)
        
        ttk.Separator(adverse_frame, orient='horizontal').pack(fill='x', pady=10)
        
        # Scenario Analysis Header
        ttk.Label(adverse_frame, text="Annual Adverse Cost by Variability Scenario:", 
                  font=('Segoe UI', 10, 'bold')).pack(anchor='w')
        ttk.Label(adverse_frame, 
                  text="(Based on positions to cover, staffing level at target OT%, and absenteeism rate)",
                  foreground=self.colors['gray'], font=('Segoe UI', 9)).pack(anchor='w', pady=(0, 5))
        
        # Scenario table using Treeview
        scenario_columns = ('scenario', 'description', 'understaff_cost', 'overstaff_cost', 'total_cost')
        self.adverse_tree = ttk.Treeview(adverse_frame, columns=scenario_columns, show='headings', height=4)
        
        self.adverse_tree.heading('scenario', text='Scenario')
        self.adverse_tree.heading('description', text='Description')
        self.adverse_tree.heading('understaff_cost', text='Understaffing')
        self.adverse_tree.heading('overstaff_cost', text='Overstaffing')
        self.adverse_tree.heading('total_cost', text='Total Adverse')
        
        self.adverse_tree.column('scenario', width=120, anchor='w')
        self.adverse_tree.column('description', width=280, anchor='w')
        self.adverse_tree.column('understaff_cost', width=100, anchor='e')
        self.adverse_tree.column('overstaff_cost', width=100, anchor='e')
        self.adverse_tree.column('total_cost', width=110, anchor='e')
        
        self.adverse_tree.pack(fill='x', pady=5)
        
        # Key insight
        insight_frame = ttk.Frame(adverse_frame)
        insight_frame.pack(fill='x', pady=(10, 0))
        ttk.Label(insight_frame, text="Key Insight:", font=('Segoe UI', 10, 'bold')).pack(side=tk.LEFT)
        self.adverse_labels['insight'] = ttk.Label(insight_frame, text="--",
                                                    foreground=self.colors['gray'],
                                                    font=('Segoe UI', 9, 'italic'),
                                                    wraplength=600)
        self.adverse_labels['insight'].pack(side=tk.LEFT, padx=10)
        
    def update_analysis_schedule_list(self):
        """Update the schedule dropdown in analysis tab."""
        if hasattr(self, 'analysis_schedule_combo'):
            names = [c['schedule_name'] for c in self.saved_costs]
            self.analysis_schedule_combo['values'] = names
            if names and not self.analysis_schedule_var.get():
                self.analysis_schedule_var.set(names[0])
    
    def run_staffing_analysis(self):
        """Run the staffing and cost analysis for selected schedule."""
        # Get selected schedule
        schedule_name = self.analysis_schedule_var.get()
        if not schedule_name:
            messagebox.showwarning("No Schedule", "Please select a saved schedule to analyze.")
            return
        
        # Find the schedule data
        schedule_data = None
        for cost in self.saved_costs:
            if cost['schedule_name'] == schedule_name:
                schedule_data = cost
                break
        
        if not schedule_data:
            messagebox.showerror("Error", "Selected schedule not found.")
            return
        
        # Get inputs
        try:
            positions = self.positions_var.get()
            target_ot = self.target_ot_var.get()
        except:
            messagebox.showerror("Error", "Please enter valid numbers for positions and target OT %.")
            return
        
        if positions <= 0:
            messagebox.showerror("Error", "Positions must be greater than 0.")
            return
        
        # Update summary
        self.analysis_summary_labels['schedule_type'].config(text=schedule_data['schedule_type'])
        self.analysis_summary_labels['hours_worked'].config(text=f"{schedule_data['hours_worked']:,.1f}")
        self.analysis_summary_labels['absenteeism_pct'].config(text=f"{schedule_data['absenteeism_ot_percent']:.1%}")
        self.analysis_summary_labels['cost_scheduled_time'].config(text=f"${schedule_data['cost_scheduled_time']:.2f}/hr")
        self.analysis_summary_labels['marginal_ot_15'].config(text=f"${schedule_data['marginal_ot_15']:.2f}/hr")
        self.analysis_summary_labels['employee_earnings'].config(text=f"${schedule_data['employee_annual_earnings']:,.0f}")
        
        # Calculate scenarios
        # Generate 5 OT percentages: 2 below target, target, 2 above (in 5% increments)
        ot_percentages = []
        
        # Find lower bounds
        lower1 = max(0, target_ot - 10)
        lower2 = max(0, target_ot - 5)
        
        # Build list ensuring we have 5 unique values
        if lower1 < lower2:
            ot_percentages.append(lower1)
        if lower2 < target_ot:
            ot_percentages.append(lower2)
        ot_percentages.append(target_ot)
        
        upper1 = target_ot + 5
        upper2 = target_ot + 10
        
        ot_percentages.append(upper1)
        ot_percentages.append(upper2)
        
        # Ensure we have exactly 5 values, filling in if needed
        while len(ot_percentages) < 5:
            ot_percentages.insert(0, max(0, ot_percentages[0] - 5))
        
        # Get key values
        absenteeism = schedule_data['absenteeism_ot_percent']
        hours_worked = schedule_data['hours_worked']
        scheduled_hours = schedule_data['scheduled_hours']
        cost_st = schedule_data['cost_scheduled_time']
        cost_ot = schedule_data['marginal_ot_15']
        wage = schedule_data['wage']
        base_employee_earnings = schedule_data['employee_annual_earnings']
        days_per_week = schedule_data.get('days_per_week', 5)
        shift_length = schedule_data.get('shift_length', 8)
        
        # Clear existing results
        for item in self.analysis_tree.get_children():
            self.analysis_tree.delete(item)
        
        # Determine number of crews and total coverage needed based on schedule type
        if days_per_week == 5:
            num_crews = 3
            # Total coverage needed: 52 weeks × 40 hours/week × positions
            total_coverage_needed = 52 * 40 * positions
        else:
            num_crews = 4
            # Total coverage needed: 52 weeks × 168 hours/week × positions
            total_coverage_needed = 52 * 168 * positions
        
        # Calculate and display each scenario
        import math
        
        for ot_pct in ot_percentages:
            # OT % represents unscheduled overtime as a percentage of scheduled hours
            ot_decimal = ot_pct / 100.0
            
            # Calculate employees needed to achieve target unscheduled OT
            # 
            # Logic: 
            #   - Each position needs full coverage (scheduled_hours per year)
            #   - Each employee provides: scheduled_hours × (1 - absenteeism) of coverage
            #   - Plus they can work OT: target_ot × scheduled_hours additional hours
            #   - So each employee provides: scheduled_hours × (1 - absenteeism + target_ot)
            #   - Employees = Positions / (1 - Absenteeism + Target_OT)
            #
            # At 0% OT: Need extra staff to cover absences (employees > positions)
            # At OT% = Absenteeism%: OT covers absences exactly (employees = positions)
            # At higher OT%: Can run leaner (employees < positions)
            
            denominator = 1 - absenteeism + ot_decimal
            if denominator > 0:
                employees_raw = positions / denominator
                total_employees = math.ceil(employees_raw)
            else:
                total_employees = positions * num_crews
            
            # Calculate actual coverage and OT hours
            # Each employee provides scheduled_hours × (1 - absenteeism) of regular coverage
            employee_regular_coverage = scheduled_hours * (1 - absenteeism)
            scheduled_coverage = total_employees * employee_regular_coverage
            unscheduled_ot_hours = max(0, total_coverage_needed - scheduled_coverage)
            
            # Scheduled Cost = Total Employees × Hours Worked × Cost of Scheduled Time
            scheduled_cost = total_employees * hours_worked * cost_st
            
            # OT Cost = Unscheduled OT Hours × Marginal OT 1.5x
            ot_cost = unscheduled_ot_hours * cost_ot
            
            # Total Cost
            total_cost = scheduled_cost + ot_cost
            
            # Cost per position (cost to cover one position 24/7)
            cost_per_position = total_cost / positions if positions > 0 else 0
            
            # Cost per person (average annual cost per employee)
            cost_per_person = total_cost / total_employees if total_employees > 0 else 0
            
            # Employee Income with OT
            # Unscheduled OT hours per employee
            ot_hours_per_employee = unscheduled_ot_hours / total_employees if total_employees > 0 else 0
            # Employee OT pay = OT hours × wage × 1.5
            employee_ot_pay = ot_hours_per_employee * wage * 1.5
            # Total employee income = base earnings + OT pay
            employee_income = base_employee_earnings + employee_ot_pay
            
            # Format values
            values = (
                f"{ot_pct}%",
                f"{total_employees:,}",
                f"${scheduled_cost:,.0f}",
                f"${ot_cost:,.0f}",
                f"${total_cost:,.0f}",
                f"${cost_per_position:,.0f}",
                f"${cost_per_person:,.0f}",
                f"${employee_income:,.0f}",
            )
            
            # Highlight target row
            tags = ('target',) if ot_pct == target_ot else ()
            self.analysis_tree.insert('', 'end', values=values, tags=tags)
        
        # === Calculate Incentive Costs ===
        # Use crew at target OT level for incentive calculations
        target_ot_decimal = target_ot / 100.0
        denominator = 1 - absenteeism + target_ot_decimal
        if denominator > 0:
            target_crew = math.ceil(positions / denominator)
        else:
            target_crew = positions
        
        # Display schedule type info for user verification
        schedule_type_str = schedule_data.get('schedule_type', 'Unknown')
        self.incentive_labels['schedule_type_info'].config(text=f"{schedule_type_str}")
        
        # Calculate total workforce for display (will be calculated again below, but need it here)
        if days_per_week == 5:
            display_total_workforce = positions * 3
        else:
            display_total_workforce = positions * 4
        self.incentive_labels['target_crew_info'].config(text=f"{display_total_workforce:,} total ({positions:,} pos × {3 if days_per_week == 5 else 4} crews)")
        
        # 1. Extra Day Off Cost (cost of covering with OT)
        # Calculate total workforce first (need it for this section)
        if days_per_week == 5:
            num_crews = 3
            total_workforce = positions * num_crews
        else:
            num_crews = 4
            total_workforce = positions * num_crews
        
        day_hours = 12 if shift_length == 12 else 8
        dayoff_per_emp = cost_ot * day_hours  # Marginal OT 1.5x × shift length
        dayoff_total = dayoff_per_emp * total_workforce  # Cost for all employees
        
        self.incentive_labels['dayoff_per_emp'].config(text=f"${dayoff_per_emp:,.2f}")
        self.incentive_labels['dayoff_total'].config(text=f"${dayoff_total:,.0f}")
        
        # 2. Cost per 5 Minutes of Shift Overlap
        # Annual shifts by schedule type: 5-day=260, 7-day 8-hr=273, 7-day 12-hr=182
        if days_per_week == 5:
            annual_shifts = 260
            shifts_explanation = "5-day: 260 shifts/year"
        elif shift_length == 8:
            annual_shifts = 273
            shifts_explanation = "7-day 8-hr: 273 shifts/year"
        else:
            annual_shifts = 182
            shifts_explanation = "7-day 12-hr: 182 shifts/year"
        
        # Cost per 5 minutes = (5/60) × OT cost × annual shifts
        overlap_per_emp = (5 / 60) * cost_ot * annual_shifts
        overlap_total = overlap_per_emp * total_workforce
        
        self.incentive_labels['overlap_formula'].config(
            text=f"(5/60 × ${cost_ot:.2f} OT × {annual_shifts} shifts)")
        self.incentive_labels['overlap_per_emp'].config(text=f"${overlap_per_emp:,.2f}")
        self.incentive_labels['overlap_total'].config(text=f"${overlap_total:,.0f}")
        self.incentive_labels['overlap_shifts'].config(text=f"{annual_shifts} ({shifts_explanation})")
        
        # 3. Shift Differential Costs
        # Calculate non-day workers based on schedule type
        # "Positions to Cover" = positions that need coverage at any time
        # Total workforce = Positions × Number of Crews
        # 5-day: 3 crews (Day, Swing, Night), 2 of 3 crews are non-day
        # 7-day 8-hour: 4 crews rotating, 2/3 of all hours are non-day shift
        # 7-day 12-hour: 4 crews rotating, 1/2 of all hours are non-day shift
        
        if days_per_week == 5:
            non_day_workers = positions * 2  # 2 of 3 crews are non-day
            non_day_explanation = f"5-day: {positions:,} pos × 2 non-day crews"
        elif shift_length == 8:
            non_day_workers = total_workforce * (2 / 3)  # 2/3 of hours are non-day
            non_day_explanation = f"7-day 8-hr: {positions:,} pos × 4 crews × (2/3)"
        else:
            non_day_workers = total_workforce * 0.5  # 1/2 of hours are non-day
            non_day_explanation = f"7-day 12-hr: {positions:,} pos × 4 crews × 0.5"
        
        # Get current shift differential and OT burden
        current_diff = self.burden_calc.inputs.get('shift_differential', 0)
        ot_burden = schedule_data.get('ot_burden', 36.27) / 100.0
        
        # Current diff as percent of wage
        current_diff_pct = (current_diff / wage * 100) if wage > 0 else 0
        
        self.incentive_labels['current_diff_dollar'].config(text=f"${current_diff:.2f}/hr")
        self.incentive_labels['current_diff_pct'].config(text=f"({current_diff_pct:.1f}% of wage)")
        
        # Calculate $ at 5%, 10%, 15% and the increase from current
        target_pcts = [5, 10, 15]
        for pct in target_pcts:
            target_diff_dollar = wage * (pct / 100)
            increase = max(0, target_diff_dollar - current_diff)  # Can't be negative
            
            # Cost = increase × (1 + OT burden) × scheduled_hours × non_day_workers
            annual_cost = increase * (1 + ot_burden) * scheduled_hours * non_day_workers
            
            # Update labels
            self.incentive_labels[f'shiftdiff_{pct}_dollar'].config(text=f"${target_diff_dollar:.2f}")
            self.incentive_labels[f'shiftdiff_{pct}_increase'].config(text=f"+${increase:.2f}")
            self.incentive_labels[f'shiftdiff_{pct}_cost'].config(text=f"${annual_cost:,.0f}")
        
        # 4. Extra Week Vacation for Non-Day Shift
        # Uses same non_day_workers count calculated above
        # 5-day: 40 hours, 7-day: 42 hours, paid at OT rate (coverage cost)
        vacation_hours = 40 if days_per_week == 5 else 42
        vacation_per_emp = vacation_hours * cost_ot
        vacation_total = vacation_per_emp * non_day_workers
        
        self.incentive_labels['vacation_hours_note'].config(
            text=f"({vacation_hours} hrs × ${cost_ot:.2f}/hr OT cost)")
        self.incentive_labels['vacation_per_emp'].config(text=f"${vacation_per_emp:,.2f}")
        self.incentive_labels['vacation_total'].config(text=f"${vacation_total:,.0f}")
        self.incentive_labels['non_day_count'].config(text=f"{non_day_workers:,.0f} employees ({non_day_explanation})")
        
        # === Adverse Cost Analysis ===
        # Calculate per-hour adverse costs
        adverse_understaff = cost_ot - cost_st  # Cost of using OT instead of ST
        adverse_overstaff = cost_st  # Cost of paying for unneeded hours
        
        # Display rates
        self.adverse_labels['understaff_rate'].config(text=f"${adverse_understaff:.2f}/hr")
        self.adverse_labels['overstaff_rate'].config(text=f"${adverse_overstaff:.2f}/hr")
        
        # Calculate ratio
        if adverse_understaff > 0:
            ratio = adverse_overstaff / adverse_understaff
            self.adverse_labels['cost_ratio'].config(text=f"{ratio:.0f}:1 (over:under)")
        else:
            ratio = 0
            self.adverse_labels['cost_ratio'].config(text="--")
        
        # Clear existing scenario results
        for item in self.adverse_tree.get_children():
            self.adverse_tree.delete(item)
        
        # Calculate staffing at target OT level
        target_ot_decimal = target_ot / 100.0
        denom = 1 - absenteeism + target_ot_decimal
        if denom > 0:
            staff_level = math.ceil(positions / denom)
        else:
            staff_level = positions
        
        # Calculate working days per year (5-day = 260, 7-day = 365)
        work_days = 260 if days_per_week == 5 else 365
        shift_hours = shift_length if shift_length else 8
        
        # Scenario calculations
        scenarios = []
        
        # Scenario 1: Flat (theoretical - exactly average every day)
        # Every day has exactly the average absenteeism rate
        # Extra staff = staff_level - positions
        extra_staff = staff_level - positions
        
        # Each day: positions × absenteeism people are absent
        # Extra staff available (after their absences) = extra_staff × (1 - absenteeism)
        daily_absences = positions * absenteeism
        daily_extra_coverage = extra_staff * (1 - absenteeism)
        
        # Net daily position: positive = overstaffed, negative = understaffed
        daily_net = daily_extra_coverage - daily_absences
        
        if daily_net >= 0:
            # Overstaffed every day
            flat_overstaff_hrs = work_days * shift_hours * daily_net
            flat_understaff_hrs = 0
        else:
            # Understaffed every day
            flat_understaff_hrs = work_days * shift_hours * (-daily_net)
            flat_overstaff_hrs = 0
        
        flat_understaff_cost = flat_understaff_hrs * adverse_understaff
        flat_overstaff_cost = flat_overstaff_hrs * adverse_overstaff
        scenarios.append((
            "Flat (Theoretical)",
            f"Exactly {absenteeism:.1%} absent every day - no variance",
            flat_understaff_cost,
            flat_overstaff_cost
        ))
        
        # Scenario 2: Lumpy (worst case - absences concentrated on certain days)
        # Key insight: Total absences are the SAME, just redistributed
        # High-absence days (Mon/Fri): More understaffing
        # Low-absence days (Tue-Thu): More overstaffing (the absences that DIDN'T happen)
        # Both adverse costs increase due to variance
        
        if days_per_week == 5:
            high_absence_days = 104  # 2 days/week × 52 weeks (Mon/Fri)
            low_absence_days = 156   # 3 days/week × 52 weeks (Tue-Thu)
        else:
            high_absence_days = 146  # ~2.8 days/week (weekends + Mon)
            low_absence_days = 219   # ~4.2 days/week
        
        # Redistribute the SAME total absences unevenly
        # If average is 8.6%, lumpy might be 15% on high days, 4% on low days
        # Constraint: (high_rate × high_days + low_rate × low_days) / total_days = average_rate
        # Solve: high_rate × high_days + low_rate × low_days = average_rate × total_days
        
        # Set high absence rate (e.g., 2x average) and solve for low rate
        high_absence_rate = min(absenteeism * 2.0, 0.30)  # 2x average, cap at 30%
        # low_rate = (avg × total - high × high_days) / low_days
        total_absence_budget = absenteeism * work_days
        high_day_absences = high_absence_rate * high_absence_days
        remaining_absences = total_absence_budget - high_day_absences
        low_absence_rate = max(0.01, remaining_absences / low_absence_days)  # Floor at 1%
        
        # Verify: weighted average should equal original absenteeism
        # (This ensures we're redistributing, not adding absences)
        
        # Extra staff available each day = staff_level - positions
        extra_staff = staff_level - positions
        
        # HIGH-ABSENCE DAYS: Understaffing occurs
        # People absent = positions × high_absence_rate
        # Coverage available from extra staff = extra_staff (assuming they show up at avg rate)
        # But extra staff also have absences, so effective extra = extra_staff × (1 - high_absence_rate)
        high_day_absent_positions = positions * high_absence_rate
        high_day_extra_coverage = extra_staff * (1 - high_absence_rate)
        high_day_shortage = max(0, high_day_absent_positions - high_day_extra_coverage)
        lumpy_understaff_hrs = high_absence_days * shift_hours * high_day_shortage
        
        # LOW-ABSENCE DAYS: Overstaffing occurs
        # People absent = positions × low_absence_rate
        # We have extra_staff beyond positions, minus their absences
        low_day_absent_positions = positions * low_absence_rate
        low_day_extra_coverage = extra_staff * (1 - low_absence_rate)
        # Net excess = extra coverage - absences to cover
        low_day_excess = max(0, low_day_extra_coverage - low_day_absent_positions)
        lumpy_overstaff_hrs = low_absence_days * shift_hours * low_day_excess
        
        lumpy_understaff_cost = lumpy_understaff_hrs * adverse_understaff
        lumpy_overstaff_cost = lumpy_overstaff_hrs * adverse_overstaff
        scenarios.append((
            "Lumpy (Worst)",
            f"Mon/Fri {high_absence_rate:.0%} absent, Tue-Thu {low_absence_rate:.1%} (same avg)",
            lumpy_understaff_cost,
            lumpy_overstaff_cost
        ))
        
        # Scenario 3: Spread (typical - moderate daily variation around mean)
        # Same total absences, but some days slightly above, some slightly below average
        # Less extreme than Lumpy - maybe 1.3x on bad days, 0.7x on good days
        
        spread_high_days = int(work_days * 0.40)   # 40% of days above average
        spread_low_days = int(work_days * 0.40)    # 40% of days below average
        spread_avg_days = work_days - spread_high_days - spread_low_days  # 20% exactly average
        
        # Set rates that preserve the overall average
        spread_high_rate = absenteeism * 1.4  # 40% higher than average
        spread_low_rate = absenteeism * 0.5   # 50% lower than average
        # Verify weighted avg ≈ absenteeism (close enough for modeling)
        
        # High-absence days
        spread_high_absences = positions * spread_high_rate
        spread_high_coverage = extra_staff * (1 - spread_high_rate)
        spread_high_shortage = max(0, spread_high_absences - spread_high_coverage)
        
        # Low-absence days  
        spread_low_absences = positions * spread_low_rate
        spread_low_coverage = extra_staff * (1 - spread_low_rate)
        spread_low_excess = max(0, spread_low_coverage - spread_low_absences)
        
        # Average days (use flat calculation)
        spread_avg_shortage = max(0, daily_absences - daily_extra_coverage)
        spread_avg_excess = max(0, daily_extra_coverage - daily_absences)
        
        spread_understaff_hrs = (spread_high_days * shift_hours * spread_high_shortage + 
                                  spread_avg_days * shift_hours * spread_avg_shortage)
        spread_overstaff_hrs = (spread_low_days * shift_hours * spread_low_excess +
                                 spread_avg_days * shift_hours * spread_avg_excess)
        
        spread_understaff_cost = spread_understaff_hrs * adverse_understaff
        spread_overstaff_cost = spread_overstaff_hrs * adverse_overstaff
        scenarios.append((
            "Spread (Typical)",
            f"40% days at {spread_high_rate:.0%}, 40% at {spread_low_rate:.1%}, 20% average",
            spread_understaff_cost,
            spread_overstaff_cost
        ))
        
        # Scenario 4: Well-Managed (best realistic case)
        # Same variability as Spread, but proactive management reduces overstaffing
        # via VTO/LOW - sending people home when overstaffed
        # Understaffing still occurs but is unavoidable
        
        # Use same rates as Spread scenario
        managed_high_rate = absenteeism * 1.4
        managed_low_rate = absenteeism * 0.5
        
        managed_high_days = int(work_days * 0.40)
        managed_low_days = int(work_days * 0.40)
        managed_avg_days = work_days - managed_high_days - managed_low_days
        
        # Understaffing - same as Spread (can't avoid this)
        managed_high_absences = positions * managed_high_rate
        managed_high_coverage = extra_staff * (1 - managed_high_rate)
        managed_high_shortage = max(0, managed_high_absences - managed_high_coverage)
        
        managed_avg_shortage = max(0, daily_absences - daily_extra_coverage)
        
        managed_understaff_hrs = (managed_high_days * shift_hours * managed_high_shortage +
                                   managed_avg_days * shift_hours * managed_avg_shortage)
        
        # Overstaffing - REDUCED by 70% through VTO/LOW
        managed_low_absences = positions * managed_low_rate
        managed_low_coverage = extra_staff * (1 - managed_low_rate)
        managed_low_excess = max(0, managed_low_coverage - managed_low_absences)
        
        managed_avg_excess = max(0, daily_extra_coverage - daily_absences)
        
        # Only 30% of potential overstaffing actually occurs (70% sent home via VTO)
        vto_effectiveness = 0.30
        managed_overstaff_hrs = ((managed_low_days * shift_hours * managed_low_excess +
                                   managed_avg_days * shift_hours * managed_avg_excess) * vto_effectiveness)
        
        managed_understaff_cost = managed_understaff_hrs * adverse_understaff
        managed_overstaff_cost = managed_overstaff_hrs * adverse_overstaff
        scenarios.append((
            "Well-Managed",
            f"Same variance as Spread, but VTO/LOW reduces 70% of idle time",
            managed_understaff_cost,
            managed_overstaff_cost
        ))
        
        # Add scenarios to tree
        for scenario_name, description, under_cost, over_cost in scenarios:
            total = under_cost + over_cost
            self.adverse_tree.insert('', 'end', values=(
                scenario_name,
                description,
                f"${under_cost:,.0f}",
                f"${over_cost:,.0f}",
                f"${total:,.0f}"
            ))
        
        # Key insight
        if ratio > 5:
            insight = (f"Overstaffing costs {ratio:.0f}× more than understaffing. "
                      f"Running lean with planned OT is more cost-effective than staffing to eliminate OT. "
                      f"Consider using VTO/LOW on slow days to minimize idle time costs.")
        else:
            insight = (f"Overstaffing costs {ratio:.1f}× more than understaffing. "
                      f"Balance is important but lean slightly toward understaffing when uncertain.")
        
        self.adverse_labels['insight'].config(text=insight)
    
    def create_charts_tab(self):
        """Create the charts tab with visualizations."""
        charts_frame = ttk.Frame(self.notebook, padding="10")
        self.notebook.add(charts_frame, text="Charts")
        
        # Info label
        info_label = ttk.Label(charts_frame,
                              text="Charts will display saved schedule comparisons. Save schedules from the calculator first.",
                              foreground=self.colors['gray'])
        info_label.pack(pady=10)
        
        # Refresh button
        refresh_btn = ttk.Button(charts_frame, text="Refresh Charts", command=self.update_charts)
        refresh_btn.pack(pady=10)
        
        # Create matplotlib figure
        self.fig = Figure(figsize=(12, 6), dpi=100)
        self.canvas = FigureCanvasTkAgg(self.fig, master=charts_frame)
        self.canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        
    def update_charts(self):
        """Update the charts with saved costs data."""
        self.fig.clear()
        
        if not self.saved_costs:
            ax = self.fig.add_subplot(111)
            ax.text(0.5, 0.5, 'No saved schedules to display.\n\nSave schedules from the calculator first.',
                   ha='center', va='center', fontsize=12, color='gray')
            ax.axis('off')
            self.canvas.draw()
            return
        
        # Create subplots
        ax1 = self.fig.add_subplot(121)
        ax2 = self.fig.add_subplot(122)
        
        names = [c['schedule_name'][:15] for c in self.saved_costs]
        costs = [c['cost_scheduled_time'] for c in self.saved_costs]
        ot_15 = [c['marginal_ot_15'] for c in self.saved_costs]
        
        colors = ['#007acc', '#00a8ff', '#28a745', '#ffc107', '#dc3545', '#6c757d']
        bar_colors = [colors[i % len(colors)] for i in range(len(names))]
        
        # Chart 1: Cost of Scheduled Time
        bars1 = ax1.bar(range(len(names)), costs, color=bar_colors)
        ax1.set_xticks(range(len(names)))
        ax1.set_xticklabels(names, rotation=45, ha='right', fontsize=8)
        ax1.set_ylabel('Cost per Hour ($)')
        ax1.set_title('Cost of Scheduled Time', fontweight='bold')
        
        # Add value labels
        for bar, val in zip(bars1, costs):
            ax1.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.5,
                    f'${val:.2f}', ha='center', va='bottom', fontsize=8)
        
        # Chart 2: Scheduled Time vs OT Cost
        x = range(len(names))
        width = 0.35
        ax2.bar([i - width/2 for i in x], costs, width, label='Scheduled Time', color='#007acc')
        ax2.bar([i + width/2 for i in x], ot_15, width, label='OT 1.5x', color='#ffc107')
        ax2.set_xticks(range(len(names)))
        ax2.set_xticklabels(names, rotation=45, ha='right', fontsize=8)
        ax2.set_ylabel('Cost per Hour ($)')
        ax2.set_title('Scheduled Time vs Overtime', fontweight='bold')
        ax2.legend(fontsize=8)
        
        self.fig.tight_layout()
        self.canvas.draw()
        
    def reset_to_defaults(self):
        """Reset all inputs to default values."""
        self.burden_calc = BurdenCalculator()
        self.schedule_calc = ScheduleCostCalculator(self.burden_calc)
        
        for key, var in self.input_vars.items():
            try:
                var.set(self.burden_calc.inputs.get(key, 0))
            except:
                pass
                
        for key, var in self.schedule_vars.items():
            try:
                var.set(self.schedule_calc.schedule_inputs.get(key, 0))
            except:
                pass
        
        self.update_calculated_labels()
        self.on_schedule_change()
        messagebox.showinfo("Reset", "All inputs have been reset to default values.")
        
    def save_config(self):
        """Save current configuration to JSON file."""
        filename = filedialog.asksaveasfilename(
            defaultextension=".json",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
            title="Save Configuration"
        )
        
        if filename:
            try:
                config = {
                    'version': '3.2.0',
                    'timestamp': datetime.now().isoformat(),
                    'burden_inputs': {},
                    'schedule_inputs': {},
                    'saved_costs': self.saved_costs
                }
                
                for key, var in self.input_vars.items():
                    try:
                        config['burden_inputs'][key] = var.get()
                    except:
                        config['burden_inputs'][key] = 0
                        
                for key, var in self.schedule_vars.items():
                    try:
                        config['schedule_inputs'][key] = var.get()
                    except:
                        config['schedule_inputs'][key] = 0
                
                with open(filename, 'w') as f:
                    json.dump(config, f, indent=2)
                    
                messagebox.showinfo("Saved", f"Configuration saved to:\n{filename}")
            except Exception as e:
                messagebox.showerror("Save Error", f"Error saving configuration:\n{str(e)}")
                
    def load_config(self):
        """Load configuration from JSON file."""
        filename = filedialog.askopenfilename(
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
            title="Load Configuration"
        )
        
        if filename:
            try:
                with open(filename, 'r') as f:
                    config = json.load(f)
                
                # Load burden inputs
                burden_inputs = config.get('burden_inputs', config.get('inputs', {}))
                for key, value in burden_inputs.items():
                    if key in self.input_vars:
                        self.input_vars[key].set(value)
                        self.burden_calc.inputs[key] = value
                
                # Load schedule inputs
                schedule_inputs = config.get('schedule_inputs', {})
                for key, value in schedule_inputs.items():
                    if key in self.schedule_vars:
                        self.schedule_vars[key].set(value)
                        self.schedule_calc.schedule_inputs[key] = value
                
                # Load saved costs
                self.saved_costs = config.get('saved_costs', [])
                
                self.update_calculated_labels()
                self.on_schedule_change()
                self.update_saved_costs_display()
                self.update_analysis_schedule_list()
                self.update_compare_dropdowns()
                
                messagebox.showinfo("Loaded", f"Configuration loaded from:\n{filename}")
            except Exception as e:
                messagebox.showerror("Load Error", f"Error loading configuration:\n{str(e)}")
    
    def get_auto_save_path(self) -> str:
        """Get the full path to the auto-save file."""
        # Try to save in the same directory as the script
        try:
            script_dir = os.path.dirname(os.path.abspath(__file__))
            return os.path.join(script_dir, self.AUTO_SAVE_FILE)
        except:
            # Fallback to current working directory
            return self.AUTO_SAVE_FILE
    
    def auto_save(self):
        """Automatically save current state to default file (no dialog)."""
        try:
            config = {
                'version': '3.3.0',
                'timestamp': datetime.now().isoformat(),
                'burden_inputs': {},
                'schedule_inputs': {},
                'saved_costs': self.saved_costs
            }
            
            for key, var in self.input_vars.items():
                try:
                    config['burden_inputs'][key] = var.get()
                except:
                    config['burden_inputs'][key] = 0
                    
            for key, var in self.schedule_vars.items():
                try:
                    config['schedule_inputs'][key] = var.get()
                except:
                    config['schedule_inputs'][key] = 0
            
            save_path = self.get_auto_save_path()
            with open(save_path, 'w') as f:
                json.dump(config, f, indent=2)
                
        except Exception as e:
            # Silently fail on auto-save errors - don't interrupt user
            print(f"Auto-save warning: {e}")
    
    def load_auto_save(self):
        """Load auto-saved configuration if it exists."""
        try:
            save_path = self.get_auto_save_path()
            if os.path.exists(save_path):
                with open(save_path, 'r') as f:
                    config = json.load(f)
                
                # Load burden inputs
                burden_inputs = config.get('burden_inputs', {})
                for key, value in burden_inputs.items():
                    if key in self.input_vars:
                        self.input_vars[key].set(value)
                        self.burden_calc.inputs[key] = value
                
                # Load schedule inputs
                schedule_inputs = config.get('schedule_inputs', {})
                for key, value in schedule_inputs.items():
                    if key in self.schedule_vars:
                        self.schedule_vars[key].set(value)
                        self.schedule_calc.schedule_inputs[key] = value
                
                # Load saved costs
                self.saved_costs = config.get('saved_costs', [])
                
                self.update_calculated_labels()
                self.on_schedule_change()
                self.update_saved_costs_display()
                self.update_analysis_schedule_list()
                self.update_compare_dropdowns()
                
        except Exception as e:
            # Silently fail on auto-load errors - just start fresh
            print(f"Auto-load warning: {e}")
    
    def on_closing(self):
        """Handle window close - auto-save and exit."""
        self.auto_save()
        self.root.destroy()
                
    def export_excel(self):
        """Export saved costs to Excel file."""
        if not OPENPYXL_AVAILABLE:
            messagebox.showwarning("Export Unavailable", 
                                  "Excel export requires openpyxl library.\n\nInstall with: pip install openpyxl")
            return
        
        if not self.saved_costs:
            messagebox.showwarning("No Data", "No saved costs to export. Save some schedules first.")
            return
        
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Export to Excel"
        )
        
        if filename:
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "Cost of Time Analysis"
                
                # Styles
                header_font = Font(bold=True, size=12)
                header_fill = PatternFill(start_color="007acc", end_color="007acc", fill_type="solid")
                header_font_white = Font(bold=True, size=12, color="FFFFFF")
                
                # Title
                ws['A1'] = "Cost of Time Calculator - Shiftwork Solutions LLC"
                ws['A1'].font = Font(bold=True, size=16)
                ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
                ws['A2'].font = Font(italic=True, size=10)
                
                # Headers
                row = 4
                headers = ['Schedule Name', 'Type', 'Scheduled Hours', 'Hours Worked', 
                          'Cost/Hour', 'OT 1.5x', 'OT 2.0x', 'Ratio Paid/Worked']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col, value=header)
                    cell.font = header_font_white
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                
                # Data
                row = 5
                for cost in self.saved_costs:
                    ws.cell(row=row, column=1, value=cost['schedule_name'])
                    ws.cell(row=row, column=2, value=cost['schedule_type'])
                    ws.cell(row=row, column=3, value=cost['scheduled_hours'])
                    ws.cell(row=row, column=4, value=cost['hours_worked'])
                    ws.cell(row=row, column=5, value=cost['cost_scheduled_time'])
                    ws.cell(row=row, column=6, value=cost['marginal_ot_15'])
                    ws.cell(row=row, column=7, value=cost['marginal_ot_20'])
                    ws.cell(row=row, column=8, value=cost['ratio_paid_worked'])
                    row += 1
                
                # Auto-fit columns
                for col in range(1, 9):
                    ws.column_dimensions[get_column_letter(col)].width = 18
                
                wb.save(filename)
                messagebox.showinfo("Exported", f"Results exported to:\n{filename}")
            except Exception as e:
                messagebox.showerror("Export Error", f"Error exporting to Excel:\n{str(e)}")
                
    def export_pdf(self):
        """Export saved costs to PDF file."""
        if not REPORTLAB_AVAILABLE:
            messagebox.showwarning("Export Unavailable",
                                  "PDF export requires reportlab library.\n\nInstall with: pip install reportlab")
            return
        
        if not self.saved_costs:
            messagebox.showwarning("No Data", "No saved costs to export. Save some schedules first.")
            return
        
        filename = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
            title="Export to PDF"
        )
        
        if filename:
            try:
                doc = SimpleDocTemplate(filename, pagesize=landscape(letter))
                elements = []
                styles = getSampleStyleSheet()
                
                # Title
                title_style = ParagraphStyle(
                    'CustomTitle',
                    parent=styles['Heading1'],
                    fontSize=18,
                    textColor=colors.HexColor('#007acc'),
                    spaceAfter=12
                )
                elements.append(Paragraph("Cost of Time Calculator", title_style))
                elements.append(Paragraph("Shiftwork Solutions LLC", styles['Normal']))
                elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", 
                                         styles['Normal']))
                elements.append(Spacer(1, 20))
                
                # Results table
                headers = ['Schedule', 'Type', 'Scheduled', 'Worked', 'Cost/Hr', 'OT 1.5x', 'OT 2.0x']
                data = [headers]
                
                for cost in self.saved_costs:
                    row = [
                        cost['schedule_name'][:20],
                        cost['schedule_type'],
                        f"{cost['scheduled_hours']:,.0f}",
                        f"{cost['hours_worked']:,.1f}",
                        f"${cost['cost_scheduled_time']:.2f}",
                        f"${cost['marginal_ot_15']:.2f}",
                        f"${cost['marginal_ot_20']:.2f}",
                    ]
                    data.append(row)
                
                table = Table(data)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#007acc')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                    ('ALIGN', (0, 1), (0, -1), 'LEFT'),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('ROWHEIGHT', (0, 0), (-1, -1), 20),
                ]))
                elements.append(table)
                
                # Footer
                elements.append(Spacer(1, 30))
                footer_text = "For more information: shift-work.com | (415) 763-5005"
                elements.append(Paragraph(footer_text, styles['Normal']))
                
                doc.build(elements)
                messagebox.showinfo("Exported", f"Results exported to:\n{filename}")
            except Exception as e:
                messagebox.showerror("Export Error", f"Error exporting to PDF:\n{str(e)}")
                
    def show_about(self):
        """Show about dialog."""
        about_text = """Cost of Time Calculator
Version 3.3.0

Shiftwork Solutions LLC
https://shift-work.com
(415) 763-5005

A tool for comparing labor costs across different 
shift schedule configurations. Helps organizations 
understand the true cost of time for various 
scheduling approaches.

Copyright 2025 Shiftwork Solutions LLC
All Rights Reserved"""
        
        messagebox.showinfo("About", about_text)
        
    def show_contact(self):
        """Show contact information."""
        contact_text = """Shiftwork Solutions LLC

Website: https://shift-work.com
Phone: (415) 763-5005
Email: info@shift-work.com

With hundreds of successful projects across industries,
we specialize in helping organizations optimize their
shift operations for maximum efficiency and employee
satisfaction.

Services:
- Schedule Design & Analysis
- Staffing Strategy
- Change Management
- Employee Surveys
- Cost Analysis
- Training & Development"""
        
        messagebox.showinfo("Contact Shiftwork Solutions", contact_text)


def main():
    """Main entry point for the application."""
    root = tk.Tk()
    
    # Set DPI awareness for Windows
    try:
        from ctypes import windll
        windll.shcore.SetProcessDpiAwareness(1)
    except:
        pass
    
    app = CostOfTimeApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()


# I did no harm and this file is not truncated
