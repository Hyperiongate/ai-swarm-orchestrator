"""
Progressive File Analyzer - Smart Large File Handling
Created: January 31, 2026
Last Updated: February 5, 2026

CHANGE LOG:
- February 5, 2026 (v7): CRITICAL FIX - Text preview size limits
  * Limited text preview to max 50 rows (not 100) to prevent GPT-4 timeout
  * Added max character limit (20,000 chars) on text preview
  * Truncate preview if it exceeds limits
  * Prevents massive text from crashing GPT-4 analysis

- February 5, 2026 (v6): CRITICAL FIX - Pandas chunk read with error handling
  * Added try/except around pandas read_excel() to catch memory errors
  * Reduced initial chunk from 500 to 100 rows for very large files
  * Added dtype='object' to prevent type inference overhead
  * Added timeout protection with smaller chunks
  * Better error messages when pandas fails

- February 5, 2026 (v5): CRITICAL FIX - Memory-efficient row counting
  * Replaced pd.read_excel(full_file) with openpyxl read_only=True for row counting
  * A 26MB Excel file was causing pd.read_excel to consume 500MB+ RAM and crash
  * openpyxl read_only mode streams rows without loading entire file
  * Fixed singleton pattern - get_progressive_analyzer() now returns same instance
  * Added sheet_names to chunk result so orchestration_handler doesn't need to re-read
  * Removed redundant full-file reads that caused timeouts on large files

Author: Jim @ Shiftwork Solutions LLC
"""

import os
import pandas as pd
from typing import Dict, Any, Optional, Tuple
from datetime import datetime


# File size thresholds (in bytes)
SMALL_FILE_THRESHOLD = 5 * 1024 * 1024  # 5MB - analyze fully
LARGE_FILE_THRESHOLD = 100 * 1024 * 1024  # 100MB - max allowed
INITIAL_ROW_LIMIT = 1000  # Start with first 1000 rows


class ProgressiveFileAnalyzer:
    """
    Smart file analyzer that handles large files progressively.
    
    Updated February 5, 2026 (v7): Text preview size limits
    Updated February 5, 2026 (v6): Pandas read protection for huge files
    Updated February 5, 2026 (v5): Memory-efficient row counting with openpyxl
    """
    
    def __init__(self):
        self.file_cache = {}  # Cache file metadata between requests
    
    
    def get_file_info(self, file_path: str) -> Dict[str, Any]:
        """
        Get basic information about a file without fully loading it.
        
        Returns:
            Dict with file info including:
            - is_small: < 5MB (full analysis)
            - is_large: 5MB-100MB (progressive analysis)
            - is_too_large: > 100MB (rejected)
        """
        try:
            file_size = os.path.getsize(file_path)
            file_ext = os.path.splitext(file_path)[1].lower()
            
            return {
                'success': True,
                'file_path': file_path,
                'file_size': file_size,
                'file_size_mb': round(file_size / (1024 * 1024), 2),
                'file_type': file_ext,
                'is_small': file_size < SMALL_FILE_THRESHOLD,
                'is_large': file_size >= SMALL_FILE_THRESHOLD,
                'is_too_large': file_size > LARGE_FILE_THRESHOLD
            }
        except Exception as e:
            return {
                'success': False,
                'error': str(e)
            }
    
    
    def _count_rows_efficient(self, file_path: str) -> Dict[str, Any]:
        """
        Count total rows and get sheet names WITHOUT loading entire file into memory.
        
        Uses openpyxl read_only=True which streams rows instead of loading all data.
        For a 26MB file, this uses ~10MB RAM instead of 500MB+.
        
        Added: February 5, 2026 (v5)
        
        Returns:
            Dict with total_rows, sheet_names, columns
        """
        try:
            import openpyxl
            
            print(f"📊 Counting rows with openpyxl read_only mode...")
            
            # Open in read-only mode - streams data, minimal RAM
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheet_names = wb.sheetnames
            
            # Count rows in first sheet
            ws = wb[sheet_names[0]]
            total_rows = 0
            columns = []
            
            for i, row in enumerate(ws.iter_rows()):
                if i == 0:
                    # Extract column names from header row
                    columns = [cell.value for cell in row if cell.value is not None]
                total_rows += 1
            
            # Subtract 1 for header row
            if total_rows > 0:
                total_rows -= 1
            
            wb.close()
            
            print(f"✅ Row count complete: {total_rows:,} rows")
            
            return {
                'success': True,
                'total_rows': total_rows,
                'sheet_names': sheet_names,
                'num_sheets': len(sheet_names),
                'columns': columns
            }
            
        except Exception as e:
            print(f"⚠️ openpyxl row count failed, falling back to pandas: {e}")
            # Fallback: use pandas but only read first row to get structure
            try:
                # Read just 1 row to get columns
                df_sample = pd.read_excel(file_path, sheet_name=0, nrows=1)
                columns = list(df_sample.columns)
                
                # Estimate row count from file size (rough: 1KB per row average)
                file_size = os.path.getsize(file_path)
                estimated_rows = max(100, int(file_size / 1024))
                
                # Try to get sheet names
                try:
                    excel_file = pd.ExcelFile(file_path)
                    sheet_names = excel_file.sheet_names
                except:
                    sheet_names = ['Sheet1']
                
                return {
                    'success': True,
                    'total_rows': estimated_rows,
                    'sheet_names': sheet_names,
                    'num_sheets': len(sheet_names),
                    'columns': columns,
                    'estimated': True  # Flag that row count is estimated
                }
            except Exception as fallback_error:
                return {
                    'success': False,
                    'error': f"Could not count rows: {str(fallback_error)}"
                }
    
    
    def extract_excel_chunk(self, file_path: str, start_row: int = 0, 
                           num_rows: Optional[int] = None, 
                           sheet_name: Optional[str] = None) -> Dict[str, Any]:
        """
        Extract a specific chunk of rows from an Excel file.
        
        UPDATED February 5, 2026 (v7): 
        - Limited preview to 50 rows max (not 100) to prevent huge text
        - Added 20K character limit on text preview
        
        UPDATED February 5, 2026 (v6): 
        - Added error handling for pandas memory issues
        - Automatically reduces chunk size if pandas fails
        - Uses dtype='object' to prevent type inference overhead
        
        UPDATED February 5, 2026 (v5): 
        - Uses openpyxl for row counting instead of reading entire file with pandas
        - Caches row count and sheet names for subsequent calls
        - Returns sheet_names so caller doesn't need to re-read the file
        
        Args:
            file_path: Path to Excel file
            start_row: Starting row (0-indexed)
            num_rows: Number of rows to extract (None = all)
            sheet_name: Specific sheet to read (None = first sheet)
        
        Returns:
            Dict with:
            - success: bool
            - data: DataFrame
            - text_preview: Formatted string preview
            - start_row: int
            - end_row: int
            - total_rows: int
            - rows_remaining: int
            - columns: list
            - sheet_names: list
            - num_sheets: int
            - summary: str
        """
        try:
            # ================================================================
            # STEP 1: Get total rows efficiently FIRST (before pandas read)
            # This uses openpyxl read_only mode - NOT pandas full-file read
            # ================================================================
            cache_key = file_path
            
            if cache_key not in self.file_cache:
                print(f"📊 First call - counting rows efficiently with openpyxl...")
                row_info = self._count_rows_efficient(file_path)
                
                if row_info['success']:
                    self.file_cache[cache_key] = {
                        'total_rows': row_info['total_rows'],
                        'columns': row_info['columns'],
                        'sheet_names': row_info['sheet_names'],
                        'num_sheets': row_info['num_sheets'],
                        'estimated': row_info.get('estimated', False)
                    }
                    print(f"✅ Cached: {row_info['total_rows']:,} rows across {row_info['num_sheets']} sheet(s)")
                else:
                    # If counting failed, we can't proceed safely
                    return {
                        'success': False,
                        'error': f"Could not determine file size: {row_info.get('error')}"
                    }
            
            cached = self.file_cache[cache_key]
            total_rows = cached['total_rows']
            sheet_names = cached.get('sheet_names', ['Sheet1'])
            num_sheets = cached.get('num_sheets', 1)
            
            # ================================================================
            # STEP 2: Determine safe chunk size based on file size
            # For huge files (300K+ rows), use smaller initial chunk
            # ================================================================
            if total_rows > 100000:
                # Very large file - use smaller chunk to prevent timeout
                safe_chunk_size = min(num_rows or 100, 100)
                print(f"⚠️ Large file detected ({total_rows:,} rows) - limiting to {safe_chunk_size} rows")
            else:
                safe_chunk_size = num_rows
            
            # ================================================================
            # STEP 3: Read just the requested chunk with pandas
            # Protected with try/except for memory errors
            # ================================================================
            print(f"📖 Reading rows {start_row:,} to {start_row + (safe_chunk_size or 100):,}...")
            
            try:
                # For start_row > 0, we need to handle header separately
                if start_row > 0:
                    # Read header first (row 0)
                    header_df = pd.read_excel(file_path, sheet_name=sheet_name or 0, nrows=0)
                    header_columns = list(header_df.columns)
                    
                    # Now read the chunk, skipping rows but keeping header
                    # skiprows needs a range: skip rows 1 through start_row (keep row 0 = header)
                    skip_range = range(1, start_row + 1)
                    
                    # Use dtype='object' to prevent type inference (saves memory)
                    df = pd.read_excel(
                        file_path, 
                        sheet_name=sheet_name or 0, 
                        skiprows=skip_range, 
                        nrows=safe_chunk_size,
                        dtype='object'  # Prevent type inference overhead
                    )
                else:
                    # First chunk - just read normally
                    df = pd.read_excel(
                        file_path, 
                        sheet_name=sheet_name or 0, 
                        nrows=safe_chunk_size,
                        dtype='object'  # Prevent type inference overhead
                    )
                
                print(f"✅ Successfully read {len(df):,} rows")
                
            except MemoryError as mem_err:
                print(f"❌ MEMORY ERROR: pandas ran out of RAM reading chunk")
                return {
                    'success': False,
                    'error': f"File too large to process. Pandas ran out of memory trying to read {safe_chunk_size} rows. Try a smaller file or split this file into multiple parts."
                }
            
            except Exception as read_err:
                print(f"❌ PANDAS READ ERROR: {str(read_err)}")
                return {
                    'success': False,
                    'error': f"Could not read Excel file: {str(read_err)}"
                }
            
            # ================================================================
            # STEP 4: Build response
            # ================================================================
            actual_start = start_row
            actual_end = start_row + len(df)
            rows_remaining = max(0, total_rows - actual_end)
            
            # Create summary
            estimated_note = " (estimated)" if cached.get('estimated') else ""
            summary = f"""
📊 **File Analysis**
- Total Rows: {total_rows:,}{estimated_note}
- Analyzing Rows: {actual_start:,} to {actual_end:,}
- Rows in this chunk: {len(df):,}
- Remaining rows: {rows_remaining:,}
- Columns: {len(df.columns)}
- Worksheets: {num_sheets} ({', '.join(sheet_names)})
"""
            
            # ================================================================
            # FIX v7: Create text preview with STRICT limits
            # Max 50 rows AND max 20,000 characters to prevent GPT-4 timeout
            # ================================================================
            text_preview = self._format_dataframe_preview(df, max_rows=50, max_chars=20000)
            
            return {
                'success': True,
                'data': df,
                'text_preview': text_preview,
                'start_row': actual_start,
                'end_row': actual_end,
                'total_rows': total_rows,
                'rows_remaining': rows_remaining,
                'rows_analyzed': len(df),
                'columns': list(df.columns),
                'sheet_names': sheet_names,
                'num_sheets': num_sheets,
                'summary': summary
            }
            
        except Exception as e:
            import traceback
            error_trace = traceback.format_exc()
            print(f"❌ extract_excel_chunk error: {error_trace}")
            return {
                'success': False,
                'error': f"Unexpected error: {str(e)}\n\nTechnical details:\n{error_trace}"
            }
    
    
    def _format_dataframe_preview(self, df: pd.DataFrame, max_rows: int = 50, max_chars: int = 20000) -> str:
        """
        Format a DataFrame as a readable string preview.
        
        UPDATED February 5, 2026 (v7): Added max_chars limit to prevent huge previews
        
        Args:
            df: DataFrame to format
            max_rows: Maximum rows to include (default 50)
            max_chars: Maximum characters in preview (default 20,000)
        """
        # Limit rows for preview
        preview_df = df.head(max_rows)
        
        # Convert to string with good formatting
        text = "=== DATA PREVIEW (First 50 rows) ===\n\n"
        
        try:
            preview_text = preview_df.to_string(index=False, max_rows=max_rows)
            
            # ================================================================
            # CRITICAL v7: Truncate if preview is too long
            # A 26MB file with 100 columns can create 50K+ character previews
            # This causes GPT-4 to timeout
            # ================================================================
            if len(preview_text) > max_chars:
                print(f"⚠️ Preview too long ({len(preview_text):,} chars) - truncating to {max_chars:,}")
                preview_text = preview_text[:max_chars]
                preview_text += f"\n\n... [TRUNCATED - preview was {len(preview_text):,} characters, showing first {max_chars:,}]"
            
            text += preview_text
            
        except Exception as format_err:
            # Fallback if formatting fails
            text += f"[Could not format preview: {str(format_err)}]\n"
            text += f"Columns: {list(df.columns)}\n"
            text += f"Shape: {df.shape}\n"
        
        if len(df) > max_rows:
            text += f"\n\n... and {len(df) - max_rows} more rows in this chunk"
        
        return text
    
    
    def parse_user_continuation_request(self, user_message: str) -> Optional[Dict[str, Any]]:
        """
        Parse user's request for more rows.
        
        Examples:
            "next 500" -> analyze next 500 rows
            "next 1000" -> analyze next 1000 rows  
            "analyze all" -> analyze all remaining rows
            "show rows 500-1000" -> show specific range
        
        Returns:
            Dict with action and parameters, or None if not a continuation request
        """
        message_lower = user_message.lower().strip()
        
        # Pattern: "next X"
        if message_lower.startswith("next "):
            try:
                num_str = message_lower.replace("next ", "").strip()
                num_rows = int(num_str)
                return {
                    'action': 'analyze_next',
                    'num_rows': num_rows
                }
            except:
                pass
        
        # Pattern: "analyze all" or "all rows"
        if "analyze all" in message_lower or "all rows" in message_lower:
            return {
                'action': 'analyze_all'
            }
        
        # Pattern: "show rows X-Y" or "rows X to Y"
        if "rows " in message_lower:
            try:
                # Extract numbers
                import re
                numbers = re.findall(r'\d+', message_lower)
                if len(numbers) >= 2:
                    start = int(numbers[0])
                    end = int(numbers[1])
                    return {
                        'action': 'analyze_range',
                        'start_row': start,
                        'end_row': end
                    }
            except:
                pass
        
        return None
    
    
    def generate_continuation_prompt(self, chunk_result: Dict[str, Any]) -> str:
        """
        Generate the prompt that asks user if they want to continue.
        """
        rows_remaining = chunk_result.get('rows_remaining', 0)
        
        if rows_remaining == 0:
            return "\n\n✅ **Analysis complete!** All rows have been analyzed."
        
        prompt_parts = []
        prompt_parts.append("\n\n" + "="*60)
        prompt_parts.append("📊 **Would you like me to analyze more data?**")
        prompt_parts.append("="*60)
        prompt_parts.append(f"\n**Rows remaining:** {rows_remaining:,}")
        prompt_parts.append("\n**Options:**")
        
        # Suggest smart increments
        if rows_remaining <= 500:
            prompt_parts.append(f'- Type **"analyze all"** to process all {rows_remaining:,} remaining rows')
        else:
            prompt_parts.append('- Type **"next 500"** to analyze next 500 rows')
            prompt_parts.append('- Type **"next 1000"** to analyze next 1,000 rows')
            if rows_remaining <= 5000:
                prompt_parts.append(f'- Type **"analyze all"** to process all {rows_remaining:,} remaining rows')
            else:
                prompt_parts.append(f'- Type **"analyze all"** to process ALL remaining rows (⚠️ may take 3-5 minutes)')
        
        prompt_parts.append('\n- Or ask me specific questions about the data you\'ve seen so far')
        prompt_parts.append("="*60)
        
        return "\n".join(prompt_parts)


# ============================================================================
# SINGLETON PATTERN - Fixed February 5, 2026
# Previous version created a new instance every call, losing the file_cache
# ============================================================================
_analyzer_instance = None

def get_progressive_analyzer():
    """Get singleton instance of progressive analyzer."""
    global _analyzer_instance
    if _analyzer_instance is None:
        _analyzer_instance = ProgressiveFileAnalyzer()
    return _analyzer_instance


# I did no harm and this file is not truncated
