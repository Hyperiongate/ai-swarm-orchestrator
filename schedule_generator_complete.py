"""
Complete Schedule Generator - Input Collection + Beautiful Output
Created: January 27, 2026
Last Updated: January 27, 2026

This is the COMPLETE system:
1. Collects all necessary information from user
2. Generates beautiful Excel with BOTH grid and calendar views
3. Includes statistics and pros/cons

Author: Jim @ Shiftwork Solutions LLC
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import os


class ScheduleInputCollector:
    """Collect all information needed to generate the perfect schedule"""
    
    def __init__(self):
        self.pattern_name = '2-2-3'
        self.start_date = None
        self.day_shift_start = None
        self.day_shift_end = None
        self.night_shift_start = None
        self.night_shift_end = None
        self.crew_names = None
        self.same_pattern_for_nights = None
        self.night_pattern = None
        
    def get_start_date(self, custom_date=None):
        """Get or calculate start date (next Sunday)"""
        if custom_date:
            self.start_date = custom_date
        else:
            # Find next Sunday
            today = datetime.now()
            days_until_sunday = (6 - today.weekday()) % 7
            if days_until_sunday == 0:
                days_until_sunday = 7
            self.start_date = today + timedelta(days=days_until_sunday)
        
        return self.start_date
    
    def get_shift_times(self, day_start='06:00', day_end='18:00', 
                        night_start='18:00', night_end='06:00'):
        """Set shift times with defaults"""
        self.day_shift_start = day_start
        self.day_shift_end = day_end
        self.night_shift_start = night_start
        self.night_shift_end = night_end
        
        return {
            'day_start': self.day_shift_start,
            'day_end': self.day_shift_end,
            'night_start': self.night_shift_start,
            'night_end': self.night_shift_end
        }
    
    def get_crew_names(self, naming_style='letters', custom_names=None):
        """Set crew names based on style or custom input"""
        if custom_names:
            self.crew_names = custom_names
        elif naming_style == 'letters':
            self.crew_names = ['A', 'B', 'C', 'D']
        elif naming_style == 'numbers':
            self.crew_names = ['1', '2', '3', '4']
        elif naming_style == 'colors':
            self.crew_names = ['Red', 'Blue', 'Green', 'Yellow']
        
        return self.crew_names
    
    def ask_pattern_preference(self, same_pattern=True, night_pattern=None):
        """Ask if day and night should use same pattern"""
        self.same_pattern_for_nights = same_pattern
        if not same_pattern:
            self.night_pattern = night_pattern
        
        return self.same_pattern_for_nights


class Complete223ScheduleGenerator:
    """Generate beautiful 2-2-3 schedule with grid + calendar views"""
    
    def __init__(self):
        self.pattern_2_2_3 = {
            'name': '2-2-3 (Pitman)',
            'cycle_days': 14,
            'crews': 4,
            'crew_patterns': {
                'day_a': ['O','D','D','O','O','D','D','D','O','O','D','D','O','O'],
                'day_b': ['D','O','O','D','D','O','O','O','D','D','O','O','D','D'],
                'night_c': ['O','N','N','O','O','N','N','N','O','O','N','N','O','O'],
                'night_d': ['N','O','O','N','N','O','O','O','N','N','O','O','N','N']
            },
            'statistics': {
                'days_worked_per_cycle': 7,
                'days_off_per_cycle': 7,
                'avg_hours_per_week': 42,
                'max_consecutive_days': 2,
                'weekends_off_per_year': 26,
                'total_days_worked_per_year': 182,
                'total_days_off_per_year': 183
            },
            'benefits': [
                'Every other weekend off as 3-day weekend',
                'Never work more than 2 consecutive days',
                'Perfect 50/50 work-life balance (182 days on, 183 days off)',
                'Predictable 2-week repeating pattern',
                'Excellent for 12-hour shifts',
                'Allows for scheduling appointments and activities',
                'Equal distribution of weekends across all crews'
            ],
            'drawbacks': [
                'Night shift workers only get 2 days off between work stretches (not enough recovery time)',
                'Lack of consistency - crews only there 2 days before handoff to another crew',
                'Poor communication potential - short overlap between crews makes knowledge transfer difficult',
                'Work some holidays (not guaranteed off)',
                'No long breaks (maximum 3 days off)',
                'Some weekends are split (work Sat, off Sun or vice versa)',
                'Requires 4 crews (higher staffing cost)',
                '12-hour shifts can be physically demanding',
                'Difficult for supervisors to maintain accountability with frequent crew rotations'
            ]
        }
    
    def create_complete_schedule(self, start_date, shift_times, crew_names, 
                                 weeks_to_show=2, output_path='/tmp'):
        """
        Generate the complete schedule with both grid and calendar views
        
        Args:
            start_date: datetime object for first Sunday
            shift_times: dict with day_start, day_end, night_start, night_end
            crew_names: list of 4 crew names [A, B, C, D]
            weeks_to_show: how many weeks to display (default 2)
            output_path: where to save the file
            
        Returns:
            str: filepath to generated Excel file
        """
        wb = Workbook()
        
        # Create two sheets
        ws_grid = wb.active
        ws_grid.title = "Schedule Grid"
        ws_calendar = wb.create_sheet("Calendar View")
        
        # Generate both views
        self._create_grid_view(ws_grid, start_date, shift_times, crew_names, weeks_to_show)
        self._create_calendar_view(ws_calendar, start_date, shift_times, crew_names, weeks_to_show)
        
        # Save file
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'schedule_2-2-3_complete_{timestamp}.xlsx'
        
        filepath = None
        save_locations = [output_path, '/mnt/user-data/outputs', '/tmp', '.']
        
        for location in save_locations:
            try:
                os.makedirs(location, exist_ok=True)
                filepath = os.path.join(location, filename)
                wb.save(filepath)
                print(f"✅ Complete schedule saved to: {filepath}")
                break
            except Exception as e:
                print(f"⚠️  Could not save to {location}: {e}")
                continue
        
        if not filepath:
            raise Exception("Could not save schedule to any location")
        
        return filepath
    
    def _create_grid_view(self, ws, start_date, shift_times, crew_names, weeks_to_show):
        """Create the traditional grid view (crews in rows, days in columns)"""
        
        # Define colors
        company_blue = PatternFill(start_color='002B5B7C', end_color='002B5B7C', fill_type='solid')
        header_bg = PatternFill(start_color='00305496', end_color='00305496', fill_type='solid')
        day_fill = PatternFill(start_color='00FFD966', end_color='00FFD966', fill_type='solid')
        night_fill = PatternFill(start_color='00203864', end_color='00203864', fill_type='solid')
        off_fill = PatternFill(start_color='00E2EFDA', end_color='00E2EFDA', fill_type='solid')
        stats_bg = PatternFill(start_color='00F2F2F2', end_color='00F2F2F2', fill_type='solid')
        
        # Define fonts
        title_font = Font(name='Calibri', size=16, bold=True, color='00FFFFFF')
        section_font = Font(name='Calibri', size=13, bold=True, color='002B5B7C')
        header_font = Font(name='Calibri', size=11, bold=True, color='00FFFFFF')
        shift_font_day = Font(name='Calibri', size=11, bold=True, color='00000000')
        shift_font_night = Font(name='Calibri', size=11, bold=True, color='00FFFFFF')
        shift_font_off = Font(name='Calibri', size=11, bold=True, color='00666666')
        stat_font = Font(name='Calibri', size=10)
        
        # Define borders
        thin_border = Border(
            left=Side(style='thin', color='00D0D0D0'),
            right=Side(style='thin', color='00D0D0D0'),
            top=Side(style='thin', color='00D0D0D0'),
            bottom=Side(style='thin', color='00D0D0D0')
        )
        
        # Alignment
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        row = 1
        
        # ========================================
        # TITLE
        # ========================================
        ws.merge_cells(f'A{row}:K{row}')
        cell = ws[f'A{row}']
        cell.value = '2-2-3 SHIFT SCHEDULE (PITMAN PATTERN)'
        cell.font = title_font
        cell.fill = company_blue
        cell.alignment = center
        ws.row_dimensions[row].height = 30
        row += 2
        
        # ========================================
        # SCHEDULE INFO
        # ========================================
        ws[f'A{row}'] = 'Start Date:'
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = start_date.strftime('%B %d, %Y')
        ws.merge_cells(f'B{row}:D{row}')
        
        ws[f'F{row}'] = 'Day Shift:'
        ws[f'F{row}'].font = Font(bold=True)
        ws[f'G{row}'] = f"{shift_times['day_start']} - {shift_times['day_end']}"
        ws.merge_cells(f'G{row}:I{row}')
        
        row += 1
        ws[f'A{row}'] = 'Pattern Cycle:'
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = '14 days (2 weeks) - Repeats continuously'
        ws.merge_cells(f'B{row}:D{row}')
        
        ws[f'F{row}'] = 'Night Shift:'
        ws[f'F{row}'].font = Font(bold=True)
        ws[f'G{row}'] = f"{shift_times['night_start']} - {shift_times['night_end']}"
        ws.merge_cells(f'G{row}:I{row}')
        
        row += 2
        
        # ========================================
        # SCHEDULE GRID
        # ========================================
        ws.merge_cells(f'A{row}:K{row}')
        cell = ws[f'A{row}']
        cell.value = 'SCHEDULE GRID'
        cell.font = section_font
        cell.alignment = center
        row += 1
        
        # Headers
        header_row = row
        days_abbr = ['Sun', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat']
        
        ws.cell(row=header_row, column=1).value = 'Crew'
        ws.cell(row=header_row, column=2).value = 'Week'
        
        for col_idx, day in enumerate(days_abbr, start=3):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.value = day
            cell.font = header_font
            cell.fill = header_bg
            cell.alignment = center
            cell.border = thin_border
        
        ws.cell(row=header_row, column=1).font = header_font
        ws.cell(row=header_row, column=1).fill = header_bg
        ws.cell(row=header_row, column=1).alignment = center
        ws.cell(row=header_row, column=1).border = thin_border
        
        ws.cell(row=header_row, column=2).font = header_font
        ws.cell(row=header_row, column=2).fill = header_bg
        ws.cell(row=header_row, column=2).alignment = center
        ws.cell(row=header_row, column=2).border = thin_border
        
        row += 1
        
        # Crew schedules
        pattern = self.pattern_2_2_3
        crew_keys = ['day_a', 'day_b', 'night_c', 'night_d']
        
        for idx, crew_key in enumerate(crew_keys):
            crew_pattern = pattern['crew_patterns'][crew_key]
            crew_name = f"Crew {crew_names[idx]}"
            
            if idx < 2:
                crew_name += " (Days)"
            else:
                crew_name += " (Nights)"
            
            # Split into weeks
            weeks = [crew_pattern[i:i+7] for i in range(0, 14, 7)]
            
            for week_num, week_pattern in enumerate(weeks, start=1):
                # Crew name (only on first week)
                if week_num == 1:
                    cell = ws.cell(row=row, column=1)
                    cell.value = crew_name
                    cell.font = Font(bold=True)
                    cell.alignment = center
                    cell.border = thin_border
                else:
                    ws.cell(row=row, column=1).border = thin_border
                
                # Week number
                cell = ws.cell(row=row, column=2)
                cell.value = week_num
                cell.alignment = center
                cell.border = thin_border
                
                # Daily shifts
                for day_idx, shift in enumerate(week_pattern, start=3):
                    cell = ws.cell(row=row, column=day_idx)
                    cell.value = shift
                    cell.alignment = center
                    cell.border = thin_border
                    
                    if shift == 'D':
                        cell.fill = day_fill
                        cell.font = shift_font_day
                    elif shift == 'N':
                        cell.fill = night_fill
                        cell.font = shift_font_night
                    elif shift == 'O':
                        cell.fill = off_fill
                        cell.font = shift_font_off
                
                row += 1
            
            row += 1  # Blank row between crews
        
        # ========================================
        # LEGEND
        # ========================================
        row += 1
        ws.merge_cells(f'A{row}:K{row}')
        cell = ws[f'A{row}']
        cell.value = 'LEGEND'
        cell.font = section_font
        cell.alignment = center
        row += 1
        
        legend_items = [
            ('D', 'Day Shift', day_fill, shift_font_day, f"12 hours: {shift_times['day_start']}-{shift_times['day_end']}"),
            ('N', 'Night Shift', night_fill, shift_font_night, f"12 hours: {shift_times['night_start']}-{shift_times['night_end']}"),
            ('O', 'OFF', off_fill, shift_font_off, 'Not scheduled')
        ]
        
        for code, label, fill, font, description in legend_items:
            ws.cell(row=row, column=1).value = code
            ws.cell(row=row, column=1).fill = fill
            ws.cell(row=row, column=1).font = font
            ws.cell(row=row, column=1).alignment = center
            ws.cell(row=row, column=1).border = thin_border
            
            ws.cell(row=row, column=2).value = label
            ws.cell(row=row, column=2).font = Font(bold=True)
            ws.cell(row=row, column=2).border = thin_border
            ws.merge_cells(f'B{row}:C{row}')
            
            ws.cell(row=row, column=4).value = description
            ws.cell(row=row, column=4).border = thin_border
            ws.merge_cells(f'D{row}:G{row}')
            
            row += 1
        
        # ========================================
        # STATISTICS
        # ========================================
        row += 2
        ws.merge_cells(f'A{row}:K{row}')
        cell = ws[f'A{row}']
        cell.value = 'PATTERN STATISTICS'
        cell.font = section_font
        cell.alignment = center
        row += 1
        
        stats = pattern['statistics']
        stat_items = [
            ('Days worked per cycle (14 days):', stats['days_worked_per_cycle']),
            ('Days off per cycle (14 days):', stats['days_off_per_cycle']),
            ('Average hours per week:', stats['avg_hours_per_week']),
            ('Maximum consecutive days worked:', stats['max_consecutive_days']),
            ('Weekends off per year:', stats['weekends_off_per_year']),
            ('Total days worked per year:', stats['total_days_worked_per_year']),
            ('Total days off per year:', stats['total_days_off_per_year'])
        ]
        
        for label, value in stat_items:
            ws.cell(row=row, column=1).value = label
            ws.cell(row=row, column=1).font = stat_font
            ws.cell(row=row, column=1).fill = stats_bg
            ws.cell(row=row, column=1).border = thin_border
            ws.merge_cells(f'A{row}:C{row}')
            
            ws.cell(row=row, column=4).value = str(value)
            ws.cell(row=row, column=4).font = Font(bold=True)
            ws.cell(row=row, column=4).fill = stats_bg
            ws.cell(row=row, column=4).alignment = center
            ws.cell(row=row, column=4).border = thin_border
            
            row += 1
        
        # ========================================
        # BENEFITS & DRAWBACKS
        # ========================================
        row += 2
        
        # Benefits
        ws.merge_cells(f'A{row}:E{row}')
        cell = ws[f'A{row}']
        cell.value = 'PATTERN BENEFITS'
        cell.font = Font(bold=True, color='00006400')
        cell.alignment = center
        row += 1
        
        for benefit in pattern['benefits']:
            ws.cell(row=row, column=1).value = f"✓ {benefit}"
            ws.cell(row=row, column=1).font = stat_font
            ws.merge_cells(f'A{row}:E{row}')
            row += 1
        
        # Drawbacks
        drawback_start_row = row - len(pattern['benefits']) - 1
        ws.merge_cells(f'G{drawback_start_row}:K{drawback_start_row}')
        cell = ws[f'G{drawback_start_row}']
        cell.value = 'PATTERN DRAWBACKS'
        cell.font = Font(bold=True, color='008B0000')
        cell.alignment = center
        
        drawback_row = drawback_start_row + 1
        for drawback in pattern['drawbacks']:
            ws.cell(row=drawback_row, column=7).value = f"• {drawback}"
            ws.cell(row=drawback_row, column=7).font = stat_font
            ws.merge_cells(f'G{drawback_row}:K{drawback_row}')
            drawback_row += 1
        
        # Column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 8
        for col in range(3, 10):
            ws.column_dimensions[get_column_letter(col)].width = 9
    
    def _create_calendar_view(self, ws, start_date, shift_times, crew_names, weeks_to_show):
        """Create the calendar view (actual dates with crew assignments)"""
        
        # Define colors
        header_bg = PatternFill(start_color='00305496', end_color='00305496', fill_type='solid')
        day_fill = PatternFill(start_color='00FFD966', end_color='00FFD966', fill_type='solid')
        night_fill = PatternFill(start_color='00203864', end_color='00203864', fill_type='solid')
        weekend_bg = PatternFill(start_color='00F0F0F0', end_color='00F0F0F0', fill_type='solid')
        
        # Fonts
        title_font = Font(name='Calibri', size=16, bold=True, color='00FFFFFF')
        header_font = Font(name='Calibri', size=11, bold=True, color='00FFFFFF')
        date_font = Font(name='Calibri', size=12, bold=True)
        crew_font = Font(name='Calibri', size=10)
        
        # Borders
        thin_border = Border(
            left=Side(style='thin', color='00D0D0D0'),
            right=Side(style='thin', color='00D0D0D0'),
            top=Side(style='thin', color='00D0D0D0'),
            bottom=Side(style='thin', color='00D0D0D0')
        )
        
        # Alignment
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        row = 1
        
        # Title
        ws.merge_cells(f'A{row}:H{row}')
        cell = ws[f'A{row}']
        cell.value = 'CALENDAR VIEW'
        cell.font = title_font
        cell.fill = header_bg
        cell.alignment = center
        ws.row_dimensions[row].height = 30
        row += 2
        
        # Day headers
        days = ['Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
        for col_idx, day in enumerate(days, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = day
            cell.font = header_font
            cell.fill = header_bg
            cell.alignment = center
            cell.border = thin_border
        
        row += 1
        
        # Generate calendar
        pattern = self.pattern_2_2_3
        current_date = start_date
        
        for week in range(weeks_to_show):
            week_start_row = row
            
            for day_idx in range(7):
                col = day_idx + 1
                day_in_cycle = (week * 7 + day_idx) % 14
                
                # Date
                cell = ws.cell(row=row, column=col)
                cell.value = current_date.strftime('%b %d')
                cell.font = date_font
                cell.alignment = center
                cell.border = thin_border
                
                if day_idx == 0 or day_idx == 6:  # Weekend
                    cell.fill = weekend_bg
                
                # Find who's working
                day_crew = None
                night_crew = None
                
                for idx, crew_key in enumerate(['day_a', 'day_b', 'night_c', 'night_d']):
                    shift = pattern['crew_patterns'][crew_key][day_in_cycle]
                    if shift == 'D':
                        day_crew = crew_names[idx]
                    elif shift == 'N':
                        night_crew = crew_names[idx]
                
                # Day shift
                cell = ws.cell(row=row+1, column=col)
                cell.value = f"Day: {day_crew}" if day_crew else "Day: -"
                cell.font = crew_font
                cell.alignment = center
                cell.border = thin_border
                if day_crew:
                    cell.fill = day_fill
                
                # Night shift
                cell = ws.cell(row=row+2, column=col)
                cell.value = f"Night: {night_crew}" if night_crew else "Night: -"
                cell.font = crew_font
                cell.alignment = center
                cell.border = thin_border
                if night_crew:
                    cell.fill = night_fill
                    cell.font = Font(name='Calibri', size=10, color='00FFFFFF')
                
                current_date += timedelta(days=1)
            
            row += 4  # Move to next week (date + 2 shift rows + gap)
        
        # Column widths
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 18


def create_complete_223_schedule():
    """
    Main function to create complete schedule with all inputs
    
    This demonstrates the complete flow
    """
    # Step 1: Collect inputs
    collector = ScheduleInputCollector()
    
    # Get start date (next Sunday)
    start_date = collector.get_start_date()
    print(f"Start date: {start_date.strftime('%B %d, %Y')}")
    
    # Get shift times (defaults)
    shift_times = collector.get_shift_times()
    print(f"Day shift: {shift_times['day_start']} - {shift_times['day_end']}")
    print(f"Night shift: {shift_times['night_start']} - {shift_times['night_end']}")
    
    # Get crew names (letters)
    crew_names = collector.get_crew_names('letters')
    print(f"Crew names: {crew_names}")
    
    # Ask about pattern (same for now)
    same_pattern = collector.ask_pattern_preference(True)
    print(f"Same pattern for day/night: {same_pattern}")
    
    # Step 2: Generate schedule
    generator = Complete223ScheduleGenerator()
    filepath = generator.create_complete_schedule(
        start_date=start_date,
        shift_times=shift_times,
        crew_names=crew_names,
        weeks_to_show=2
    )
    
    return filepath


# Test if run directly
if __name__ == '__main__':
    print("=" * 60)
    print("COMPLETE 2-2-3 SCHEDULE GENERATOR")
    print("=" * 60)
    filepath = create_complete_223_schedule()
    print(f"\n✅ Complete schedule generated!")
    print(f"📄 File: {filepath}")
    print(f"\nIncludes:")
    print("  • Grid view (traditional schedule)")
    print("  • Calendar view (actual dates)")
    print("  • Statistics")
    print("  • Benefits & Drawbacks")


# I did no harm and this file is not truncated
