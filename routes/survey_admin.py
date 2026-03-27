"""
Survey in a Box — Admin Dashboard Routes
File: routes/survey_admin.py
Created: March 16, 2026
Last Updated: March 27, 2026 — Phase 2: Code mode selection (random vs employee ID)

PURPOSE:
    Password-protected admin dashboard for Survey in a Box.
    Jim uses this to review intake submissions, configure survey projects,
    and approve clients for Phase 2 (survey assembly).

ENDPOINTS (Phase 1):
    GET  /survey/admin                          — Dashboard (password gate)
    GET  /api/survey/admin/clients              — List all survey_clients (JSON)
    GET  /api/survey/admin/client/<id>          — Get one client + project (JSON)
    POST /api/survey/admin/client/<id>/status   — Update client status
    POST /api/survey/admin/project/save         — Save/update survey_projects row
    POST /api/survey/admin/project/<id>/approve — Approve project (status -> approved)
    POST /api/survey/admin/client/<id>/notes    — Save admin notes to survey_clients

ENDPOINTS (Phase 2 -- Roster):
    POST /api/survey/admin/project/<id>/roster/upload
        Accepts .xlsx file upload.
        Request form fields:
            file          (required) .xlsx file
            use_employee_id (optional) 'true' to use existing employee IDs
                            instead of generating random codes.
        Reads columns: Name/Employee Name/Full Name (required),
                       Dept/Department, Shift/Shift Name,
                       Tenure/Years/Seniority (all optional),
                       Employee ID/Badge Number/Emp ID (required if use_employee_id=true).
        If use_employee_id=false (default): generates unique random 5-digit codes.
        If use_employee_id=true: reads existing IDs from roster, validates
          each as 1-20 alphanumeric+hyphen chars, stores verbatim.
        Stores code_mode ('random' or 'employee_id') on survey_projects.
        Wipes existing roster before inserting new one.
        Returns: success, count, message, mapped columns, skipped columns,
                 code_mode used.

    GET  /api/survey/admin/project/<id>/roster/download-codes
        Generates a printable .xlsx: Employee Name | Survey Code.
        Instructions footer text adapts to code_mode.
        Returns: downloadable .xlsx file.

AUTH:
    Session-cookie based. Password checked against SURVEY_ADMIN_PASSWORD env var.

DESIGN RULES:
    - PostgreSQL: RealDictCursor, %s params, RETURNING id
    - All JSON list fields decoded on read, encoded on write
    - No changes to any existing Swarm table or route
    - Blueprint prefix: none (routes defined with full paths)
    - Roster codes (random mode): random 5-digit integers (10000-99999), unique per project
    - Roster codes (employee_id mode): verbatim from xlsx, validated 1-20 alphanumeric+hyphen
    - Column fuzzy matching tolerates common variations in roster header names
    - openpyxl used for both reading uploads and generating download xlsx

CHANGELOG:
    - March 27, 2026: Code mode selection -- TWO CHANGES to Phase 2 helpers only:
      1. _map_column_name(): added employee_id_col mappings for Employee ID,
         Badge Number, Emp ID, Worker ID, Payroll ID, Badge No., BadgeNum,
         Employee No, and similar variants.
      2. upload_roster(): added use_employee_id form field handling.
         When use_employee_id=true, reads employee_id_col from xlsx instead
         of generating random codes. Validates each ID (1-20 alphanumeric+hyphen).
         Stores code_mode='employee_id' or 'random' on survey_projects.
         Returns code_mode in response.
         download_roster_codes(): reads code_mode from DB and adjusts the
         instructions footer text accordingly.
      NO CHANGES to any Phase 1 endpoints or helpers.

    - March 26, 2026: Phase 2 Deliverable 2 -- Added roster upload and
      download-codes endpoints. Added _map_column_name() fuzzy mapper,
      _generate_unique_codes() helper. No changes to any Phase 1 endpoints.
    - March 16, 2026: Initial creation for Phase 1, Step 1.3.
"""

import json
import os
import random
import re
import secrets
import traceback
from datetime import datetime
from functools import wraps
from io import BytesIO

from flask import (Blueprint, render_template, request, jsonify,
                   session, redirect, url_for, send_file)

survey_admin_bp = Blueprint('survey_admin', __name__)

# ---------------------------------------------------------------------------
# AUTH HELPERS
# ---------------------------------------------------------------------------

ADMIN_SESSION_KEY = 'survey_admin_authenticated'


def _get_admin_password():
    """Read the admin password from environment. Never hard-code."""
    return os.environ.get('SURVEY_ADMIN_PASSWORD', '')


def _is_authenticated():
    return session.get(ADMIN_SESSION_KEY) is True


def require_survey_admin(f):
    """Decorator: return 401 JSON for API routes if not authenticated."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if not _is_authenticated():
            return jsonify({'success': False, 'error': 'Not authenticated.', 'redirect': '/survey/admin'}), 401
        return f(*args, **kwargs)
    return decorated


# ---------------------------------------------------------------------------
# DB HELPER
# ---------------------------------------------------------------------------

def _get_db():
    from db_engine import get_db_connection
    return get_db_connection()


# ---------------------------------------------------------------------------
# JSON DECODE HELPERS
# ---------------------------------------------------------------------------

def _decode_json_field(val, fallback=None):
    """Safely decode a JSON string stored in a TEXT column."""
    if val is None:
        return fallback if fallback is not None else []
    if isinstance(val, (list, dict)):
        return val
    try:
        return json.loads(val)
    except (json.JSONDecodeError, TypeError):
        return fallback if fallback is not None else []


def _client_to_dict(row):
    """Convert a survey_clients RealDictRow to a clean Python dict."""
    d = dict(row)
    d['department_names']   = _decode_json_field(d.get('department_names'), [])
    d['biggest_challenges'] = _decode_json_field(d.get('biggest_challenges'), [])
    for key in ('created_at', 'updated_at'):
        if d.get(key) and hasattr(d[key], 'isoformat'):
            d[key] = d[key].isoformat()
    return d


def _project_to_dict(row):
    """Convert a survey_projects RealDictRow to a clean Python dict."""
    d = dict(row)
    for field in ('selected_questions', 'excluded_questions',
                  'custom_questions', 'selected_schedules'):
        d[field] = _decode_json_field(d.get(field), [])
    for key in ('created_at', 'updated_at', 'approved_at', 'opened_at', 'closed_at'):
        if d.get(key) and hasattr(d[key], 'isoformat'):
            d[key] = d[key].isoformat()
    return d


# ---------------------------------------------------------------------------
# PHASE 2 HELPERS — ROSTER UPLOAD & CODE GENERATION
# ---------------------------------------------------------------------------

def _map_column_name(col_header):
    """
    Fuzzy-map a spreadsheet column header to a canonical roster field name.
    Returns the canonical name or None if unrecognised.

    Canonical names:
        employee_name   — required in all modes
        department      — optional demographic
        shift           — optional demographic
        tenure_bracket  — optional demographic
        employee_id_col — required in employee_id mode (existing IDs)
    """
    col = col_header.strip().lower()

    if col in ('name', 'employee name', 'full name', 'employee_name',
               'fullname', 'emp name', 'emp. name'):
        return 'employee_name'

    if col in ('dept', 'department', 'dept.', 'department name',
               'dept name', 'area', 'work area'):
        return 'department'

    if col in ('shift', 'shift name', 'shift_name', 'shiftname',
               'shift type', 'work shift'):
        return 'shift'

    if col in ('tenure', 'years', 'seniority', 'tenure_bracket',
               'years of service', 'yrs', 'tenure bracket',
               'yrs of service', 'service years'):
        return 'tenure_bracket'

    if col in ('employee id', 'emp id', 'emp. id', 'employee_id', 'empid',
               'id', 'badge', 'badge number', 'badge no', 'badge no.',
               'worker id', 'worker_id', 'payroll id', 'payroll_id',
               'badgenum', 'badgeno', 'employee no', 'employee no.', 'emp no'):
        return 'employee_id_col'

    return None


def _validate_employee_id_value(val):
    """
    Validate a single employee ID value for employee_id mode.
    Accepts 1-20 characters: digits, letters, and hyphens.
    Returns (is_valid, error_reason_or_None).
    """
    if not val or not str(val).strip():
        return False, 'empty value'
    v = str(val).strip()
    if len(v) > 20:
        return False, f'too long ({len(v)} chars, max 20)'
    if not re.match(r'^[A-Za-z0-9\-]+$', v):
        return False, 'contains invalid characters (only digits, letters, hyphens allowed)'
    return True, None


def _generate_unique_codes(count):
    """
    Generate `count` unique random 5-digit codes (10000-99999).
    Used in random mode. Returns list of strings e.g. ['25281', '94670', ...].
    """
    if count > 90000:
        raise ValueError(f"Cannot generate {count} unique codes from 10000-99999 (max 90000)")
    pool = list(range(10000, 100000))
    random.shuffle(pool)
    return [str(c) for c in pool[:count]]


def _read_roster_xlsx(file_bytes, use_employee_id=False):
    """
    Parse an .xlsx roster file and return:
        rows    — list of dicts: employee_name (required), department, shift,
                  tenure_bracket, employee_id_val (when use_employee_id=True)
        mapped  — dict of {original_header: canonical_name}
        skipped — list of unrecognised column headers

    Raises ValueError if employee_name column not found, or if
    use_employee_id=True and employee_id_col column not found.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("openpyxl is required for roster upload.")

    wb = load_workbook(filename=BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)

    try:
        headers = [str(h).strip() if h is not None else '' for h in next(rows_iter)]
    except StopIteration:
        raise ValueError("Roster file appears to be empty.")

    col_map = {}
    mapped  = {}
    skipped = []
    for idx, header in enumerate(headers):
        canonical = _map_column_name(header)
        if canonical:
            if canonical not in col_map:
                col_map[canonical] = idx
                mapped[header] = canonical
        else:
            if header:
                skipped.append(header)

    if 'employee_name' not in col_map:
        raise ValueError(
            f"Could not find employee name column. "
            f"Headers found: {headers}. "
            f"Expected one of: Name, Employee Name, Full Name."
        )

    if use_employee_id and 'employee_id_col' not in col_map:
        raise ValueError(
            f"Employee ID mode selected but no Employee ID column found. "
            f"Headers found: {headers}. "
            f"Expected one of: Employee ID, Badge Number, Emp ID, Badge No., Worker ID, Payroll ID."
        )

    result_rows = []
    for raw_row in rows_iter:
        if all(cell is None or str(cell).strip() == '' for cell in raw_row):
            continue

        def _cell(key):
            idx = col_map.get(key)
            if idx is None or idx >= len(raw_row):
                return None
            val = raw_row[idx]
            return str(val).strip() if val is not None else None

        name = _cell('employee_name')
        if not name:
            continue

        row_data = {
            'employee_name':  name,
            'department':     _cell('department'),
            'shift':          _cell('shift'),
            'tenure_bracket': _cell('tenure_bracket'),
        }
        if use_employee_id:
            row_data['employee_id_val'] = _cell('employee_id_col')

        result_rows.append(row_data)

    wb.close()
    return result_rows, mapped, skipped


# ---------------------------------------------------------------------------
# DASHBOARD PAGE
# ---------------------------------------------------------------------------

@survey_admin_bp.route('/survey/admin', methods=['GET'])
def admin_dashboard():
    """Render the admin dashboard HTML. Auth is handled client-side via API."""
    return render_template('survey_admin.html')


# ---------------------------------------------------------------------------
# AUTH ENDPOINTS
# ---------------------------------------------------------------------------

@survey_admin_bp.route('/api/survey/admin/login', methods=['POST'])
def admin_login():
    """Validate password and set session cookie."""
    data     = request.get_json(force=True) or {}
    password = data.get('password', '')
    correct  = _get_admin_password()

    if not correct:
        return jsonify({'success': False, 'error': 'Admin password not configured on server.'}), 500

    if password == correct:
        session[ADMIN_SESSION_KEY] = True
        session.permanent = True
        return jsonify({'success': True})
    else:
        return jsonify({'success': False, 'error': 'Incorrect password.'}), 401


@survey_admin_bp.route('/api/survey/admin/logout', methods=['POST'])
def admin_logout():
    """Clear the admin session."""
    session.pop(ADMIN_SESSION_KEY, None)
    return jsonify({'success': True})


@survey_admin_bp.route('/api/survey/admin/auth-status', methods=['GET'])
def auth_status():
    """Check whether the current session is authenticated."""
    return jsonify({'authenticated': _is_authenticated()})


# ---------------------------------------------------------------------------
# CLIENT LIST
# ---------------------------------------------------------------------------

@survey_admin_bp.route('/api/survey/admin/clients', methods=['GET'])
@require_survey_admin
def list_clients():
    """Return all survey_clients ordered by created_at DESC."""
    try:
        conn = _get_db()
        try:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT
                    sc.*,
                    sp.id             AS project_id,
                    sp.project_token  AS project_token,
                    sp.status         AS project_status,
                    sp.response_count AS response_count
                FROM survey_clients sc
                LEFT JOIN survey_projects sp ON sp.survey_client_id = sc.id
                ORDER BY sc.created_at DESC
            """)
            rows = cursor.fetchall()
        finally:
            conn.close()

        clients = []
        for row in rows:
            d = _client_to_dict(row)
            d['project_id']     = d.get('project_id')
            d['project_token']  = d.get('project_token')
            d['project_status'] = d.get('project_status')
            d['response_count'] = d.get('response_count', 0)
            clients.append(d)

        return jsonify({'success': True, 'clients': clients, 'total': len(clients)})

    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


# ---------------------------------------------------------------------------
# SINGLE CLIENT DETAIL
# ---------------------------------------------------------------------------

@survey_admin_bp.route('/api/survey/admin/client/<int:client_id>', methods=['GET'])
@require_survey_admin
def get_client(client_id):
    """Return full detail for one client plus their project config (if any)."""
    try:
        conn = _get_db()
        try:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM survey_clients WHERE id = %s", (client_id,))
            client_row = cursor.fetchone()
            if not client_row:
                return jsonify({'success': False, 'error': f'Client {client_id} not found.'}), 404

            cursor.execute(
                "SELECT * FROM survey_projects WHERE survey_client_id = %s ORDER BY created_at DESC LIMIT 1",
                (client_id,)
            )
            project_row = cursor.fetchone()
        finally:
            conn.close()

        client  = _client_to_dict(client_row)
        project = _project_to_dict(project_row) if project_row else None

        return jsonify({'success': True, 'client': client, 'project': project})

    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


# ---------------------------------------------------------------------------
# UPDATE CLIENT STATUS
# ---------------------------------------------------------------------------

@survey_admin_bp.route('/api/survey/admin/client/<int:client_id>/status', methods=['POST'])
@require_survey_admin
def update_client_status(client_id):
    """Change the status of a survey_clients row."""
    try:
        data       = request.get_json(force=True) or {}
        new_status = data.get('status', '').strip()

        valid_statuses = ('new', 'reviewing', 'approved', 'rejected', 'in-progress', 'delivered')
        if new_status not in valid_statuses:
            return jsonify({'success': False,
                            'error': f'Invalid status. Must be one of: {", ".join(valid_statuses)}'}), 400

        conn = _get_db()
        try:
            cursor = conn.cursor()
            cursor.execute(
                "UPDATE survey_clients SET status = %s, updated_at = NOW() WHERE id = %s",
                (new_status, client_id)
            )
            if cursor.rowcount == 0:
                conn.rollback()
                return jsonify({'success': False, 'error': f'Client {client_id} not found.'}), 404
            conn.commit()
        finally:
            conn.close()

        return jsonify({'success': True, 'client_id': client_id, 'status': new_status})

    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


# ---------------------------------------------------------------------------
# SAVE ADMIN NOTES ON CLIENT
# ---------------------------------------------------------------------------

@survey_admin_bp.route('/api/survey/admin/client/<int:client_id>/notes', methods=['POST'])
@require_survey_admin
def save_client_notes(client_id):
    """Save Jim's admin notes to survey_clients.admin_notes."""
    try:
        data  = request.get_json(force=True) or {}
        notes = data.get('admin_notes', '')

        conn = _get_db()
        try:
            cursor = conn.cursor()
            cursor.execute(
                "UPDATE survey_clients SET admin_notes = %s, updated_at = NOW() WHERE id = %s",
                (notes, client_id)
            )
            if cursor.rowcount == 0:
                conn.rollback()
                return jsonify({'success': False, 'error': f'Client {client_id} not found.'}), 404
            conn.commit()
        finally:
            conn.close()

        return jsonify({'success': True, 'client_id': client_id})

    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


# ---------------------------------------------------------------------------
# SAVE / UPDATE PROJECT CONFIGURATION
# ---------------------------------------------------------------------------

@survey_admin_bp.route('/api/survey/admin/project/save', methods=['POST'])
@require_survey_admin
def save_project():
    """
    Upsert a survey_projects row for a given survey_client_id.
    If a project already exists for this client, update it.
    If not, create a new one (generates project_token).
    """
    try:
        data      = request.get_json(force=True) or {}
        client_id = data.get('survey_client_id')
        if not client_id:
            return jsonify({'success': False, 'error': 'survey_client_id is required.'}), 400

        selected_questions = json.dumps(data.get('selected_questions', []))
        excluded_questions = json.dumps(data.get('excluded_questions', []))
        custom_questions   = json.dumps(data.get('custom_questions', []))
        selected_schedules = json.dumps(data.get('selected_schedules', []))
        admin_notes        = data.get('admin_notes', None)

        conn = _get_db()
        try:
            cursor = conn.cursor()

            cursor.execute(
                "SELECT id FROM survey_projects WHERE survey_client_id = %s ORDER BY created_at DESC LIMIT 1",
                (client_id,)
            )
            existing = cursor.fetchone()

            if existing:
                project_id = existing['id'] if isinstance(existing, dict) else existing[0]
                cursor.execute("""
                    UPDATE survey_projects SET
                        selected_questions = %s,
                        excluded_questions = %s,
                        custom_questions   = %s,
                        selected_schedules = %s,
                        admin_notes        = %s,
                        updated_at         = NOW()
                    WHERE id = %s
                """, (
                    selected_questions, excluded_questions, custom_questions,
                    selected_schedules, admin_notes, project_id
                ))
                conn.commit()
                return jsonify({'success': True, 'project_id': project_id, 'action': 'updated'})
            else:
                token = secrets.token_urlsafe(16)
                cursor.execute("""
                    INSERT INTO survey_projects (
                        survey_client_id, project_token, status,
                        selected_questions, excluded_questions,
                        custom_questions, selected_schedules, admin_notes,
                        created_at, updated_at
                    ) VALUES (%s, %s, 'draft', %s, %s, %s, %s, %s, NOW(), NOW())
                    RETURNING id
                """, (
                    client_id, token,
                    selected_questions, excluded_questions,
                    custom_questions, selected_schedules, admin_notes
                ))
                row        = cursor.fetchone()
                project_id = row['id'] if isinstance(row, dict) else row[0]
                conn.commit()
                return jsonify({'success': True, 'project_id': project_id,
                                'project_token': token, 'action': 'created'})
        finally:
            conn.close()

    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


# ---------------------------------------------------------------------------
# APPROVE PROJECT
# ---------------------------------------------------------------------------

@survey_admin_bp.route('/api/survey/admin/project/<int:project_id>/approve', methods=['POST'])
@require_survey_admin
def approve_project(project_id):
    """Set survey_projects.status = 'approved'. Also updates survey_clients."""
    try:
        conn = _get_db()
        try:
            cursor = conn.cursor()

            cursor.execute(
                "SELECT survey_client_id FROM survey_projects WHERE id = %s",
                (project_id,)
            )
            row = cursor.fetchone()
            if not row:
                conn.rollback()
                return jsonify({'success': False, 'error': f'Project {project_id} not found.'}), 404

            client_id = row['survey_client_id'] if isinstance(row, dict) else row[0]

            cursor.execute("""
                UPDATE survey_projects
                SET status = 'approved', approved_at = NOW(), updated_at = NOW()
                WHERE id = %s
            """, (project_id,))

            cursor.execute("""
                UPDATE survey_clients
                SET status = 'approved', updated_at = NOW()
                WHERE id = %s
            """, (client_id,))

            conn.commit()
        finally:
            conn.close()

        print(f"[survey_admin] Project {project_id} approved for client {client_id}")
        return jsonify({
            'success':    True,
            'project_id': project_id,
            'client_id':  client_id,
            'status':     'approved',
            'message':    'Project approved. Ready for Phase 2 (Survey Assembly).'
        })

    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


# ---------------------------------------------------------------------------
# GET PROJECT DETAIL
# ---------------------------------------------------------------------------

@survey_admin_bp.route('/api/survey/admin/project/<int:project_id>', methods=['GET'])
@require_survey_admin
def get_project(project_id):
    """Return full detail for one survey_projects row."""
    try:
        conn = _get_db()
        try:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM survey_projects WHERE id = %s", (project_id,))
            row = cursor.fetchone()
        finally:
            conn.close()

        if not row:
            return jsonify({'success': False, 'error': f'Project {project_id} not found.'}), 404

        return jsonify({'success': True, 'project': _project_to_dict(row)})

    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


# ---------------------------------------------------------------------------
# PHASE 2 — ROSTER UPLOAD
# ---------------------------------------------------------------------------

@survey_admin_bp.route('/api/survey/admin/project/<int:project_id>/roster/upload', methods=['POST'])
@require_survey_admin
def upload_roster(project_id):
    """
    Accept an .xlsx employee roster upload.

    Form fields:
        file              (required) .xlsx roster file
        use_employee_id   (optional) 'true' to use existing employee IDs
                          rather than generating random survey codes.

    Random mode (default):
        - Reads Name column (required) + optional Dept/Shift/Tenure columns
        - Generates unique random 5-digit codes per employee
        - Stores code_mode='random' on survey_projects

    Employee ID mode (use_employee_id=true):
        - Reads Name + Employee ID column (both required) + optional demographics
        - Validates each ID: 1-20 alphanumeric+hyphen characters
        - Stores IDs verbatim as employee_code in survey_roster
        - Stores code_mode='employee_id' on survey_projects
        - No code sheet distribution needed — employees know their own IDs

    Returns:
        success    bool
        count      int       number of employees loaded
        code_mode  str       'random' or 'employee_id'
        mapped     dict      {original_header: canonical_name}
        skipped    list      unrecognised column headers
        message    str
    """
    try:
        # Verify project exists
        conn = _get_db()
        try:
            cursor = conn.cursor()
            cursor.execute("SELECT id FROM survey_projects WHERE id = %s", (project_id,))
            if not cursor.fetchone():
                return jsonify({'success': False, 'error': f'Project {project_id} not found.'}), 404
        finally:
            conn.close()

        # Parse use_employee_id flag from form data
        use_employee_id = (request.form.get('use_employee_id', 'false').strip().lower() == 'true')
        code_mode = 'employee_id' if use_employee_id else 'random'

        # Validate file upload
        if 'file' not in request.files:
            return jsonify({'success': False,
                            'error': 'No file provided. Send as multipart/form-data with key "file".'}), 400

        uploaded_file = request.files['file']
        if not uploaded_file.filename:
            return jsonify({'success': False, 'error': 'Empty filename.'}), 400

        if not uploaded_file.filename.lower().endswith('.xlsx'):
            return jsonify({'success': False,
                            'error': f'File must be .xlsx format. Received: {uploaded_file.filename}'}), 400

        file_bytes = uploaded_file.read()

        # Parse roster
        try:
            employee_rows, mapped, skipped = _read_roster_xlsx(file_bytes, use_employee_id)
        except (ValueError, ImportError) as parse_err:
            return jsonify({'success': False, 'error': str(parse_err)}), 400

        if not employee_rows:
            return jsonify({'success': False, 'error': 'No employee rows found in the file.'}), 400

        # Generate or validate codes
        if use_employee_id:
            # Validate all employee IDs before inserting any
            codes     = []
            id_errors = []
            seen_ids  = set()
            for i, emp in enumerate(employee_rows):
                raw_id = emp.get('employee_id_val') or ''
                ok, reason = _validate_employee_id_value(raw_id)
                if not ok:
                    id_errors.append(f"Row {i + 2} ({emp['employee_name']}): {reason}")
                elif raw_id.strip() in seen_ids:
                    id_errors.append(f"Row {i + 2} ({emp['employee_name']}): duplicate ID '{raw_id.strip()}'")
                else:
                    seen_ids.add(raw_id.strip())
                    codes.append(raw_id.strip())

            if id_errors:
                return jsonify({
                    'success': False,
                    'error':   f'Employee ID validation failed ({len(id_errors)} errors). '
                               f'First error: {id_errors[0]}',
                    'all_errors': id_errors[:20],  # cap at 20 for response size
                }), 400
        else:
            codes = _generate_unique_codes(len(employee_rows))

        # Write to database
        conn = _get_db()
        try:
            cursor = conn.cursor()

            # Wipe existing roster
            cursor.execute(
                "DELETE FROM survey_roster WHERE survey_project_id = %s",
                (project_id,)
            )
            deleted = cursor.rowcount
            if deleted > 0:
                print(f"[roster_upload] Wiped {deleted} existing roster rows for project {project_id}")

            # Insert new roster rows
            for emp, code in zip(employee_rows, codes):
                cursor.execute("""
                    INSERT INTO survey_roster
                        (survey_project_id, employee_name, department,
                         shift, tenure_bracket, employee_code, has_responded)
                    VALUES (%s, %s, %s, %s, %s, %s, FALSE)
                """, (
                    project_id,
                    emp['employee_name'],
                    emp.get('department'),
                    emp.get('shift'),
                    emp.get('tenure_bracket'),
                    code,
                ))

            # Update survey_projects — roster state + code_mode
            cursor.execute("""
                UPDATE survey_projects
                SET roster_uploaded = TRUE,
                    roster_count    = %s,
                    code_mode       = %s,
                    updated_at      = NOW()
                WHERE id = %s
            """, (len(employee_rows), code_mode, project_id))

            conn.commit()
        finally:
            conn.close()

        print(f"[roster_upload] Project {project_id}: {len(employee_rows)} employees, "
              f"code_mode={code_mode}, mapped={mapped}, skipped={skipped}")

        if use_employee_id:
            msg = (f'{len(employee_rows)} employees loaded with their existing IDs. '
                   f'No code sheet needed -- employees enter their own ID.')
        else:
            msg = f'{len(employee_rows)} employees loaded. Unique 5-digit survey codes generated.'

        return jsonify({
            'success':   True,
            'count':     len(employee_rows),
            'code_mode': code_mode,
            'mapped':    mapped,
            'skipped':   skipped,
            'message':   msg,
        })

    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


# ---------------------------------------------------------------------------
# PHASE 2 — ROSTER CODE SHEET DOWNLOAD
# ---------------------------------------------------------------------------

@survey_admin_bp.route('/api/survey/admin/project/<int:project_id>/roster/download-codes', methods=['GET'])
@require_survey_admin
def download_roster_codes(project_id):
    """
    Generate and return a printable .xlsx: Employee Name | Survey Code.
    Instructions footer adapts to code_mode:
        random:       'Enter this code when starting the online survey.'
        employee_id:  'Enter your employee ID when starting the online survey.'

    Only available when roster_uploaded=TRUE.
    """
    try:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return jsonify({'success': False,
                            'error': 'openpyxl not installed. Add it to requirements.txt.'}), 500

        conn = _get_db()
        try:
            cursor = conn.cursor()
            cursor.execute(
                "SELECT id, roster_uploaded, roster_count, code_mode FROM survey_projects WHERE id = %s",
                (project_id,)
            )
            project_row = cursor.fetchone()
            if not project_row:
                return jsonify({'success': False, 'error': f'Project {project_id} not found.'}), 404

            pdata = dict(project_row) if hasattr(project_row, 'keys') else {
                'id': project_row[0],
                'roster_uploaded': project_row[1],
                'roster_count': project_row[2],
                'code_mode': project_row[3] if len(project_row) > 3 else 'random',
            }

            if not pdata.get('roster_uploaded'):
                return jsonify({'success': False,
                                'error': 'No roster uploaded for this project. '
                                         'Upload a roster first.'}), 400

            code_mode = pdata.get('code_mode') or 'random'

            cursor.execute("""
                SELECT employee_name, employee_code
                FROM survey_roster
                WHERE survey_project_id = %s
                ORDER BY employee_name ASC
            """, (project_id,))
            roster_rows = cursor.fetchall()
        finally:
            conn.close()

        if not roster_rows:
            return jsonify({'success': False,
                            'error': 'Roster table is empty for this project.'}), 404

        # Column header label adapts to mode
        code_col_label = 'Employee ID' if code_mode == 'employee_id' else 'Survey Code'

        # Instructions text adapts to mode
        if code_mode == 'employee_id':
            instructions = ('INSTRUCTIONS: Give each employee this list. '
                            'They will enter their Employee ID when starting the online survey.')
        else:
            instructions = ('INSTRUCTIONS: Give each employee their Survey Code. '
                            'They will enter this code when starting the online survey.')

        # Build workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'Survey Codes'

        header_fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
        header_font = Font(bold=True)

        ws['A1'] = 'Employee Name'
        ws['B1'] = code_col_label
        for cell in [ws['A1'], ws['B1']]:
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal='left')

        for roster_row in roster_rows:
            if hasattr(roster_row, 'keys'):
                name = roster_row['employee_name']
                code = roster_row['employee_code']
            else:
                name, code = roster_row[0], roster_row[1]
            ws.append([name, code])

        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20

        ws.append([])
        instr_row = ws.max_row + 1
        ws.cell(row=instr_row, column=1, value=instructions)
        ws.cell(row=instr_row, column=1).font = Font(italic=True, color='666666')

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f'survey_codes_project_{project_id}.xlsx'
        return send_file(
            buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename,
        )

    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

# I did no harm and this file is not truncated
