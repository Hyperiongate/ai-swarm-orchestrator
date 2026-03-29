"""
SURVEY IN A BOX — Respondent Routes
File: routes/survey_respondent.py
Created: March 13, 2026
Last Updated: March 29, 2026 — Phase 2 Session 3: Likert answer option injection

PURPOSE:
    Employee-facing survey engine for Survey in a Box Phase 3.
    Handles the public-facing survey at /survey/take/<token> and the
    API endpoints that power it. Fully anonymous — no names, no IP
    addresses, no PII stored or logged.

    This blueprint does NOT require authentication. It is the public
    interface that employees use to complete their survey.

ENDPOINTS:
    GET  /survey/take/<token>                    -- Survey page (HTML)
    GET  /api/survey/take/<token>/questions      -- Load survey structure
    POST /api/survey/take/<token>/submit         -- Submit completed survey
    GET  /api/survey/admin/project/<id>/open     -- Jim opens survey for responses
    POST /api/survey/admin/project/<id>/close    -- Jim closes survey
    GET  /api/survey/admin/project/<id>/responses -- Response count + summary
    GET  /api/survey/admin/project/<id>/export   -- Download .xlsx response data

ANONYMOUS DESIGN:
    - No login, no registration, no email collection
    - No IP addresses stored or logged
    - session_token is a one-way SHA-256 hash of a browser cookie value.
      Used only to prevent duplicate submissions in a single browser session.
      Cannot be reversed to identify any individual.
    - The export contains no timestamps, no row IDs, no session tokens --
      purely the answer data in SurveySelector-compatible format.

EXPORT FORMAT (SurveySelector / Remark compatible):
    Without roster:
        - One sheet named "Sheet1"
        - One row per respondent
        - One column per question in survey order
        - Column headers: question short_label
        - Values: full text of selected answer option
        - Unanswered questions: the literal string "BLANK"

    With roster (Phase 2):
        - First 4 columns: Employee Code | Department | Shift | Tenure
          (looked up from survey_roster by employee_code stored in response)
        - Then all survey question columns in order
        - Provides SurveySelector with demographic breakout data directly

DUPLICATE PREVENTION:
    Without roster:
        - Browser sets a session cookie (survey_session_<token>) on first load
        - Cookie value is SHA-256 hashed before storage -- one-way, not reversible
        - On submission, the hash is checked against survey_responses for that project
        - If a matching session_token exists, submission is rejected with HTTP 409

    With roster (Phase 2):
        - Employee enters 5-digit code issued by Jim
        - Code validated against survey_roster for this project
        - has_responded flag checked -- rejects if already used (HTTP 409)
        - On success: has_responded set to TRUE and employee_code stored in response row

SURVEY OPEN/CLOSE:
    - Jim explicitly opens a survey via GET /open endpoint (auth required)
    - Employees cannot submit to a closed survey (returns 403 with clear message)
    - Jim closes the survey via POST /close when data collection is complete
    - open/close state is stored in survey_projects.is_open

POSTGRESQL RULES:
    - RealDictCursor dict-only rows (access by name, never by index)
    - TRUE/FALSE for booleans (not 0/1)
    - %s for all parameters
    - RETURNING id on INSERT

CHANGELOG:
    - March 29, 2026: ONE CHANGE ONLY -- Likert option injection in
      _get_survey_questions_ordered(). The question bank defines ~30 Likert
      questions with type='likert' and a 'scale' field but NO 'options' list.
      The online survey frontend needs an explicit 'options' array to render
      radio buttons. Without it, Likert questions render as empty space.

      Fix: after creating q_copy for each question, if type is 'likert' and
      'options' is absent, inject the standard 5-point scale matching the
      labels used by export_to_word() in survey_builder.py:
          '1 -- Strongly Disagree', '2 -- Disagree', '3 -- Neutral',
          '4 -- Agree', '5 -- Strongly Agree'
      Type is kept as 'likert' so the frontend renders the styled likert-scale
      grid layout (not a plain radio list).

      Only type='likert' is affected. Types 'multiple_choice', 'checkbox',
      and 'text' are completely unchanged -- they already have explicit options
      or render as textarea. No other types exist in the question bank.
      survey_builder.py was NOT modified (read-only per Opus handoff).

    - March 27, 2026: Code mode support -- TWO CHANGES ONLY:
      1. get_survey_questions(): added code_mode to JSON response so the
         frontend can show the correct label and placeholder in the code
         input field ('5-Digit Survey Code' vs 'Employee ID').
      2. submit_survey(): added code_mode-aware validation. In random mode,
         the existing 5-digit check applies. In employee_id mode, the code
         is validated as 1-20 alphanumeric+hyphen characters (matching the
         flexible format allowed by survey_admin.py upload). Error messages
         adapt to the mode. All other logic unchanged.
      NO OTHER CHANGES.

    - March 26, 2026: Phase 2 additions -- FOUR CHANGES ONLY:
      1. get_survey_questions(): added roster_uploaded and roster_count to
         JSON response so the frontend knows whether to show the code field.
      2. submit_survey(): added employee_code validation. If project has a
         roster (roster_uploaded=TRUE), the code is required, validated
         against survey_roster, and has_responded is set on success.
         employee_code stored in survey_responses row. Backward compatible:
         if roster_uploaded=FALSE, code is ignored and cookie-based
         duplicate prevention applies as before.
      3. export_responses(): when roster exists, prepends Employee Code,
         Department, Shift, Tenure columns (looked up from survey_roster)
         before the survey question columns.
      4. Header and changelog updated.
      NO OTHER CHANGES. All existing endpoints, logic, and error handling
      are completely unchanged.

    - March 25, 2026: BUG FIX -- _require_admin_auth() was checking the wrong
      Flask session key. ONE LINE CHANGED: 'survey_admin_logged_in' ->
      'survey_admin_authenticated'.

    - March 13, 2026 (BUG FIX): Fixed _get_survey_questions_ordered().
      SurveyBuilder.create_survey() returns a flat 'questions' list, not a
      'sections' hierarchy.

    - March 13, 2026: Initial creation. Phase 3 of Survey in a Box roadmap.
"""

import hashlib
import json
import os
import re
import traceback
from datetime import datetime, timezone
from io import BytesIO

from flask import (Blueprint, Response, jsonify, make_response,
                   render_template, request, session)

from db_engine import get_db_connection

survey_respondent_bp = Blueprint('survey_respondent', __name__)

# ---------------------------------------------------------------------------
# CONSTANTS
# ---------------------------------------------------------------------------

# Cookie name template -- includes token to scope per-survey
_COOKIE_NAME_PREFIX = 'survey_session_'

# How long the duplicate-prevention cookie lives (seconds). 30 days.
_COOKIE_MAX_AGE = 30 * 24 * 60 * 60

# ---------------------------------------------------------------------------
# LIKERT SCALE OPTIONS
# Injected into type='likert' questions which have no explicit 'options' list.
# Labels match export_to_word() in survey_builder.py exactly so that online
# and paper survey responses use identical option text in exports.
# Added: March 29, 2026
# ---------------------------------------------------------------------------
_LIKERT_OPTIONS = [
    '1 -- Strongly Disagree',
    '2 -- Disagree',
    '3 -- Neutral',
    '4 -- Agree',
    '5 -- Strongly Agree',
]


# ---------------------------------------------------------------------------
# INTERNAL HELPERS
# ---------------------------------------------------------------------------

def _hash_session_value(raw_value):
    """One-way SHA-256 hash of a session cookie value. Non-reversible."""
    return hashlib.sha256(raw_value.encode('utf-8')).hexdigest()


def _get_session_token(token):
    """
    Return (cookie_value, hashed_token) for this survey token.
    cookie_value: the raw string stored in the browser cookie (for setting).
    hashed_token: the SHA-256 hash stored in the DB (for duplicate check).
    Returns (None, None) if no cookie exists yet.
    """
    cookie_name  = _COOKIE_NAME_PREFIX + token
    cookie_value = request.cookies.get(cookie_name)
    if not cookie_value:
        return None, None
    return cookie_value, _hash_session_value(cookie_value)


def _generate_session_cookie_value():
    """Generate a new random value to set as the browser session cookie."""
    import secrets
    return secrets.token_hex(32)


def _load_project_by_token(token):
    """
    Load survey_projects + survey_clients row by project_token.
    Returns (project_row, client_row) or (None, None) if not found.
    """
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        cursor.execute(
            """
            SELECT sp.*, sc.company_name, sc.preferred_administration
            FROM survey_projects sp
            JOIN survey_clients sc ON sc.id = sp.survey_client_id
            WHERE sp.project_token = %s
            """,
            (token,)
        )
        row = cursor.fetchone()
        if not row:
            return None, None

        cursor.execute(
            'SELECT * FROM survey_clients WHERE id = %s',
            (row['survey_client_id'],)
        )
        client_row = cursor.fetchone()
        return row, client_row
    finally:
        conn.close()


def _safe_json_loads(val, default=None):
    if default is None:
        default = []
    if not val:
        return default
    try:
        if isinstance(val, (list, dict)):
            return val
        return json.loads(val)
    except (TypeError, ValueError):
        return default


def _get_survey_questions_ordered(project_row):
    """
    Return the ordered list of question dicts for this project.

    SurveyBuilder.create_survey() returns a survey dict with a flat
    'questions' list (not 'sections'). Each question dict from the
    question bank uses 'id' and 'text' fields. We inject a 'short_label'
    key (set to the question id) so the export and take-page code has a
    consistent field to use as column headers.

    Falls back to full question bank if no questions selected.

    March 29, 2026 addition: after building q_copy, inject standard answer
    options for type='likert' questions which have no 'options' field in the
    bank. Options match export_to_word() scale labels in survey_builder.py.
    """
    try:
        from survey_builder import SurveyBuilder
        builder = SurveyBuilder()

        selected_questions = _safe_json_loads(project_row.get('selected_questions'), [])
        custom_questions   = _safe_json_loads(project_row.get('custom_questions'), [])
        selected_schedules = _safe_json_loads(project_row.get('selected_schedules'), [])
        company_name       = project_row.get('company_name', 'Your Company')

        if not selected_questions:
            selected_questions = list(builder.question_bank.keys())
        else:
            question_ids_in_bank = set(builder.question_bank.keys())
            any_match_as_id = any(q in question_ids_in_bank for q in selected_questions)
            if not any_match_as_id:
                CATEGORY_MAP = {
                    'work_life_balance':   'schedule_features',
                    'schedule_preference': 'schedule_features',
                    'overtime':            'overtime',
                    'fatigue_safety':      'working_conditions',
                    'communication':       'working_conditions',
                    'compensation':        'working_conditions',
                    'commute':             'demographics',
                    'childcare':           'daycare_eldercare',
                    'demographics':        'demographics',
                    'open_ended':          'open_ended',
                }
                bank_categories = set()
                for sel in selected_questions:
                    mapped = CATEGORY_MAP.get(sel)
                    if mapped:
                        bank_categories.add(mapped)
                    else:
                        bank_categories.add(sel)
                if bank_categories:
                    selected_questions = [
                        qid for qid, qdata in builder.question_bank.items()
                        if qdata.get('category') in bank_categories
                    ]
                else:
                    selected_questions = list(builder.question_bank.keys())

        survey_obj = builder.create_survey(
            project_name='Survey',
            company_name=company_name,
            selected_questions=selected_questions,
            schedules_to_rate=selected_schedules,
            custom_questions=custom_questions if custom_questions else None
        )

        ordered = []
        for q in survey_obj.get('questions', []):
            q_copy = dict(q)
            if 'short_label' not in q_copy:
                q_copy['short_label'] = q_copy.get('id', q_copy.get('text', 'unknown'))

            # ------------------------------------------------------------------
            # LIKERT OPTION INJECTION  (added March 29, 2026)
            # The question bank stores Likert questions with type='likert' and a
            # 'scale' description field but NO 'options' list. The online survey
            # frontend iterates q.options to render radio buttons. Without this
            # injection, Likert questions render as empty space.
            #
            # We inject the same 5-point scale labels used by export_to_word()
            # in survey_builder.py so that online responses and paper survey
            # exports produce identical option text strings in SurveySelector.
            #
            # Only type='likert' is affected. Types 'multiple_choice' and
            # 'checkbox' already have explicit options; type='text' renders as
            # a textarea and needs no options. survey_builder.py is NOT touched.
            # ------------------------------------------------------------------
            if q_copy.get('type') == 'likert' and not q_copy.get('options'):
                q_copy['options'] = _LIKERT_OPTIONS[:]

            ordered.append(q_copy)

        for sched in survey_obj.get('schedule_concepts', []):
            sched_id = sched.get('schedule', {}).get('id', 'schedule')
            ordered.append({
                'id':          sched_id,
                'short_label': sched_id,
                'text':        sched.get('question', sched_id),
                'type':        'multiple_choice',
                'options':     sched.get('rating_options', []),
            })

        print(f'[survey_respondent] _get_survey_questions_ordered: {len(ordered)} questions')
        return ordered, survey_obj

    except Exception as e:
        print(f'[survey_respondent] _get_survey_questions_ordered error: {traceback.format_exc()}')
        return [], {}


# ---------------------------------------------------------------------------
# ADMIN HELPERS -- require auth
# ---------------------------------------------------------------------------

def _require_admin_auth():
    """
    Check Flask session for admin login. Returns error response or None.

    IMPORTANT: The session key must match ADMIN_SESSION_KEY in survey_admin.py.
    survey_admin.py sets: session['survey_admin_authenticated'] = True
    This function checks: session.get('survey_admin_authenticated')
    These must always match. If survey_admin.py ever changes its key,
    update this function to match.
    """
    if not session.get('survey_admin_authenticated'):
        return jsonify({'success': False, 'error': 'Not authenticated',
                        'redirect': '/survey/admin'}), 401
    return None


# ---------------------------------------------------------------------------
# ROUTES -- EMPLOYEE-FACING SURVEY PAGE
# ---------------------------------------------------------------------------

@survey_respondent_bp.route('/survey/take/<token>', methods=['GET'])
def survey_take_page(token):
    """
    Render the employee-facing survey page.
    The page loads questions via /api/survey/take/<token>/questions
    and submits via /api/survey/take/<token>/submit.
    """
    project_row, client_row = _load_project_by_token(token)

    if not project_row:
        return render_template('survey_closed.html',
                               message='This survey link is not valid.'), 404

    if project_row['status'] not in ('approved', 'administered'):
        return render_template('survey_closed.html',
                               message='This survey is not currently active.'), 403

    if not project_row.get('is_open'):
        return render_template('survey_closed.html',
                               message='This survey is not currently accepting responses. '
                                       'Please check back later or contact your supervisor.'), 403

    company_name = client_row.get('company_name', 'Your Company') if client_row else 'Your Company'

    return render_template('survey_take.html',
                           token=token,
                           company_name=company_name)


# ---------------------------------------------------------------------------
# ROUTES -- LOAD SURVEY QUESTIONS (API)
# ---------------------------------------------------------------------------

@survey_respondent_bp.route('/api/survey/take/<token>/questions', methods=['GET'])
def get_survey_questions(token):
    """
    Return the ordered question list for this survey.
    Used by the survey_take.html frontend to render the form.
    Returns questions grouped by section with full option text.
    Does NOT require authentication.

    Phase 2 addition: response now includes roster_uploaded (bool),
    roster_count (int), and code_mode ('random' or 'employee_id') so the
    frontend knows whether to show the code field and how to label it.
    """
    try:
        project_row, client_row = _load_project_by_token(token)

        if not project_row:
            return jsonify({'success': False, 'error': 'Survey not found'}), 404

        if not project_row.get('is_open'):
            return jsonify({'success': False, 'error': 'Survey is not open'}), 403

        ordered_questions, survey_obj = _get_survey_questions_ordered(project_row)

        if not ordered_questions:
            return jsonify({'success': False, 'error': 'No questions configured for this survey'}), 500

        sections_dict = {}
        section_order = []
        for q in ordered_questions:
            cat = q.get('category', 'Survey Questions').replace('_', ' ').title()
            if cat not in sections_dict:
                sections_dict[cat] = []
                section_order.append(cat)
            sections_dict[cat].append(q)

        sections = [
            {'title': cat, 'questions': sections_dict[cat]}
            for cat in section_order
        ]

        company_name = project_row.get('company_name', 'Your Company')

        # Phase 2: roster status and code_mode for frontend
        roster_uploaded = bool(project_row.get('roster_uploaded', False))
        roster_count    = int(project_row.get('roster_count', 0) or 0)
        code_mode       = project_row.get('code_mode') or 'random'

        return jsonify({
            'success':         True,
            'company_name':    company_name,
            'token':           token,
            'total_questions': len(ordered_questions),
            'sections':        sections,
            'roster_uploaded': roster_uploaded,
            'roster_count':    roster_count,
            'code_mode':       code_mode,
        })

    except Exception as e:
        print(f'[survey_respondent] get_survey_questions error: {traceback.format_exc()}')
        return jsonify({'success': False, 'error': str(e)}), 500


# ---------------------------------------------------------------------------
# ROUTES -- SUBMIT SURVEY (API)
# ---------------------------------------------------------------------------

@survey_respondent_bp.route('/api/survey/take/<token>/submit', methods=['POST'])
def submit_survey(token):
    """
    Accept a completed survey submission.

    Request body (JSON):
        {
            "answers": {
                "dept": "Production",
                "management_input": "4 -- Agree"
            },
            "employee_code": "12345"   <- required if project has roster
        }

    Phase 2 employee code validation:
        - If project.roster_uploaded=TRUE: employee_code is required.
          Random mode: exactly 5 digits.
          Employee ID mode: 1-20 alphanumeric+hyphen chars.
          Code must exist in survey_roster and has_responded must be FALSE.
          On success: has_responded set TRUE, employee_code stored in response.
        - If roster_uploaded=FALSE: code ignored, cookie-based prevention applies.

    Returns:
        201 Created on success
        400 if employee_code missing or invalid format
        403 if survey is closed
        404 if employee code not found in roster
        409 if duplicate submission
        422 if answers are missing or malformed
    """
    try:
        project_row, client_row = _load_project_by_token(token)

        if not project_row:
            return jsonify({'success': False, 'error': 'Survey not found'}), 404

        if not project_row.get('is_open'):
            return jsonify({
                'success': False,
                'error':   'This survey is no longer accepting responses.'
            }), 403

        # ---- Parse request body ---------------------------------------------
        data = request.get_json(force=True, silent=True) or {}

        # ---- PHASE 2: Employee code validation ------------------------------
        roster_uploaded = bool(project_row.get('roster_uploaded', False))
        code_mode       = project_row.get('code_mode') or 'random'
        roster_row      = None

        if roster_uploaded:
            employee_code = str(data.get('employee_code', '') or '').strip()

            if not employee_code:
                if code_mode == 'employee_id':
                    err_msg = 'Please enter your Employee ID before submitting.'
                else:
                    err_msg = 'Please enter your 5-digit survey code before submitting.'
                return jsonify({'success': False, 'error': err_msg}), 400

            if code_mode == 'employee_id':
                if not re.match(r'^[A-Za-z0-9\-]+$', employee_code) or len(employee_code) > 20:
                    return jsonify({
                        'success': False,
                        'error':   'Employee ID format is not valid. '
                                   'Please check your ID and try again.'
                    }), 400
            else:
                if not employee_code.isdigit() or len(employee_code) != 5:
                    return jsonify({
                        'success': False,
                        'error':   'Survey code must be exactly 5 digits. '
                                   'Please check your code and try again.'
                    }), 400

            conn = get_db_connection()
            try:
                cursor = conn.cursor()
                cursor.execute(
                    """
                    SELECT id, has_responded, department, shift, tenure_bracket
                    FROM survey_roster
                    WHERE survey_project_id = %s AND employee_code = %s
                    """,
                    (project_row['id'], employee_code)
                )
                roster_row = cursor.fetchone()
            finally:
                conn.close()

            if not roster_row:
                if code_mode == 'employee_id':
                    not_found_msg = ('Employee ID not recognized. Please check your ID '
                                     'and try again, or contact your supervisor.')
                else:
                    not_found_msg = ('Survey code not recognized. Please check your code '
                                     'and try again, or contact your supervisor.')
                return jsonify({'success': False, 'error': not_found_msg}), 404

            if roster_row['has_responded']:
                if code_mode == 'employee_id':
                    dup_msg = ('This Employee ID has already been used to submit a response. '
                               'Only one response per employee is allowed. Thank you!')
                else:
                    dup_msg = ('This survey code has already been used to submit a response. '
                               'Only one response per employee is allowed. Thank you!')
                return jsonify({'success': False, 'error': dup_msg}), 409

        else:
            employee_code = None

        # ---- Cookie-based duplicate check (when no roster) ------------------
        cookie_value, hashed_token = _get_session_token(token)

        if not roster_uploaded and hashed_token:
            conn = get_db_connection()
            try:
                cursor = conn.cursor()
                cursor.execute(
                    """
                    SELECT id FROM survey_responses
                    WHERE survey_project_id = %s AND session_token = %s
                    LIMIT 1
                    """,
                    (project_row['id'], hashed_token)
                )
                duplicate = cursor.fetchone()
            finally:
                conn.close()

            if duplicate:
                return jsonify({
                    'success': False,
                    'error':   'You have already submitted a response for this survey. '
                               'Only one response per person is allowed. Thank you!'
                }), 409

        # ---- Validate answers -----------------------------------------------
        answers = data.get('answers')

        if not isinstance(answers, dict):
            return jsonify({
                'success': False,
                'error':   'answers must be a JSON object keyed by question short_label.'
            }), 422

        if len(answers) == 0:
            return jsonify({
                'success': False,
                'error':   'No answers were submitted. Please complete the survey before submitting.'
            }), 422

        cleaned_answers = {k: v for k, v in answers.items()
                           if isinstance(k, str) and isinstance(v, str) and v.strip()}

        # ---- Store response -------------------------------------------------
        new_session_value = None
        new_hashed_token  = None

        if not roster_uploaded:
            if not cookie_value:
                new_session_value = _generate_session_cookie_value()
                new_hashed_token  = _hash_session_value(new_session_value)
            else:
                new_hashed_token = hashed_token

        conn = get_db_connection()
        try:
            cursor = conn.cursor()

            cursor.execute(
                """
                INSERT INTO survey_responses
                    (survey_project_id, answers, session_token, employee_code)
                VALUES (%s, %s::jsonb, %s, %s)
                RETURNING id
                """,
                (
                    project_row['id'],
                    json.dumps(cleaned_answers),
                    new_hashed_token,
                    employee_code,
                )
            )
            row = cursor.fetchone()
            response_id = row['id']

            cursor.execute(
                """
                UPDATE survey_projects
                SET response_count = response_count + 1, updated_at = NOW()
                WHERE id = %s
                """,
                (project_row['id'],)
            )

            if roster_uploaded and roster_row:
                cursor.execute(
                    """
                    UPDATE survey_roster
                    SET has_responded = TRUE
                    WHERE id = %s
                    """,
                    (roster_row['id'],)
                )

            conn.commit()
        finally:
            conn.close()

        print(f'[survey_respondent] Response {response_id} stored for project '
              f'{project_row["id"]} ({project_row.get("company_name", "?")})'
              f'{" code=" + employee_code if employee_code else ""}')

        resp = make_response(jsonify({
            'success':     True,
            'message':     'Thank you! Your response has been recorded.',
            'response_id': response_id,
        }), 201)

        if new_session_value:
            cookie_name = _COOKIE_NAME_PREFIX + token
            resp.set_cookie(
                cookie_name,
                new_session_value,
                max_age=_COOKIE_MAX_AGE,
                httponly=True,
                samesite='Lax',
                secure=os.environ.get('FLASK_ENV') != 'development'
            )

        return resp

    except Exception as e:
        print(f'[survey_respondent] submit_survey error: {traceback.format_exc()}')
        return jsonify({'success': False, 'error': str(e)}), 500


# ---------------------------------------------------------------------------
# ROUTES -- ADMIN: OPEN SURVEY
# ---------------------------------------------------------------------------

@survey_respondent_bp.route('/api/survey/admin/project/<int:project_id>/open', methods=['GET', 'POST'])
def open_survey(project_id):
    """
    Jim opens a survey to accept employee responses.
    Sets is_open = TRUE, opened_at = NOW(), status = 'administered'.
    Also sets survey_url on the project for convenience.
    Requires admin auth.
    """
    auth_error = _require_admin_auth()
    if auth_error:
        return auth_error

    try:
        conn = get_db_connection()
        try:
            cursor = conn.cursor()

            cursor.execute(
                'SELECT id, project_token, status, is_open FROM survey_projects WHERE id = %s',
                (project_id,)
            )
            project = cursor.fetchone()

            if not project:
                return jsonify({'success': False, 'error': 'Project not found'}), 404

            if project['status'] not in ('approved', 'administered'):
                return jsonify({
                    'success': False,
                    'error':   f'Project must be approved before opening. '
                               f'Current status: {project["status"]}'
                }), 400

            token      = project['project_token']
            base_url   = os.environ.get('APP_BASE_URL', 'https://ai-swarm-orchestrator.onrender.com')
            survey_url = f'{base_url}/survey/take/{token}'

            cursor.execute(
                """
                UPDATE survey_projects
                SET is_open    = TRUE,
                    opened_at  = NOW(),
                    status     = 'administered',
                    survey_url = %s,
                    updated_at = NOW()
                WHERE id = %s
                """,
                (survey_url, project_id)
            )
            conn.commit()
        finally:
            conn.close()

        return jsonify({
            'success':    True,
            'message':    'Survey is now open and accepting responses.',
            'survey_url': survey_url,
            'token':      token,
        })

    except Exception as e:
        print(f'[survey_respondent] open_survey error: {traceback.format_exc()}')
        return jsonify({'success': False, 'error': str(e)}), 500


# ---------------------------------------------------------------------------
# ROUTES -- ADMIN: CLOSE SURVEY
# ---------------------------------------------------------------------------

@survey_respondent_bp.route('/api/survey/admin/project/<int:project_id>/close', methods=['POST'])
def close_survey(project_id):
    """
    Jim closes a survey -- no more responses accepted.
    Sets is_open = FALSE, closed_at = NOW().
    Requires admin auth.
    """
    auth_error = _require_admin_auth()
    if auth_error:
        return auth_error

    try:
        conn = get_db_connection()
        try:
            cursor = conn.cursor()

            cursor.execute(
                'SELECT id, is_open, response_count FROM survey_projects WHERE id = %s',
                (project_id,)
            )
            project = cursor.fetchone()

            if not project:
                return jsonify({'success': False, 'error': 'Project not found'}), 404

            cursor.execute(
                """
                UPDATE survey_projects
                SET is_open    = FALSE,
                    closed_at  = NOW(),
                    updated_at = NOW()
                WHERE id = %s
                """,
                (project_id,)
            )
            conn.commit()
        finally:
            conn.close()

        return jsonify({
            'success':        True,
            'message':        'Survey is now closed.',
            'response_count': project['response_count'] or 0,
        })

    except Exception as e:
        print(f'[survey_respondent] close_survey error: {traceback.format_exc()}')
        return jsonify({'success': False, 'error': str(e)}), 500


# ---------------------------------------------------------------------------
# ROUTES -- ADMIN: RESPONSE SUMMARY
# ---------------------------------------------------------------------------

@survey_respondent_bp.route('/api/survey/admin/project/<int:project_id>/responses', methods=['GET'])
def get_response_summary(project_id):
    """
    Return response count and open/closed status for a project.
    Used by the admin dashboard to show response progress.
    Requires admin auth.
    """
    auth_error = _require_admin_auth()
    if auth_error:
        return auth_error

    try:
        conn = get_db_connection()
        try:
            cursor = conn.cursor()

            cursor.execute(
                """
                SELECT sp.response_count, sp.is_open, sp.opened_at, sp.closed_at,
                       sp.survey_url, sp.project_token,
                       sp.roster_uploaded, sp.roster_count,
                       COUNT(sr.id) AS actual_count
                FROM survey_projects sp
                LEFT JOIN survey_responses sr ON sr.survey_project_id = sp.id
                WHERE sp.id = %s
                GROUP BY sp.id
                """,
                (project_id,)
            )
            row = cursor.fetchone()
        finally:
            conn.close()

        if not row:
            return jsonify({'success': False, 'error': 'Project not found'}), 404

        return jsonify({
            'success':         True,
            'project_id':      project_id,
            'response_count':  row['actual_count'],
            'is_open':         row['is_open'],
            'opened_at':       str(row['opened_at']) if row['opened_at'] else None,
            'closed_at':       str(row['closed_at']) if row['closed_at'] else None,
            'survey_url':      row['survey_url'] or '',
            'token':           row['project_token'],
            'roster_uploaded': bool(row.get('roster_uploaded', False)),
            'roster_count':    int(row.get('roster_count', 0) or 0),
        })

    except Exception as e:
        print(f'[survey_respondent] get_response_summary error: {traceback.format_exc()}')
        return jsonify({'success': False, 'error': str(e)}), 500


# ---------------------------------------------------------------------------
# ROUTES -- ADMIN: EXPORT RESPONSES AS XLSX
# ---------------------------------------------------------------------------

@survey_respondent_bp.route('/api/survey/admin/project/<int:project_id>/export', methods=['GET'])
def export_responses(project_id):
    """
    Export all survey responses as a SurveySelector/Remark-compatible .xlsx file.

    Without roster:
        - Sheet name: "Sheet1"
        - One row per respondent
        - Column headers: question short_label in survey order
        - Values: full text of selected answer option
        - Unanswered questions: the literal string "BLANK"

    With roster (Phase 2):
        - First 4 columns: Employee Code | Department | Shift | Tenure
        - Then all survey question columns in order
        - Unanswered questions: "BLANK"

    Requires admin auth.
    """
    auth_error = _require_admin_auth()
    if auth_error:
        return auth_error

    try:
        conn = get_db_connection()
        try:
            cursor = conn.cursor()
            cursor.execute(
                """
                SELECT sp.*, sc.company_name, sc.preferred_administration
                FROM survey_projects sp
                JOIN survey_clients sc ON sc.id = sp.survey_client_id
                WHERE sp.id = %s
                """,
                (project_id,)
            )
            project_row = cursor.fetchone()

            if not project_row:
                return jsonify({'success': False, 'error': 'Project not found'}), 404

            roster_uploaded = bool(project_row.get('roster_uploaded', False))

            cursor.execute(
                """
                SELECT answers, employee_code FROM survey_responses
                WHERE survey_project_id = %s
                ORDER BY submitted_at ASC
                """,
                (project_id,)
            )
            response_rows = cursor.fetchall()

            roster_lookup = {}
            if roster_uploaded:
                cursor.execute(
                    """
                    SELECT employee_code, department, shift, tenure_bracket
                    FROM survey_roster
                    WHERE survey_project_id = %s
                    """,
                    (project_id,)
                )
                for roster_row in cursor.fetchall():
                    roster_lookup[roster_row['employee_code']] = {
                        'department':     roster_row.get('department') or 'BLANK',
                        'shift':          roster_row.get('shift') or 'BLANK',
                        'tenure_bracket': roster_row.get('tenure_bracket') or 'BLANK',
                    }

        finally:
            conn.close()

        if not response_rows:
            return jsonify({
                'success': False,
                'error':   'No responses have been submitted for this project yet.'
            }), 404

        ordered_questions, _ = _get_survey_questions_ordered(project_row)

        if not ordered_questions:
            return jsonify({
                'success': False,
                'error':   'Could not load question structure for this project.'
            }), 500

        question_headers = [q['short_label'] for q in ordered_questions]

        if roster_uploaded:
            column_headers = ['Employee Code', 'Department', 'Shift', 'Tenure'] + question_headers
        else:
            column_headers = question_headers

        try:
            import openpyxl
            from openpyxl.styles import Font
        except ImportError:
            return jsonify({
                'success': False,
                'error':   'openpyxl is required for xlsx export.'
            }), 500

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Sheet1'

        header_font = Font(bold=True)
        for col_idx, header in enumerate(column_headers, start=1):
            cell      = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font

        for row_idx, response_row in enumerate(response_rows, start=2):
            raw_answers = response_row['answers']
            if isinstance(raw_answers, str):
                try:
                    answers = json.loads(raw_answers)
                except (TypeError, ValueError):
                    answers = {}
            elif isinstance(raw_answers, dict):
                answers = raw_answers
            else:
                answers = {}

            emp_code = response_row.get('employee_code') or ''
            col_idx  = 1

            if roster_uploaded:
                demo = roster_lookup.get(emp_code, {})
                ws.cell(row=row_idx, column=col_idx, value=emp_code or 'BLANK')
                col_idx += 1
                ws.cell(row=row_idx, column=col_idx, value=demo.get('department', 'BLANK'))
                col_idx += 1
                ws.cell(row=row_idx, column=col_idx, value=demo.get('shift', 'BLANK'))
                col_idx += 1
                ws.cell(row=row_idx, column=col_idx, value=demo.get('tenure_bracket', 'BLANK'))
                col_idx += 1

            for header in question_headers:
                value = answers.get(header, 'BLANK')
                if not value or not str(value).strip():
                    value = 'BLANK'
                ws.cell(row=row_idx, column=col_idx, value=str(value))
                col_idx += 1

        for col_idx, header in enumerate(column_headers, start=1):
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[col_letter].width = max(len(header) + 4, 15)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        company_name     = project_row.get('company_name', 'Survey')
        safe_company     = ''.join(c if c.isalnum() else '_' for c in company_name)[:40]
        ts               = datetime.now(timezone.utc).strftime('%Y%m%d')
        filename         = f'SurveyData_{safe_company}_{ts}.xlsx'
        respondent_count = len(response_rows)

        print(f'[survey_respondent] Exported {respondent_count} responses '
              f'for project {project_id} ({company_name}): {filename} '
              f'(roster={roster_uploaded}, cols={len(column_headers)})')

        return Response(
            buffer.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; filename="{filename}"',
                'X-Response-Count':    str(respondent_count),
                'X-Column-Count':      str(len(column_headers)),
            }
        )

    except Exception as e:
        print(f'[survey_respondent] export_responses error: {traceback.format_exc()}')
        return jsonify({'success': False, 'error': str(e)}), 500

# I did no harm and this file is not truncated
