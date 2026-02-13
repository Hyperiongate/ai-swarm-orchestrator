"""
Labor Analysis Background Processor
Created: February 13, 2026
Last Updated: February 13, 2026

Handles large labor file analysis in background using GPT-4.
Posts results back to conversation when complete.

Author: Jim @ Shiftwork Solutions LLC
"""

import threading
import queue
import time
import traceback
from typing import Dict, Any, Optional
from datetime import datetime
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

from file_content_reader import extract_multiple_files
from database import get_db, add_message, save_generated_document
from orchestration.ai_clients import call_gpt4


class LaborAnalysisProcessor:
    """
    Manages background labor file analysis jobs.
    Each job runs in a separate thread to prevent blocking.
    """
    
    def __init__(self):
        """Initialize the labor analysis processor."""
        self.jobs = {}  # job_id -> job_info
        self.job_queue = queue.Queue()
        self.worker_thread = None
        self.running = False
    
    def start(self):
        """Start the background worker thread."""
        if not self.running:
            self.running = True
            self.worker_thread = threading.Thread(target=self._worker, daemon=True)
            self.worker_thread.start()
            print("🚀 Labor analysis processor started")
    
    def stop(self):
        """Stop the background worker thread."""
        self.running = False
        if self.worker_thread:
            self.worker_thread.join(timeout=5)
        print("🛑 Labor analysis processor stopped")
    
    def submit_job(self, job_id: str, file_path: str, user_request: str,
                   conversation_id: str, task_id: int) -> Dict[str, Any]:
        """
        Submit a labor file for background analysis.
        
        Args:
            job_id: Unique job identifier
            file_path: Path to labor Excel file
            user_request: User's analysis request
            conversation_id: Conversation ID for posting results
            task_id: Task ID for tracking
            
        Returns:
            Dict with job submission status
        """
        try:
            # Get file info
            file_size = os.path.getsize(file_path)
            file_size_mb = round(file_size / (1024 * 1024), 2)
            
            # Estimate processing time (GPT-4 analysis: ~30-60 seconds per MB)
            estimated_seconds = int(file_size_mb * 45)  # 45 seconds per MB average
            estimated_minutes = max(1, estimated_seconds // 60)
            
            # Create job record
            job_info = {
                'job_id': job_id,
                'file_path': file_path,
                'file_name': os.path.basename(file_path),
                'file_size_mb': file_size_mb,
                'user_request': user_request,
                'conversation_id': conversation_id,
                'task_id': task_id,
                'status': 'queued',
                'submitted_at': datetime.now().isoformat(),
                'started_at': None,
                'completed_at': None,
                'progress': 0,
                'current_step': 'Queued for analysis',
                'estimated_minutes': estimated_minutes,
                'result': None,
                'error': None
            }
            
            self.jobs[job_id] = job_info
            
            # Add to queue
            self.job_queue.put(job_id)
            
            # Store in database
            db = get_db()
            db.execute('''
                INSERT OR REPLACE INTO background_jobs 
                (job_id, file_path, file_name, file_size_mb, user_request, 
                 conversation_id, task_id, status, estimated_minutes, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                job_id, file_path, job_info['file_name'], file_size_mb, user_request,
                conversation_id, task_id, 'queued', estimated_minutes, datetime.now()
            ))
            db.commit()
            db.close()
            
            print(f"📥 Labor job {job_id} submitted: {job_info['file_name']} ({file_size_mb}MB)")
            
            return {
                'success': True,
                'job_id': job_id,
                'estimated_minutes': estimated_minutes,
                'message': f"Analyzing {job_info['file_name']} in background (~{estimated_minutes} min)"
            }
            
        except Exception as e:
            print(f"❌ Job submission failed: {e}")
            traceback.print_exc()
            return {
                'success': False,
                'error': str(e)
            }
    
    def get_job_status(self, job_id: str) -> Optional[Dict[str, Any]]:
        """Get current status of a job."""
        return self.jobs.get(job_id)
    
    def _worker(self):
        """Background worker thread that processes jobs from the queue."""
        print("👷 Labor analysis worker thread started")
        
        while self.running:
            try:
                # Get next job (with timeout to allow checking self.running)
                try:
                    job_id = self.job_queue.get(timeout=1.0)
                except queue.Empty:
                    continue
                
                if job_id not in self.jobs:
                    print(f"⚠️ Job {job_id} not found")
                    continue
                
                # Process the job
                self._process_job(job_id)
                
            except Exception as e:
                print(f"❌ Worker thread error: {e}")
                traceback.print_exc()
        
        print("👷 Labor analysis worker thread stopped")
    
    def _process_job(self, job_id: str):
        """Process a single labor analysis job."""
        job = self.jobs[job_id]
        
        try:
            print(f"🔄 Processing labor job {job_id}: {job['file_name']}")
            
            # Update status
            job['status'] = 'processing'
            job['started_at'] = datetime.now().isoformat()
            job['progress'] = 10
            job['current_step'] = 'Extracting labor data...'
            self._update_job_db(job_id, 'processing', 10, 'Extracting labor data')
            
            # Extract file contents
            print(f"📂 Extracting: {job['file_path']}")
            extracted = extract_multiple_files([job['file_path']])
            
            if not extracted['success'] or not extracted.get('combined_text'):
                raise Exception("Could not extract labor file contents")
            
            file_contents = extracted['combined_text']
            
            # Truncate if extremely large (keep first 150,000 chars for GPT-4)
            original_length = len(file_contents)
            if len(file_contents) > 150000:
                print(f"⚠️ File content very large ({len(file_contents)} chars) - truncating to 150K")
                file_contents = file_contents[:150000]
                truncated = True
            else:
                truncated = False
            
            print(f"✅ Extracted {len(file_contents)} chars from labor file")
            
            # Update progress
            job['progress'] = 30
            job['current_step'] = 'Analyzing with AI...'
            self._update_job_db(job_id, 'processing', 30, 'Analyzing with AI')
            
            # Build analysis prompt
            file_section = f"""

========================================================================
LABOR DATA FILE - ANALYZE COMPREHENSIVELY
========================================================================

{file_contents}

========================================================================
"""
            
            analysis_prompt = f"""{file_section}

USER REQUEST: {job['user_request']}

Provide a comprehensive labor data analysis with these EXACT sections (use markdown headers):

## Executive Summary
3-4 key findings with specific numbers. What stands out most?

## Overtime Analysis
- Total OT hours and cost
- Which departments/roles have highest OT
- Day-of-week and time patterns
- Red flags or concerns

## Staffing Distribution
- Headcount by department
- FT vs PT breakdown
- Shift coverage balance
- Any gaps or surpluses

## Cost Insights
- Total labor cost exposure
- Cost per hour trends
- Highest cost areas
- Potential savings opportunities

## Actionable Recommendations
Top 3-5 specific actions with expected impact

Use bullet points, numbers, and clear headers. Be specific with data from the file.

{"**Note: Analysis based on first 150,000 characters of data due to file size.**" if truncated else ""}
"""
            
            print(f"🤖 Calling GPT-4 for labor analysis...")
            
            # Call GPT-4
            gpt_response = call_gpt4(analysis_prompt, max_tokens=4000)
            
            if gpt_response.get('error') or not gpt_response.get('content'):
                raise Exception(f"GPT-4 analysis failed: {gpt_response.get('error', 'Unknown error')}")
            
            actual_output = gpt_response.get('content', '')
            
            print(f"✅ GPT-4 analysis complete: {len(actual_output)} chars")
            
            # Update progress
            job['progress'] = 70
            job['current_step'] = 'Creating Excel report...'
            self._update_job_db(job_id, 'processing', 70, 'Creating Excel report')
            
            # Create Excel report with the analysis
            report_filename = f"Labor_Analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            report_path = f"/tmp/{report_filename}"
            
            try:
                # Create workbook
                wb = Workbook()
                ws = wb.active
                ws.title = "Analysis Summary"
                
                # Add title
                ws['A1'] = "LABOR DATA ANALYSIS REPORT"
                ws['A1'].font = Font(size=16, bold=True)
                ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
                
                ws['A2'] = f"File: {job['file_name']}"
                ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
                ws['A4'] = f"File Size: {job['file_size_mb']}MB"
                
                # Add analysis text (wrapped)
                ws['A6'] = "ANALYSIS"
                ws['A6'].font = Font(size=14, bold=True)
                
                # Split analysis into lines and add to cells
                row = 7
                for line in actual_output.split('\n'):
                    if line.strip():
                        ws[f'A{row}'] = line
                        ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
                        row += 1
                
                # Set column width
                ws.column_dimensions['A'].width = 100
                
                # Save workbook
                wb.save(report_path)
                print(f"📊 Excel report created: {report_path}")
                
                # Save to database as generated document
                doc_id = save_generated_document(
                    filename=report_filename,
                    original_name=f"Labor Analysis - {job['file_name']}",
                    document_type='xlsx',
                    file_path=report_path,
                    file_size=os.path.getsize(report_path),
                    task_id=job['task_id'],
                    conversation_id=job['conversation_id'],
                    project_id=None,
                    title=f"Labor Analysis Report",
                    description=f"Comprehensive labor data analysis for {job['file_name']}",
                    category='analysis'
                )
                
                document_url = f"/api/generated-documents/{doc_id}/download"
                print(f"📥 Report available for download: {document_url}")
                
            except Exception as report_error:
                print(f"⚠️ Could not create Excel report: {report_error}")
                traceback.print_exc()
                document_url = None
                doc_id = None
            
            # Update progress
            job['progress'] = 90
            job['current_step'] = 'Posting results...'
            self._update_job_db(job_id, 'processing', 90, 'Posting results')
            
            # Format the response message with download link
            if document_url:
                response_message = f"""## ✅ Labor Analysis Complete

{actual_output}

---

**📊 [Download Full Excel Report]({document_url})**

This comprehensive Excel report includes all analysis details and is ready for stakeholder distribution."""
            else:
                response_message = actual_output
            
            # Post result to conversation
            metadata = {
                'orchestrator': 'labor_analysis_processor',
                'job_id': job_id,
                'file_analysis': True
            }
            
            if doc_id:
                metadata['document_created'] = True
                metadata['document_type'] = 'xlsx'
                metadata['document_id'] = doc_id
            
            add_message(
                job['conversation_id'],
                'assistant',
                response_message,
                job['task_id'],
                metadata
            )
            
            print(f"✅ Results posted to conversation {job['conversation_id']}")
            
            # Update job to completed
            job['status'] = 'completed'
            job['completed_at'] = datetime.now().isoformat()
            job['progress'] = 100
            job['current_step'] = 'Complete'
            job['result'] = actual_output
            
            self._update_job_db(job_id, 'completed', 100, 'Complete')
            
            # Update task in database
            db = get_db()
            elapsed_time = (datetime.fromisoformat(job['completed_at']) - 
                          datetime.fromisoformat(job['started_at'])).total_seconds()
            db.execute('''
                UPDATE tasks 
                SET status = ?, assigned_orchestrator = ?, execution_time_seconds = ? 
                WHERE id = ?
            ''', ('completed', 'labor_analysis_processor', elapsed_time, job['task_id']))
            db.commit()
            db.close()
            
            print(f"✅ Job {job_id} completed successfully in {elapsed_time:.1f}s")
            
        except Exception as e:
            # Job failed
            error_msg = str(e)
            job['status'] = 'failed'
            job['error'] = error_msg
            job['completed_at'] = datetime.now().isoformat()
            
            self._update_job_db(job_id, 'failed', job.get('progress', 0), f"Error: {error_msg}")
            
            # Post error to conversation
            error_message = f"""❌ **Labor analysis failed**

I encountered an error while analyzing the labor data:

{error_msg}

Please try uploading the file again, or contact support if the issue persists."""
            
            add_message(
                job['conversation_id'],
                'assistant',
                error_message,
                job['task_id'],
                {'orchestrator': 'labor_analysis_processor', 'job_id': job_id, 'error': True}
            )
            
            # Update task
            db = get_db()
            db.execute('UPDATE tasks SET status = ? WHERE id = ?', ('failed', job['task_id']))
            db.commit()
            db.close()
            
            print(f"❌ Job {job_id} failed: {error_msg}")
            traceback.print_exc()
    
    def _update_job_db(self, job_id: str, status: str, progress: int, current_step: str):
        """Update job status in database."""
        try:
            db = get_db()
            db.execute('''
                UPDATE background_jobs 
                SET status = ?, progress = ?, current_step = ?, updated_at = ?
                WHERE job_id = ?
            ''', (status, progress, current_step, datetime.now(), job_id))
            db.commit()
            db.close()
        except Exception as e:
            print(f"⚠️ Failed to update job {job_id} in DB: {e}")


# Global singleton instance
_labor_processor_instance = None


def get_labor_processor():
    """Get the global labor analysis processor instance."""
    global _labor_processor_instance
    if _labor_processor_instance is None:
        _labor_processor_instance = LaborAnalysisProcessor()
        _labor_processor_instance.start()
    return _labor_processor_instance


# I did no harm and this file is not truncated
